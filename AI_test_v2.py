import sys
import time
import datetime
import telegram

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic

import pandas as pd
import pandas_datareader as pdr
from pandas import Series, DataFrame

import sqlite3
import talib as ta
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import winsound

form_class = uic.loadUiType("AI_test_v2_form.ui")[0]

class AI_test_v2(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.timer = QTimer(self)
        self.timer.start(1000)
        self.timer.timeout.connect(self.timeout)
        self.recommanded_done = False

        self.pushButton_1.clicked.connect(self.execute_query)  # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행
        self.pushButton_2.clicked.connect(self.download_basic_data)  # 일봉 다운로드
        self.pushButton_3.clicked.connect(self.add_tech_data)       # 기술적 지표 추가
        self.pushButton_4.clicked.connect(self.download_jisu)  # 지수 다운로드
        self.pushButton_5.clicked.connect(self.condition_filter)  # 조건 필터 추가
        self.pushButton_10.clicked.connect(self.test___)
        self.sound_play()

    def timeout(self):
        print("test")

    def codeNname_load(self, category='Default'):      # 인자를 받아서 코드 불러오기
        if category == 'Default':
            kospi_kosdaq = self.comboBox_1.currentText()
        else:
            kospi_kosdaq = category

        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")
        sheet = file.get_sheet_by_name(kospi_kosdaq)  #인자 받는 부분 나중에 수정할 것, 앞으로 지원하지 않는 기능임
        code_list = []
        name_list = []
        index_list = []

        for r in sheet.rows:
            index_list.append(r[0].value)
            code_list.append(r[1].value)
            name_list.append(r[2].value)

        codeNname = {"index": index_list, "code": code_list, "name": name_list}

        df = DataFrame(codeNname)
        return df

    def sound_play(self):
        # 도레미파솔라시 Hz
        so1 = {'do': 261, 're': 293, 'mi': 329, 'pa': 349, 'sol': 391, 'ra': 440, 'si': 493}
        # mel = ['do', 're', 'mi', 'pa', 'sol', 'ra', 'si']
        mel = ['do', 'mi', 'sol']
        dur = [2, 2, 4]             # 연주 시간

        music = zip(mel, dur)

        for melody, duration in music:
            winsound.Beep(so1[melody], 100 * duration)

    def simul_data_to_excel(self, data, save_name="data_result"):  # contine : 이어쓰기 여부 1이면 이어쓰기임   @@@@@@@@@@@@ 수정필요함

        path = "c:/users/백/save_excel.xlsx"
        wb = load_workbook(path, data_only=True)
        worksheet_1 = wb.active

        # Dataframe을 엑셀로 뿌린다.
        maxrow = worksheet_1.max_row

        if maxrow == 1:
            header = True
        else:
            header = False

        for i, row in enumerate(dataframe_to_rows(data, index=True, header=header)):  #한줄씩 엑셀로
            if len(row) > 1:
                worksheet_1.append(row)
        save_name = str(save_name)
        save_path = "c:/users/백/" + save_name + ".xlsx"
        wb.save(save_path)
        maxrow = worksheet_1.max_row
        print("최대행은 %s행 입니다." % maxrow)
        print("saved to file name %s.xlsx!!" % save_name)
        self.sound_play()


    def execute_query(self, save_file_name="Default"):
        kospi_kosdaq = self.comboBox_1.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_tec.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        # con = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        query_ = self.lineEdit.text()  # 입력된 값을 저장

        df2 = DataFrame()
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print("%s 진입 %s / %s" % (name, i + 1, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query_, con, index_col='index')
            df2 = df2.append(df)

        con.close()
        print("쿼리 실행 완료")
        self.simul_data_to_excel(df2, save_file_name)

    def download_basic_data(self):

        start_date = '20110102'
        kospi_kosdaq = self.comboBox_1.currentText()
        code_name = self.codeNname_load()  # 종목 코드로드
        codes = code_name['code']  # 코드네임을 리스트로 저장
        names = code_name['name']  # 코드네임을 리스트로 저장
        # codes = ['A058860']  # 코드네임을 리스트로 저장
        # names = ['KTis']  # 코드네임을 리스트로 저장

        file_name = kospi_kosdaq + "data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        for i, code in enumerate(codes):  # 뺑뺑이 돌리기
            date_list = []
            name = names[i]
            print(name + " 진입")
            code = code[1:]
            # data = pdr.get_data_yahoo(code + '.KS', start='20190101')
            data = pdr.naver.NaverDailyReader(code, start=start_date).read()

            date_index = data.index
            for date in date_index:
                date = str(date)
                date = date[:4] + date[5:7] + date[8:10]
                date = int(date)
                date_list.append(date)

            df = pd.DataFrame(data)
            df['index'] = date_list
            df['Open'] = pd.to_numeric(df['Open'])
            df['Close'] = pd.to_numeric(df['Close'])
            df['High'] = pd.to_numeric(df['High'])
            df['Low'] = pd.to_numeric(df['Low'])
            df['Change'] = df['Close'].diff()
            df['change_ratio'] = df['Close'].pct_change()  # 변화량을 퍼센테이지로 구한다.
            df['change_ratio'] = df['change_ratio'].round(5)
            df['Volume'] = pd.to_numeric(df['Volume'])
            df.to_sql(name, con, if_exists="replace")

            print(i + 1, "/", len(names), name, " 완료")
        con.close()

        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()

    def tech__(self, name, kospi_kosdaq='etf_real'):    #한종목의 기술적 지표 계산
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_1.currentText()
        else:
            pass

        file_name = kospi_kosdaq + "data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        try:
            df = pd.read_sql("SELECT * FROM "+ "'"+ name +"' ", con, index_col='index') #df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        except:
            return

        if len(df['Open']) > 0:  # 값이 없는것을 제외시킨다
            df['name'] = name
            op = df['Open']
            cl = df['Close']
            cl_list = list(cl)
            shift_close = cl.shift()    # 전일 종가
            hi = df['High']
            lo = df['Low']
            vo = df['Volume']

            df['Open_R'] = (op - shift_close) / shift_close
            df['High_R'] = (hi - shift_close) / shift_close
            df['Low_R'] = (lo - shift_close) / shift_close
            df['Vol_R'] = vo.pct_change()
            df['Vol_R'] = df['Vol_R'].round(5)      # 전일 대비 거래량

            dfma3 = ta.SMA(cl, 3)
            dfma5 = ta.SMA(cl, 5)
            dfma10 = ta.SMA(cl, 10)
            dfma20 = ta.SMA(cl, 20)
            dfma60 = ta.SMA(cl, 60)

            tr = ta.TRANGE(hi, lo, cl)
            df['atr20(N)'] = ta.SMA(tr, 20)

            df['vma20_R'] = vo / ta.SMA(vo, 20)
            df['dp_ma3'] = cl / dfma3
            df['dp_ma5'] = cl / dfma5
            df['dp_ma10'] = cl / dfma10
            df['dp_ma20'] = cl / dfma20
            df['dp_ma60'] = cl / dfma60
            # df['dp_ma120'] = cl / dfma120
            # # df['dp_ma480'] = cl / dfma480

            bbands20 = pd.Series(ta.BBANDS(cl, timeperiod=20, nbdevup=2, nbdevdn=2))  # 밴드는 또 Series를 만들었다가
            # df['bbands20_upR'] = cl/bbands20[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # df['bbands30_mov'] = bbands30[1]
            # df['bbands20_downR'] = cl/bbands20[2]
            df['bbands_width'] = (bbands20[0] - bbands20[2]) / cl

            df['bong'] = (df['Close'] - df['Open']) / df['Open']  # 봉크기
            df['bong'] = df['bong'].round(5)
            df['N20_R'] = df['atr20(N)'] / cl
            df['N20_R'] = df['N20_R'].round(5)

            # df['dm1_open'] = op.shift(1)
            # df['dm1_high'] = hi.shift(1)
            # df['dm1_low'] = lo.shift(1)
            # df['dm1_close'] = cl.shift(1)
            # df['dm1_open_R'] = (df['dm1_open'] - cl) / cl
            # df['dm1_high_R'] = (df['dm1_high'] - cl) / cl
            # df['dm1_low_R'] = (df['dm1_low'] - cl) / cl
            # df['dm1_close_R'] = df['change_ratio'].shift(1)

            df['d1_open'] = op.shift(-1)
            df['d1_high'] = hi.shift(-1)
            df['d1_low'] = lo.shift(-1)
            df['d1_close'] = cl.shift(-1)
            df['d1_open_R'] = (df['d1_open'] - cl) / cl
            df['d1_high_R'] = (df['d1_high'] - cl) / cl
            df['d1_low_R'] = (df['d1_low'] - cl) / cl
            df['d1_close_R'] = df['change_ratio'].shift(-1)

            df['d2_open'] = op.shift(-2)
            df['d2_high'] = hi.shift(-2)
            df['d2_low'] = lo.shift(-2)
            df['d2_close'] = cl.shift(-2)
            df['d2_open_R'] = df['d1_open_R'].shift(-1)
            df['d2_high_R'] = df['d1_high_R'].shift(-1)
            df['d2_low_R'] = df['d1_low_R'].shift(-1)
            df['d2_close_R'] = df['change_ratio'].shift(-2)

            max_10_list = df['Close'].rolling(window=10).max()
            min_10_list = df['Close'].rolling(window=10).min()
            max_20_list = df['Close'].rolling(window=20).max()
            min_20_list = df['Close'].rolling(window=20).min()
            max_55_list = df['Close'].rolling(window=55).max()
            min_55_list = df['Close'].rolling(window=55).min()
            max_60_list = df['Close'].rolling(window=60).max()
            min_60_list = df['Close'].rolling(window=60).min()

            max_10_R_list = (df['Close'] - max_10_list) / max_10_list   # 10일 최고가대비 얼마나 하락했는가
            min_10_R_list = (df['Close'] - min_10_list) / min_10_list  # 10일 최저가대비 얼마나 상승했는가

            max_20_R_list = (df['Close'] - max_20_list) / max_20_list
            min_20_R_list = (df['Close'] - min_20_list) / min_20_list

            max_55_R_list = (df['Close'] - max_55_list) / max_55_list
            min_55_R_list = (df['Close'] - min_55_list) / min_55_list

            max_60_R_list = (df['Close'] - max_60_list) / max_60_list
            min_60_R_list = (df['Close'] - min_60_list) / min_60_list


            df['max_10_R'] = max_10_R_list
            df['max_10_R'] = df['max_10_R'].round(5)
            df['min_10_R'] = min_10_R_list
            df['min_10_R'] = df['min_10_R'].round(5)
            df['max_20_R'] = max_20_R_list
            df['max_20_R'] = df['max_20_R'].round(5)
            df['min_20_R'] = min_20_R_list
            df['min_20_R'] = df['min_20_R'].round(5)
            df['max_55_R'] = max_55_R_list
            df['max_55_R'] = df['max_55_R'].round(5)
            df['min_55_R'] = min_55_R_list
            df['min_55_R'] = df['min_55_R'].round(5)
            df['max_60_R'] = max_60_R_list
            df['max_60_R'] = df['max_60_R'].round(5)
            df['min_60_R'] = min_60_R_list
            df['min_60_R'] = df['min_60_R'].round(5)

            # 열순서 정렬
            df = df[['name', 'Open', 'High', 'Low', 'Close', 'Volume', 'Change', 'change_ratio', 'Open_R',
                     'High_R', 'Low_R', 'Vol_R', 'atr20(N)', 'vma20_R', 'dp_ma3', 'dp_ma5', 'dp_ma10', 'dp_ma20',
                     'dp_ma60', 'bbands_width', 'bong', 'N20_R', 'd1_open', 'd1_high', 'd1_low', 'd1_close',
                     'd1_open_R', 'd1_high_R', 'd1_low_R', 'd1_close_R',
                     'd2_open', 'd2_high', 'd2_low', 'd2_close', 'd2_open_R', 'd2_high_R', 'd2_low_R', 'd2_close_R',
                     'max_10_R', 'min_10_R', 'max_20_R', 'min_20_R', 'max_55_R', 'min_55_R', 'max_60_R', 'min_60_R']]

        return df

    def add_tech_data(self, kospi_kosdaq='etf_real'):         # 전체 종목 대상 자동 기술적지표 구하기
        print(kospi_kosdaq)
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_1.currentText()
        else:
            print(kospi_kosdaq)

        file_name = kospi_kosdaq + "_data_add_tec.db"
        con2 = sqlite3.connect("c:/users/백/" + file_name)  # 키움증권 다운로드 종목 데이터 베이스

        code_name = self.codeNname_load(kospi_kosdaq)   # 종목 코드로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.tech__(name, kospi_kosdaq)
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()

    def download_jisu(self):
        tickers = ['^KS11', '^KQ11']  # 코스피, 코스닥
        names = ['kospi', 'kosdaq']
        file_name = "jisu_data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        for i, code in enumerate(tickers):  # 뺑뺑이 돌리기
            date_list = []
            name = names[i]
            print(name + " 진입")

            data = pdr.get_data_yahoo(code, start='20190101')

            print(i + 1, "/", len(names), name, " 완료")
            date_index = data.index
            for date in date_index:
                date = str(date)
                date = date[:4] + date[5:7] + date[8:10]
                date = int(date)
                date_list.append(date)
            df = pd.DataFrame(data)
            df['index'] = date_list
            df['Open'] = pd.to_numeric(df['Open'])
            df['Close'] = pd.to_numeric(df['Close'])
            df['High'] = pd.to_numeric(df['High'])
            df['Low'] = pd.to_numeric(df['Low'])
            df['Change'] = df['Close'].diff()
            df['Volume'] = pd.to_numeric(df['Volume'])
            df.to_sql(name, con, if_exists="replace")

        con.close()

        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()

    def condition_filter(self, category="Default"):  # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 터틀트레이딩

        #####
        self.sound_play()
        # max_hold_day = 1        # 보유기간
        stopgain = 1.5  # 익절값
        constant_stoploss = 0.1  # 절대 손절값, 2N과 비교하여 더 작은값을 손절값으로
        condition_low = -0.03  # 매수조건 전일 하락값
        condition_open = -0.01  # 매수조건 당일 시초가

        if category is False:  # 나중에 왜 이렇게 되는지 확인할 것
            kospi_kosdaq = self.comboBox_1.currentText()
        else:
            kospi_kosdaq = category

        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % kospi_kosdaq)

        # name_list = code_name["name"]  # 코드네임을 리스트로 저장
        name_list = ["셀트리온", "현대차", "삼성전자"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["KODEX 200", "KODEX 200TR"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['Open'])
            high_list = list(df['High'])
            low_list = list(df['Low'])  # 저가
            close_list = list(df['Close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])
            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []  # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            position_list = []
            hold_day_list = []  # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            buy_type_list = []

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                low = low_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                position = 0
                stoploss_price = 0
                buy_type = 0

                if j == 0:
                    position_list.append(position)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                elif max_20_R == 0 and min_10_R == 0:
                    position_list.append(position)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                else:

                    position = position_list[j - 1]
                    if position == 1:

                        stopgain_price = buy_price_list[j - 1] * stopgain  # 익절값 설정
                        stoploss_price = stoploss_price_list[j - 1]
                        hold_day = hold_day_list[j - 1] + 1
                        buy_type = buy_type_list[j - 1]
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j - 1]
                    d1_open = open_list[j - 1]
                    d1_close = close_list[j - 1]
                    d1_high = high_list[j - 1]
                    d1_low = low_list[j - 1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j - 1]

                    # try:
                    #     d1_u_tail_ratio = (d1_open - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 윗꼬리 가격상 비율
                    #     d1_l_tail_ratio = (d1_close - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 아랫꼬리 가격상 비율
                    # except ZeroDivisionError:
                    #     d1_u_tail_ratio = 0
                    #     d1_l_tail_ratio = 0

                    # if (open_list[j] < d1_low and close > open_list[j] and close > (low + high) / 2 and 0.1 > change > 0.04
                    #         and position == 0):

                    if max_20_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 0 and position == 0:
                        # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
                        # if max_20_R == 0 and N != None and N_ratio > 0.01 and winnloss == 0 and position == 0:
                        #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거

                        position = 1
                        position_list.append(position)
                        hold_day_list.append(hold_day)

                        condition = "buy_s1"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)  # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        # stoploss_price = close - (2 * N)
                        # stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S1"
                        buy_type_list.append(buy_type)


                    elif max_55_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 1 and position == 0:
                        # S2 매수조건 진입, 60이평 위 조건 추가
                        # elif max_55_R == 0 and N != None and N_ratio > 0.01 and winnloss == 1 and position == 0:
                        #     # S2 매수조건 진입, 60이평 위 조건 제거

                        position = 1
                        position_list.append(position)
                        hold_day_list.append(hold_day)

                        condition = "buy_s2"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)  # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        # stoploss_price = close - (2 * N)
                        # stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S2"
                        buy_type_list.append(buy_type)

                    # elif position == 1 and low < stoploss_price or open_list[j] < stoploss_price:      # 손절(즉시)
                    elif position == 1 and close < stoploss_price:  # 손절(종가기준)

                        position = 0
                        position_list.append(position)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j - 1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        # if open_list[j] < stoploss_price:       # 즉시 손절
                        #     sell_price = open_list[j]
                        # else:
                        #     sell_price = stoploss_price
                        # sell_price_list.append(sell_price)

                        sell_price = close  # 종가 기준 손절
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_10_R == 0 and buy_type == "S1" and position == 1:  # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        position = 0
                        position_list.append(position)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j - 1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = close  # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1  # 이득
                        else:
                            winnloss = 0  # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_20_R == 0 and buy_type == "S2" and position == 1:  # S2 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        position = 0
                        position_list.append(position)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j - 1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        sell_price = close  # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1  # 이득
                        else:
                            winnloss = 0  # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif position == 1:  # 홀드

                        position = 1
                        position_list.append(position)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j - 1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    else:  # 그외는 관망

                        position = 0
                        position_list.append(position)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")  # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        buy_type = 0
                        buy_type_list.append(buy_type)

            df_result['date'] = df['Date']
            df_result['name'] = df['name']
            df_result['open'] = df['Open']

            df_result['high'] = df['High']
            df_result['low'] = df['Low']
            df_result['close'] = df['Close']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = df['dp_ma60']
            # df_result['open_ratio'] = open_ratio_list
            df_result['volume'] = df['Volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            df_result['max_20R'] = df['max_20_R']
            df_result['min_10R'] = df['min_10_R']
            df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            # df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']

            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list
            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['position'] = position_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['buy_type'] = buy_type_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i + 1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()  # 터틀 트레이딩 전략 끝

    def test___(self):
        print("test")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    ai_test_v2 = AI_test_v2()
    ai_test_v2.show()
    app.exec_()