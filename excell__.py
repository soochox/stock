from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

df = pd.read_excel("financials.xlsx")


wb = Workbook()
ws = wb.create_sheet()

for row in dataframe_to_rows(df, index=True, header=True):
    if len(row)>1:
        ws.append(row)
        print(row)

ws1 = wb.create_sheet("Mysheet")
ws.title = "New title"
df2 = df["구분"]
df2 = pd.DataFrame(df2)
print(df2)
for row in dataframe_to_rows(df2, index=True, header=True):
    if len(row)>1:
        ws1.append(row)
        print(row)

wb.save("c:/users/백/test.xlsx")

wb = Workbook()
ws = wb.active  # 현재 활성화 되어있는 시트를 가져옴
ws1 = wb.create_sheet("Mysheet")  # 새로시트를 만드는데 위치는 맨끝(디폴트 값임)
ws2 = wb.create_sheet("Mysheet1", 0)  # 새로시트를 만드는데 위치는 첫번째
ws3 = wb.create_sheet("Mysheet2", -1)  # 새로시트를 만드는데 위치는 끝에서 두번째
ws.title = "New Title"  # 이름바꾸기
ws4 = wb["New Title"]  # 이름이 정해지면 그걸로 변수지정가능하다
print(wb.sheetnames)  # 어떤시트가 있는지 확인

for sheet in wb:  # 워크북을 루프 태울수있다. 워크북은 워크시트가 모두 합쳐진것이다.
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)  # 워크시트를 복사할 수 있다. (셀값, 스타일, 코멘트만 복사된다. 이미지나 표는 안됨)
# 셀에 접근
c = ws['A4']
ws['A4'] = 4  # 값을 바로 입력
d = ws.cell(row=4, column=2, value=10)  # 셀 메서드로 값을 입력하고 변수할당함
e = ws.cell(4, 3, 100)  # 참고 : 워크시트가 메모리에 만들어졌을때 셀은 없다. 셀에 처음 억세스했을 때 그때 셀이 만들어짐

# 주의 위의 이런 특징 때문에 셀에 값을 할당안하고 단지 스크롤링하는 것만으로 메모리를 차지할 수 있다.
# 예)
# for x in range(1,101):
# for y in range(1,101):
# ws.cell(row=x, column=y)
# 이렇게 하면 아무값이 없어도 메모리는 100x100셀만큼 차지하게 된다.

# 다수의 셀에 접근
cell_range = ws['A1:C2']  # 원래 엑셀과 같음
col_C = ws['C']  # C 한개열 전체 선택
col_range = ws['C:D']  # 2개열 선택
row10 = ws[10]  # 10행 전체 선택
row_range = ws[5:10]  # 5행~10행 선택
# iter_rows 매서드
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# iter_columns 매서드   #iter_row와 다른 점은 column 을 리턴한다.
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
        # 참고 : iter_cols 는 성능상의 이유로 read-only모드에서는 동작이 안된다.

# 메모리에 읽혀진 모든 셀에 억세스할때 worksheet.rows
ws['C9'] = "hello"
print(ws.rows)

# Values only 값만 옮기고 싶을때 worksheet.values 를 사용
for row in ws.values:
    for value in row:
        print(value)

# Worksheet.iter_rows() 의 values_only 옵션으로 같은 기능을 사용할 수 있다.
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

f = ws['C5']
f.value = 10  # 변수에 값할당은 이렇게 해야함

wb.save("c:/users/백/test_excel.xlsx")  # 파일의 저장, 이명령은 경고없이 바로 파일을 덮어쓴다.

# 파일 로딩 load_workbook
wb0 = load_workbook("c:/users/백/test_excel.xlsx")
print(wb0.sheetnames)