import sys
from pywinauto import application
import win32com.client

from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import *
from PyQt5 import uic
import os

import pandas as pd
from pandas import Series, DataFrame
import pandas.io.sql as pd_sql
import sqlite3
import talib as ta
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from talib.abstract import *
import FinanceDataReader as fdr
import numpy as np

form_class = uic.loadUiType("Serch_UI.ui")[0]

class baek_data_analysis2(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.execute_query)
        self.pushButton_2.clicked.connect(self.add_tech)
        self.pushButton_3.clicked.connect(self.add_tech_manual)
        self.pushButton_4.clicked.connect(self.add_D_N_ohlc)
        self.pushButton_5.clicked.connect(self.merge___)
        self.pushButton_6.clicked.connect(self.drop_data)
        #self.pushButton_7.clicked.connect(self.add_one_col_tech)
        self.pushButton_8.clicked.connect(self.test___)
        self.pushButton_9.clicked.connect(self.add_past_Day_ohlc)
        self.pushButton_10.clicked.connect(self.add_day_max_min)
        self.pushButton_11.clicked.connect(self.add_condition_judge)
        self.pushButton_12.clicked.connect(self.add_month_change)
        self.pushButton_13.clicked.connect(self.add_month_index_all)
        self.pushButton_14.clicked.connect(self.add_min_tech_button)
        self.pushButton_15.clicked.connect(self.get_minute_info_form_day_info)  # 선택된 일봉정보의 분봉정보로 가져오기

    def simul_data_to_excel(self, data):
        # 전략 시뮬레이션 결과를 데이터를 엑셀로 저장, 받은 데이터는 반드시 Dataframe형식이고 "date"열을 포함해야함

        wb = Workbook()
        ws = wb.create_sheet()
        #sheet2_name = "수익률 분석"

        # Dataframe을 엑셀로 뿌린다.
        for i, row in enumerate(dataframe_to_rows(data, index=True, header=True)):  #한줄씩 엑셀로

            if len(row) > 1:
                ws.append(row)

        wb.save("c:/users/백/save_excel.xlsx")
        print("save to excell finish!!")

    def codeNname_load(self):
        kospi_kosdaq = self.comboBox_2.currentText()

        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")

        sheet = file.get_sheet_by_name(kospi_kosdaq)  #인자 받는 부분 나중에 수정할 것, 앞으로 지원하지 않는 기능임
        code_list = []
        name_list = []
        index_list = []

        for r in sheet.rows:
            index_list.append(r[0].value)
            code_list.append(r[1].value)
            name_list.append(r[2].value)

        codeNname = {"index": index_list, "code": code_list, "name": name_list}

        df = DataFrame(codeNname)
        return df

    def tech__(self, name):    #한종목의 기술적 지표 계산
        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        # con = sqlite3.connect("c:/users/백/%s_data_month_index.db" % self.kospi_kosdaq)
        try:
            df = pd.read_sql("SELECT * FROM "+ "'"+ name +"' ", con, index_col='index') #df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        except:
            return

        df.sort_index(inplace=True, ascending=True)  # 인덱스 기준으로 역으로 정렬

        if len(df['open']) > 0:  #값이 없는것을 제외시킨다

            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            vo = df['volume'] * 0.1 * 10

            dfma3 = ta.SMA(cl, 3)
            dfma5 = ta.SMA(cl, 5)
            dfma10 = ta.SMA(cl, 10)
            dfma20 = ta.SMA(cl, 20)
            # dfma60 = ta.SMA(cl, 60)
            # dfma120 = ta.SMA(cl, 120)
            # # dfma480 = ta.SMA(cl, 480)

            df['vma20_R'] = vo / ta.SMA(vo, 20)
            df['dp_ma3'] = cl / dfma3
            df['dp_ma5'] = cl / dfma5
            df['dp_ma10'] = cl / dfma10
            df['dp_ma20'] = cl / dfma20
            # df['dp_ma60'] = cl / dfma60
            # df['dp_ma120'] = cl / dfma120
            # # df['dp_ma480'] = cl / dfma480

            # bbands30 = pd.Series(ta.BBANDS(cl, timeperiod=30, nbdevup=1.8, nbdevdn=1.8))  # 밴드는 또 Series를 만들었다가
            # df['bbands30_upR'] = cl/bbands30[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # #df['bbands30_mov'] = bbands30[1]
            # df['bbands30_downR'] = cl/bbands30[2]
            # df['bbands_width'] = (bbands30[0] - bbands30[2]) / cl
            # print("완료")
            #
            # macd = pd.Series(ta.MACD(cl, 5, 34, 7))  # 밴드는 또 Series를 만들었다가
            # #df['macd_line5,34,7'] = macd[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # #df['macd_sig5,34,7'] = macd[1]
            # #df['macd_histo5,34,7'] = macd[2]
            #
            # #df['macd_line5,34,7'] = macd[0] / dfma5  # 5일 이평기준 어느정도에 있는가?
            # #df['macd_sig5,34,7'] = macd[1] / dfma5
            # #df['macd_histo5,34,7'] = macd[2] / dfma5  #히스토그램 구하는 공식은 단기이평-장기 이평임
            #
            # #df['rsi5'] = ta.RSI(cl, 5)
            # #df['rsi10'] = ta.RSI(cl, 10)
            # #df['rsi20'] = ta.RSI(cl, 20)
            # #df['rsi60'] = ta.RSI(cl, 60)
            # #df['rsi120'] = ta.RSI(cl, 60)

        return df


    def add_tech(self):   #전체 종목 대상 자동 기술적지표 구하기

        self.kospi_kosdaq = self.comboBox_2.currentText()
        file_name = self.kospi_kosdaq + "_data_add_auto.db"
        con2 = sqlite3.connect("c:/users/백/" + file_name)  # 키움증권 다운로드 종목 데이터 베이스

        code_name = self.codeNname_load()   #종목 코드로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.tech__(name)
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("최종완료, %s 로 저장 완료" %file_name)

    def tech_manual(self, name):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db")
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con,
                         index_col='index')  # df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)

        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다

            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            vo = df['volume'] * 0.1 * 10
            dp_ma20 = df['dp_ma20'] * 0.1 * 10

            # bband__ = df['bbands30_upR']

            bong_list = (cl-op)/op  # 봉크기
            # gap_ratio_list = []  #갭 비율
            # ma20_grd_5_list = [] #20일 이평선의 기울기(5일기준)
            u_tail_R_list = []  #윗꼬리 비율
            l_tail_R_list = []   #아랫꼬리 비율


            vol_d1_R = []  # 전일대비 거래량 비율 구하기
            D1_profit_list = []  # 익일 종가 - 당일종가 비율
            D1_open_list = []   #익일 시가
            D1_high_list = []  #익일 고가
            D1_low_list = []   #익일 저가
            D1_close_list = []  #익일 종가
            #D1_bband_list = [] #익일 볼린저밴드

            for i in range(len(op)):

                if i == 0:  # 첫번째 데이터는 전일 정보가 없으므로 0추가
                    vol_d1_R.append(0)
                    gap_ratio_list.append(0)

                elif vo[i - 1] == 0:  # 0으로 나누는 경우 error방지
                    vol_d1_R.append(0)
                    d0_open = op[i]  # 당일 시가
                    d0_close = cl[i]  # 당일 종가
                    d_1_open = op[i - 1]  # 전일 시가
                    d_1_close = cl[i - 1]  # 전일 종가

                    d_1_max_open_close = max(d_1_open, d_1_close)  # 전일 시가, 종가 중 큰값
                    d_1_min_open_close = min(d_1_open, d_1_close)  # 전일 시가, 종가 중 작은값

                    d0_max_open_close = max(d0_open, d0_close)  # 당일 시가, 종가 중 큰값
                    d0_min_open_close = min(d0_open, d0_close)  # 당일 시가, 종가 중 작은값


                    # 여기 아래는 갭구하는 코드
                    # if d0_min_open_close > d_1_max_open_close:  # 상승갭이 있다 = 금일 시가와 종가 중 작은값이 전일 시가와 종가 중 큰 값보다 크다
                    #     gap = d0_min_open_close - d_1_max_open_close
                    #     gap_ratio = gap / d_1_max_open_close  # 전일기준 갭사이즈 비율
                    #     gap_ratio_list.append(gap_ratio)
                    #
                    # elif d0_max_open_close < d_1_min_open_close:  # 하락갭이 있다 = 금일 시가 종가 중 큰값이 전일 시가 종가 중 작은 값보다 작다
                    #     gap = d0_max_open_close - d_1_min_open_close
                    #     gap_ratio = gap / d_1_min_open_close  # 전일기준 갭사이즈 비율
                    #     gap_ratio_list.append(gap_ratio)
                    #
                    # else:
                    #     gap_ratio_list.append(0)

                else:
                    vol_d1_Ratio = vo[i] / vo[i - 1]  # i=오늘, i-1 = 어제 i가 커질수록 과거에서 미래로 가는것임. 전일대비 거래량
                    vol_d1_R.append(vol_d1_Ratio)

                    d0_open = op[i]  # 당일 시가
                    d0_close = cl[i]  # 당일 종가
                    d_1_open = op[i - 1]  # 전일 시가
                    d_1_close = cl[i - 1]  # 전일 종가

                    d_1_max_open_close = max(d_1_open, d_1_close)  # 전일 시가, 종가 중 큰값
                    d_1_min_open_close = min(d_1_open, d_1_close)  # 전일 시가, 종가 중 작은값

                    d0_max_open_close = max(d0_open, d0_close)  # 당일 시가, 종가 중 큰값
                    d0_min_open_close = min(d0_open, d0_close)  # 당일 시가, 종가 중 작은값

                    if d0_min_open_close > d_1_max_open_close:  # 상승갭이 있다 = 금일 시가와 종가 중 작은값이 전일 시가와 종가 중 큰 값보다 크다
                        gap = d0_min_open_close - d_1_max_open_close
                        gap_ratio = gap / d_1_max_open_close  # 전일기준 갭사이즈 비율
                        gap_ratio_list.append(gap_ratio)

                    elif d0_max_open_close < d_1_min_open_close:  # 하락갭이 있다 = 금일 시가 종가 중 큰값이 전일 시가 종가 중 작은 값보다 작다
                        gap = d0_max_open_close - d_1_min_open_close
                        gap_ratio = gap / d_1_min_open_close  # 전일기준 갭사이즈 비율
                        gap_ratio_list.append(gap_ratio)

                    else:
                        gap_ratio_list.append(0)


            for i in range(len(op)):

                u_tail_R = (hi[i] - max(op[i],cl[i]))/max(op[i],cl[i])
                u_tail_R_list.append(u_tail_R)
                l_tail_R = (min(op[i], cl[i]) - lo[i]) / min(op[i], cl[i])
                l_tail_R_list.append(l_tail_R)

                if i == len(op) - 1:  # 가장 최근 데이터는 익일 정보가 없으므로 0추가

                    D1_profit_list.append(0)
                    D1_open_list.append(0)
                    D1_high_list.append(0)
                    D1_low_list.append(0)
                    D1_close_list.append(0)
                    #D1_bband_list.append(0)

                elif cl[i] == 0:  # 0으로 나누는 경우 error방지
                    D1_profit_list.append(0)

                else:
                    D1_profit = (cl[i + 1] - cl[i]) / cl[i]
                    D1_profit_list.append(D1_profit)
                    D1_open_list.append(op[i+1])
                    D1_high_list.append(hi[i+1])
                    D1_low_list.append(lo[i+1])
                    D1_close_list.append(cl[i+1])
                    #D1_bband_list.append(bband__[i+1])


            df['bong_ratio'] = bong_list
            df['gap_ratio'] = gap_ratio_list
            # df['ma20_grd'] = ma20_grd_5_list
            df['u_tail_R'] = u_tail_R_list
            df['l_tail_R'] = l_tail_R_list

            df['vol_d1_R'] = vol_d1_R
            df['D1_open'] = D1_open_list
            df['D1_high'] = D1_high_list
            df['D1_low'] = D1_low_list
            df['D1_close'] = D1_close_list
            df['D1_profit'] = D1_profit_list  # 종가거래 하루 수익
            #df['D1_bbands30_upR'] = D1_bband_list

        return df

    def add_tech_manual(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.tech_manual(name)
            df['name'] = name   #종목명 삽입
            # df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            #df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
            #con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            #print(df.columns)
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        print("tech_manual 최종완료")

    def D_N_ohlc(self, name):  # 1일~5일 ohlc확인  삭제 검토 할 것

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data.db")   #기술적지표 추가된거에 추가
        #con = sqlite3.connect("c:/users/백/trace_" + self.kospi_kosdaq + ".db")   #원본에 추가(나중에 join으로 합치기)

        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)

        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다

            date = df['index']
            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            ch = df['change_ratio'] * 0.1 * 10

            D1_open_list = []
            D1_close_list = []
            D1_high_list = []
            D1_low_list = []
            D1_change_R_list = []

            D1_v20maR_list = []

            D2_open_list = []
            D2_close_list = []
            D2_high_list = []
            D2_low_list = []
            D2_change_R_list = []

            D3_open_list = []
            D3_close_list = []
            D3_high_list = []
            D3_low_list = []
            D3_change_R_list = []

            v20mar = df['vma20_R']
            for i in range(len(op)):

                if i > len(op) - 1 - 1:  # 가장 최근 데이터는 익일 정보가 없으므로 0추가 (과거 → 현재), 0부터 1을 더뺀다
                    D1_open_list.append(0)
                    D1_close_list.append(0)
                    D1_high_list.append(0)
                    D1_low_list.append(0)
                    D1_change_R_list.append(0)

                    D1_v20maR_list.append(0)

                    D2_open_list.append(0)
                    D2_close_list.append(0)
                    D2_high_list.append(0)
                    D2_low_list.append(0)
                    D2_change_R_list.append(0)

                    D3_open_list.append(0)
                    D3_close_list.append(0)
                    D3_high_list.append(0)
                    D3_low_list.append(0)
                    D3_change_R_list.append(0)

                elif i > len(op) - 2 - 1:

                    D1_open_list.append(op[i+1])
                    D1_close_list.append(cl[i+1])
                    D1_high_list.append(hi[i+1])
                    D1_low_list.append(lo[i+1])
                    D1_change_R_list.append(ch[i+1])
                    D1_v20maR_list.append(v20mar[i+1])

                    D2_open_list.append(0)
                    D2_close_list.append(0)
                    D2_high_list.append(0)
                    D2_low_list.append(0)
                    D2_change_R_list.append(0)

                    D3_open_list.append(0)
                    D3_close_list.append(0)
                    D3_high_list.append(0)
                    D3_low_list.append(0)
                    D3_change_R_list.append(0)

                elif i > len(op) - 3 - 1:

                    D1_open_list.append(op[i + 1])
                    D1_close_list.append(cl[i + 1])
                    D1_high_list.append(hi[i + 1])
                    D1_low_list.append(lo[i + 1])
                    D1_change_R_list.append(ch[i + 1])

                    D1_v20maR_list.append(v20mar[i + 1])

                    D2_open_list.append(op[i + 2])
                    D2_close_list.append(cl[i + 2])
                    D2_high_list.append(hi[i + 2])
                    D2_low_list.append(lo[i + 2])
                    D2_change_R_list.append(ch[i + 2])

                    D3_open_list.append(0)
                    D3_close_list.append(0)
                    D3_high_list.append(0)
                    D3_low_list.append(0)
                    D3_change_R_list.append(0)

                else:

                    D1_open_list.append(op[i + 1])
                    D1_close_list.append(cl[i + 1])
                    D1_high_list.append(hi[i + 1])
                    D1_low_list.append(lo[i + 1])
                    D1_change_R_list.append(ch[i + 1])
                    D1_v20maR_list.append(v20mar[i + 1])

                    D2_open_list.append(op[i + 2])
                    D2_close_list.append(cl[i + 2])
                    D2_high_list.append(hi[i + 2])
                    D2_low_list.append(lo[i + 2])
                    D2_change_R_list.append(ch[i + 2])

                    D3_open_list.append(op[i + 3])
                    D3_close_list.append(cl[i + 3])
                    D3_high_list.append(hi[i + 3])
                    D3_low_list.append(lo[i + 3])
                    D3_change_R_list.append(ch[i + 3])

                
            df['D1_open'] = D1_open_list
            df['D1_close'] = D1_close_list
            df['D1_high'] = D1_high_list
            df['D1_low'] = D1_low_list
            df['D1_changeR'] = D1_change_R_list

            df['D1_v20maR'] = D1_v20maR_list

            df['D2_open'] = D2_open_list
            df['D2_close'] = D2_close_list
            df['D2_high'] = D2_high_list
            df['D2_low'] = D2_low_list
            df['D2_changeR'] = D2_change_R_list

            df['D3_open'] = D3_open_list
            df['D3_close'] = D3_close_list
            df['D3_high'] = D3_high_list
            df['D3_low'] = D3_low_list
            df['D3_changeR'] = D3_change_R_list

        con.close()

        return df


    def add_D_N_ohlc(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()

        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data.db")  # 키움증권 다운로드 종목 데이터 베이스
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
        #for i in [523,524,525,526]:  # 뺑뺑이 돌리기
            #name = name_list[i]
            print(name + " 진입")
            df = self.D_N_ohlc(name)
            print(len(df['open']))
            if len(df['open']) > 0:
                df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
                df.to_sql(name, con2, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("D+N_high_low 최종완료")

    def past_Day_ohlc(self, name):  #전시점 tech 넣기

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data.db")  # 기술적지표 추가된거에 추가
        # con = sqlite3.connect("c:/users/백/trace_" + self.kospi_kosdaq + ".db")   #원본에 추가(나중에 join으로 합치기)
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)
        print(len(df['open']))
        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다
            print("진입")
            df['date'] = df['index']
            df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            df.drop("index", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.

            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            ch = df['change_ratio'] * 0.1 * 10
            v20 = df['vma20_R']

            #dp_ma3 = df['dp_ma3']
            dp_ma5 = df['dp_ma5']
            dp_ma10 = df['dp_ma10']
            dp_ma20 = df['dp_ma20']

            # dp_ma60 = df['dp_ma60']
            # dp_ma120 = df['dp_ma120']
            #dp_ma480= df['dp_ma480']

            vol_d1_R = df['vol_d1_R']
            # bband30_upR = df['bbands30_upR']
            # bband30_downR = df['bbands30_downR']
            # bbands_width= df['bbands_width']
            # macd_line = df['macd_line5,34,7']
            # macd_sig = df['macd_sig5,34,7']
            # macd_histo = df['macd_histo5,34,7']
            print("완료2")

            # rsi20 = df['rsi20']

            bf1_open_list = []
            bf1_close_list = []
            bf1_high_list = []
            bf1_low_list = []
            bf1_change_R_list = []
            bf1_v20maR_list = []

            #bf1_dp_ma3 = []
            bf1_dp_ma5 = []
            bf1_dp_ma10 = []
            bf1_dp_ma20 = []

            # bf1_dp_ma60 = []
            # bf1_dp_ma120 = []
            #bf1_dp_ma480 = []

            bf1_vol_d1_R = []
            bf1_bband30_upR = []
            bf1_bband30_downR = []
            bf1_bbands_width = []
            bf1_macd_line = []
            bf1_macd_sig = []
            bf1_macd_histo = []

            bf1_rsi20 = []

            for i in range(len(op)):

                if i == 0:  # 가장 과거 데이터는 전일 정보가 없으므로 0추가 (과거 → 현재), 0부터 1을 더뺀다
                    bf1_open_list.append(0)
                    bf1_close_list.append(0)
                    bf1_high_list.append(0)
                    bf1_low_list.append(0)
                    bf1_change_R_list.append(0)
                    bf1_v20maR_list.append(0)

                    #bf1_dp_ma3.append(0)
                    bf1_dp_ma5.append(0)
                    bf1_dp_ma10.append(0)
                    bf1_dp_ma20.append(0)

                    # bf1_dp_ma60.append(0)
                    # bf1_dp_ma120.append(0)
                    #bf1_dp_ma480.append(0)

                    bf1_vol_d1_R.append(0)
                    bf1_bband30_upR.append(0)
                    bf1_bband30_downR.append(0)
                    bf1_bbands_width.append(0)
                    bf1_macd_line.append(0)
                    bf1_macd_sig.append(0)
                    bf1_macd_histo.append(0)

                    bf1_rsi20.append(0)

                else:  #둘째날 데이터

                    bf1_open_list.append(op[i-1])
                    bf1_close_list.append(cl[i-1])
                    bf1_high_list.append(hi[i-1])
                    bf1_low_list.append(lo[i-1])
                    bf1_change_R_list.append(ch[i-1])
                    bf1_v20maR_list.append(v20[i-1])

                    #bf1_dp_ma3.append(dp_ma3[i-1])
                    bf1_dp_ma5.append(dp_ma5[i-1])
                    bf1_dp_ma10.append(dp_ma10[i-1])
                    bf1_dp_ma20.append(dp_ma20[i-1])

                    # bf1_dp_ma60.append(dp_ma60[i-1])
                    #
                    # bf1_dp_ma120.append(dp_ma120[i-1])
                    #bf1_dp_ma480.append(dp_ma480[i-1])

                    bf1_vol_d1_R.append(vol_d1_R[i-1])
                    bf1_bband30_upR.append(bband30_upR[i-1])
                    bf1_bband30_downR.append(bband30_downR[i-1])
                    bf1_bbands_width.append(bbands_width[i-1])
                    bf1_macd_line.append(macd_line[i-1])
                    bf1_macd_sig.append(macd_sig[i-1])
                    bf1_macd_histo.append(macd_histo[i-1])

                    bf1_rsi20.append(rsi20[i-1])

            df['bf1_open'] = bf1_open_list
            df['bf1_close'] = bf1_close_list
            df['bf1_high'] = bf1_high_list
            df['bf1_low'] = bf1_low_list
            df['bf1_change_R'] =bf1_change_R_list
            df['bf1_v20maR'] =bf1_v20maR_list
            #df['bf1_dp_ma3'] = bf1_dp_ma3
            df['bf1_dp_ma5'] = bf1_dp_ma5
            df['bf1_dp_ma10'] = bf1_dp_ma10
            df['bf1_dp_ma20'] =bf1_dp_ma20

            # df['bf1_dp_ma60'] =bf1_dp_ma60
            # df['bf1_dp_ma120'] =bf1_dp_ma120

            df['bf1_vol_d1_R'] =bf1_vol_d1_R
            df['bf1_bband30_upR'] =bf1_bband30_upR
            df['bf1_bband30_downR'] =bf1_bband30_downR
            df['bf1_bbands_width'] =bf1_bbands_width

            df['bf1_rsi20'] =bf1_rsi20

        con.close()
        return df

    def condition_judge(self,df): #조건 판단열 삽입, 통합전에 삽입해야함
        ch_r_list = df['change_ratio']
        condition_list = []
        print("여기 진행중3")
        for i in range(len(ch_r_list)):
            ch_r = ch_r_list[i]
            if ch_r > 0.1 and ch_r < 0.2:
                condition_list.append(1)
            else:
                condition_list.append(0)

        condition_listA = []
        for j in range(len(condition_list)):
            condi_list = condition_list[j:j+5]     #5일 안에 조건을 만족하면 모두다 1이라 쓴다.
            if max(condi_list) == 1:
                condition_listA.append(1)
            else:
                condition_listA.append(0)
        print(len(condition_list))
        print(len(condition_listA))
        df['condition'] = condition_listA
        df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
        print("내부 함수 완료")
        return df

    def add_condition_judge(self):  # 조건 판단열 삽입, 통합전에 삽입해야함
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db")  # 데이터 로드
        name_list = code_name['name']
        for i, name in enumerate(name_list):
            print(name)
            df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)
            df2 = self.condition_judge(df)
            df2.to_sql(name, con, if_exists="replace")
            print(name,"완료")
        con.close()

    def add_past_Day_ohlc(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data.db")  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
        #for i in [523, 524, 525, 526]:  # 뺑뺑이 돌리기
            #name = name_list[i]
            print(name + " 진입")
            df = self.past_Day_ohlc(name)
            #df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.

            if len(df['open']) > 0:    #값이 없는 것은 pass
                df.to_sql(name, con2, if_exists="replace")
                print("완료5")
            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("past_day_ohlcv 최종완료")


    def drop_data(self): #한열 삭제
        self.kospi_kosdaq = self.comboBox_2.currentText()
        con2 = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        code_name = pd.read_sql("SELECT * FROM kosdaq", con2, index_col="index")
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        con2.close()
        con = sqlite3.connect("c:/users/백/kosdaq_data2.db")  # 키움증권 다운로드 종목 데이터 베이스
        col_name = self.lineEdit_2.text()
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col= "index")

            df.drop(col_name, axis=1, inplace=True)   #column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            # dataframe.drop("컬럼 이름", axis=1, inplace=True)
            # 여기서 axis는 축을 의미하고 0은 row를 1은 column을 의미하게 됩니다.
            # inplace=True라고 설정을 해야지 바로 drop의 동작을 실행합니다.

            #df.rename(index={"index", "date"}, inplace=True)
            #df.rename(columns={"index": "date"}, inplace=True)  #칼럼명 변경
            #print(df.head())
            df.to_sql(name, con, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        print("최종완료")
        con.close()


    def merge___(self):   #모든 데이터 한테이블 통합
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"
        file = "c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db"

        con2 = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        # con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_add_manual.db")  # 키움증권 다운로드 종목 데이터 베이스
        # con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_month_index.db")  # 키움증권 다운로드 종목 데이터 베이스

        c = con2.cursor()

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            if i > 0:
                df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con2, index_col="index")

                if len(df['open']) > 0:   #빈데이터를 패스하기 위하여
                    print("%d / %d %s 진입" %(i, len(name_list), name))
                    if self.kospi_kosdaq == "kospi":
                        c.execute("INSERT INTO 동화약품 SELECT * FROM " + "'" + name + "'")   #코스피
                        con2.commit()
                    else:
                        c.execute("INSERT INTO 삼천당제약 SELECT * FROM " + "'" + name + "'")
                        con2.commit()

                    print(name+"추가 완료")
                c.execute("DROP TABLE " + "'" + name + "'")  # 테이블 삭제
                print(name + "삭제 완료")
        c.close()
        con2.close()
        #os.rename("kosdaq_data_add_auto.db", "kosdaq_data_one_tabel.db")
        print("한테이블 통합 최종완료, %s로 저장" % file)

    def add_minute_tech(self, name):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_minu_data.db")  #분봉데이터 로드
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)
        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다

            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            vo = df['volume'] * 0.1 * 10
            index_list = []

            df['date'] = df['index']
            df.drop("index", axis=1, inplace=True)   # axis=1 은 열을 의미한다.

            for i in range(len(df['date'])):
                date = df['date'][i]
                time = df['time'][i]
                index = str(date) + ";" + str(time)
                index_list.append(index)
            df['index'] = index_list
            cl_list = list(cl)
            cl_list2 = cl_list[1:]
            cl_list2.append(cl_list2[-1])
            df['cl2'] = cl_list2
            df['change_2'] = df['close']-df['cl2']
            df = DataFrame(df, columns= ['index', 'date', 'time', 'name', 'open', 'high', 'low', 'close', 'change', 'volume', 'change_ratio'])
        return df

    def add_min_tech_button(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        file_name = self.kospi_kosdaq + "_minu_data.db"
        con = sqlite3.connect("c:/users/백/" + file_name)  # 분봉데이터 로드

        code_name = self.codeNname_load()  # 종목 코드로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.add_minute_tech(name)
            df.to_sql(name, con, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con.close()
        print("최종완료, %s 로 저장 완료" % file_name)




    def add_month_change(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for j, name in enumerate(name_list):

            print("현재 작성중인 종목 : %s (%s / %s)" %(name, j, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col='index')  # 데이터 저장순서는 현재부터 과거로다
            cl_list = list(df["close"])
            change_list = []
            change_ratio_list = []
            data_lenth = len(df["change"])
            # df.drop("change", axis=1, inplace=True)
            # df.drop("change_ratio", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            # dataframe.drop("컬럼 이름", axis=1, inplace=True)
            # 여기서 axis는 축을 의미하고 0은 row를 1은 column을 의미하게 됩니다.
            # inplace=True라고 설정을 해야지 바로 drop의 동작을 실행합니다.

            # df.rename(index={"index", "date"}, inplace=True)
            # df.rename(columns={"index": "date"}, inplace=True)  #칼럼명 변경
            # print(df.head())

            for i in range(len(cl_list)):
                if i == data_lenth-1:
                    change_list.append(0)
                    change_ratio_list.append(0)

                else:
                    new_change = cl_list[i] - cl_list[i+1]
                    change_list.append(new_change)
                    if cl_list[i+1] == 0:               #종가가 0인경우 대비
                        change_ratio_list.append(0)
                    else:
                        new_change_ratio = new_change / cl_list[i+1]
                        change_ratio_list.append(new_change_ratio)
            df["change"] = change_list
            df["change_ratio"] = change_ratio_list
            df.to_sql(name, con, if_exists="replace")
        con.close()


        # for i, name in enumerate(name_list):  # 뺑뺑이 돌리기

    def execute_query(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        query_ = self.lineEdit.text()   #입력된 값을 저장
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db")  # 키움증권 다운로드 종목 데이터 베이스

        if self.kospi_kosdaq == "kospi":
            print("완료1")
            df = pd.read_sql("SELECT * FROM 동화약품 WHERE " + query_, con, index_col=None)   #코스피

        else:
            print("완료1")
            df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)

        con.close()
        print("완료2")
        self.simul_data_to_excel(df)


    def add_high_low(self, name, day):
        self.kospi_kosdaq = self.comboBox_2.currentText()

        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        try:
            df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con,
                             index_col=None)  # df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        except:
            return

        day_max_list = []
        day_min_list = []
        final_close_list = [] #최종 종가

        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다
            hi = list(df['high'] * 0.1 * 10)
            lo = list(df['low'] * 0.1 * 10)
            cl = list(df['close'] * 0.1 * 10)

            for i in range(len(hi)):

                if i == len(hi)-1:
                    day_max_list.append(0)
                    day_min_list.append(0)
                    final_close_list.append(0)

                else:
                    day_high_list = hi[i + 1: i + day+1]  # 다음날 부터 + 향후 X일간의 리스트
                    day_high_list.sort(reverse=True)  # 최대값은 리스트의 첫번째 값
                    day_max_list.append(day_high_list[0])
                    last_close_list = cl[i+1:i+1+day]
                    last_close = last_close_list[-1]
                    final_close_list.append(last_close)

                    day_low_list = lo[i + 1: i + day+1]  # 다음날 부터 + 향후 X일간의 리스트
                    day_low_list.sort()  # 최소값은 리스트의 첫번째 값
                    day_min_list.append(day_low_list[0])

            df['hi_max'] = day_max_list
            df['low_min'] = day_min_list
            df['D+close'] = final_close_list

        con.close()
        return df

    def add_day_max_min(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/%s_data_high_low.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        day = self.spinBox.value()

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.add_high_low(name, day)
            # df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("high low 최종완료")


    def add_month_index(self, name):   #종목 월검색 열 추가

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        month_index_list = []
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)  # 데이터 저장순서는 현재부터 과거로다
        date_list = df['index']
        name_ = df['name'][0]
        print(name_)
        for date in date_list:
            month_index = str(date)[:6]
            name_month = name+";"+month_index
            month_index_list.append(name_month)
        con.close()
        df['month_index'] = month_index_list
        print(name_+"완료")
        return df


    def add_month_index_all(self):
        code_name = self.codeNname_load()  # 종목 코드 로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con2 = sqlite3.connect("c:/users/백/%s_data_month_index.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        for i, name in enumerate(name_list):
            df = self.add_month_index(name)
            df.to_sql(name, con2, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("월봉인덱스 삽입 완료")

    def read_lookup_data_from_excel(self, filename):
        wb = load_workbook(filename=filename, data_only=True)
        sheet_range = wb["test1"]

        data1 = sheet_range['B:B']  #종목명
        data2 = sheet_range['C:C']  #날짜

        stock_list = []
        date_list = []

        for i, temp_stock in enumerate(data1):
            stock = temp_stock.value
            date = data2[i].value
            if stock == None:
                break
            stock_list.append(stock)
            date_list.append(date)
        df_dic = {'date': date_list, 'name': stock_list}
        df = DataFrame(df_dic)

        return df


    def read_day_info_from_month(self):   #엑셀의 월봉데이터에 해당하는 일봉데이터를 엑셀로 다시 저장
        self.kospi_kosdaq = self.comboBox_2.currentText()
        compare_list = self.read_excel()  #불러올 종목+월
        file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"

        con = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        size = len(compare_list)
        df2__ = {}
        df2 = DataFrame(df2__)
        for i, compare_ in enumerate(compare_list):
            query_ = "month_index == '" + compare_ + "'"
            print(query_)
            if self.kospi_kosdaq == "kospi":
                full_query = "SELECT * FROM 동화약품 WHERE " + query_
                print(full_query)
                df = pd.read_sql(full_query, con, index_col=None)  # 코스피
            else:
                df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)
            print("진행 %s / %s" % (i, size))
            if i == 0:
                df2 = df
            else:
                df2 = df2.append(df)
        print(df2)
        con.close()
        print("완료2")
        self.simul_data_to_excel(df2)

    def read_day_info_from_month_new(self):   #엑셀의 월봉데이터에 해당하는 일봉데이터를 엑셀로 다시 저장
        self.kospi_kosdaq = self.comboBox_2.currentText()
        date_and_name = self.read_lookup_data_from_excel()  #불러올 종목+월
        date_list = list(date_and_name['date'])
        print(date_list)
        name_list = list(date_and_name['name'])

        # file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"
        file = "c:/users/백/" + self.kospi_kosdaq + "_30minute.db"
        print(file)

        con = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        size = len(date_list)
        df2__ = {}
        df2 = DataFrame(df2__)
        for i, date_ in enumerate(date_list):
            date = str(date_)
            name = name_list[i]
            print(name)
            query_ = "date == '" + date + "'"
            # query_ = "'index' == " + date

            final_query = "SELECT * FROM " + name + " WHERE " + query_

            print(query_)
            print(final_query)

            # df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)

            df = pd.read_sql(final_query, con, index_col=None)
            print("진행 %s / %s" % (i, size))
            if i == 0:
                df2 = df
            else:
                df2 = df2.append(df)
        print(df2)
        con.close()

        self.simul_data_to_excel(df2)

    def del_col(self):   #수정 필요함
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/%s_data_high_low.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        # name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
        #     print(name + " 진입")
        #     df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
        #     df.to_sql(name, con2, if_exists="replace")
        #
        #     print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("삭제 최종완료")

    def get_minute_info_form_day_info(self):   #일봉정보를 분봉정보로 가져오기

        minute_file = "c:/users/백/minute_test.db"    #분봉 파일
        con1 = sqlite3.connect(minute_file)  # 분봉 데이터 베이스

        filename = "C:\\Users\\백\\PycharmProjects\\gaebal\\test.xlsx"
        day_data = self.read_lookup_data_from_excel(filename)    #엑셀 데이터 읽기
        date_list = day_data['date']
        df2 = DataFrame()

        for i in range(len(date_list)):
            date = day_data['date'][i]
            name = day_data['name'][i]
            query = "date = "+ str(date)
            # query = "index = "+ str(date)
            print(name, query)
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query, con1, index_col=None)  # 코스피
            df2 = df2.append(df)   #df2.append(df)만 하면 안된다.
        print(df2)
        self.simul_data_to_excel(df2)
        con1.close()


    def test___(self):
        # self.add_month_change()
        # self.tech__("삼천당제약")  # 한종목의 기술적 지표 계산
        print("시작")
        # self.add_month_index_all()
        # df = self.read_lookup_data_from_excel()
        # self.read_day_info_from_month_new()
        # self.read_day_info_from_month()
        self.test2()
        # self.add_minute_tech("삼천당제약")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    baek_data_= baek_data_analysis2()
    baek_data_.show()
    app.exec_()
