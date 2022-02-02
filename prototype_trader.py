import sys
import pandas as pd
import pandas_datareader as pdr

import sqlite3
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic

form_class = uic.loadUiType("prototype.ui")[0]

class prototype_trader(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.timer = QTimer(self)
        self.timer.start(1000)
        self.timer.timeout.connect(self.timeout)

        self.pushButton.clicked.connect(self.DownloadStockCode)

    def timeout(self):
        current_time = QTime.currentTime()
        text_time = current_time.toString("hh:mm:ss")
        time_msg = "현재시간: " + text_time
        self.statusbar.showMessage(time_msg)

    def DownloadStockCode(self):
        df = pd.read_html('http://kind.krx.co.kr/corpgeneral/corpList.do?method=download', header=0)[
            0]  # 한국거래소 홈페이지에서 자동으로 종목코드를 다운로드한다.
        df_columns = df.columns
        df.drop([df_columns[3], df_columns[4], df_columns[5], df_columns[6], df_columns[7], df_columns[8]], axis=1,
                inplace=True)
        df.rename(columns={'회사명': 'name', '종목코드': 'code', '업종': 'category'}, inplace=True)
        df['code'] = df['code'].astype('str')  # 코드를 문자열 타입으로 바꾸기
        df['code'] = df['code'].str.zfill(6)  # 6 자리만큼 0으로 채운다.
        df = df.query("category != '신탁업 및 집합투자업' and category != '금융 지원 서비스업' \
        and category != '기타 금융업'")  # 스팩, 지주, 투자회사 제거
        msg = "종목코드 다운로드 완료"
        self.statusbar.showMessage(msg)
        return df

    # 삼성전자 - 005930
    def download_basic_data(self, code='005930', start_date='20130102'):  # 데이터 다운로드
        # data = pdr.get_data_yahoo(code + '.KS', start='20190101')

        jisu = Download_Jisu(start=start_date)  # 지수 다운로드

        data = pdr.naver.NaverDailyReader(code, start=start_date).read()

        df = pd.DataFrame(data)
        df.reset_index(inplace=True)
        df['Date'] = df['Date'].dt.strftime('%Y%m%d')  # 20220128 이런 형식으로 변경
        df['Open'] = pd.to_numeric(df['Open'])
        df['Close'] = pd.to_numeric(df['Close'])
        df['High'] = pd.to_numeric(df['High'])
        df['Low'] = pd.to_numeric(df['Low'])
        df['Change'] = df['Close'].diff()
        df['change_ratio'] = df['Close'].pct_change()  # 변화량을 퍼센테이지로 구한다.
        df['change_ratio'] = df['change_ratio'].round(5)
        df['Volume'] = pd.to_numeric(df['Volume'])  # 데이터 타입을 숫자형으로 바꾸기

        df = pd.merge(df, jisu, on='Date', how='left')

        # df = df.set_index(keys='index', drop=True)          # 인덱스 이름을 index로 바꾸기, drop=False/True ~ 원본 index열 유지/제거.
        return df

    def calculate_technical_index(self, df):  # 한종목의 기술적 지표 추가
        op = df['Open']
        cl = df['Close']
        shift_close = cl.shift()  # 전일 종가
        hi = df['High']
        lo = df['Low']
        vo = df['Volume']

        df['Open_R'] = (op - shift_close) / shift_close
        df['High_R'] = (hi - shift_close) / shift_close
        df['Low_R'] = (lo - shift_close) / shift_close
        df['Vol_R'] = vo.pct_change()
        df['Vol_R'] = df['Vol_R'].round(5)  # 전일 대비 거래량

        # 기준 기간 중 최고 최저가
        max_10_list = df['Close'].rolling(window=10).max()
        min_10_list = df['Close'].rolling(window=10).min()
        max_20_list = df['Close'].rolling(window=20).max()
        min_20_list = df['Close'].rolling(window=20).min()
        max_55_list = df['Close'].rolling(window=55).max()
        min_55_list = df['Close'].rolling(window=55).min()
        max_60_list = df['Close'].rolling(window=60).max()
        min_60_list = df['Close'].rolling(window=60).min()

        # 기준 기간 중 최고/최저가 비율
        df['max_10_R'] = (df['Close'] - max_10_list) / max_10_list  # 10일 최고가대비 얼마나 하락했는가
        df['min_10_R'] = (df['Close'] - min_10_list) / min_10_list  # 10일 최저가대비 얼마나 상승했는가
        df['max_20_R'] = (df['Close'] - max_20_list) / max_20_list
        df['min_20_R'] = (df['Close'] - min_20_list) / min_20_list
        df['max_55_R'] = (df['Close'] - max_55_list) / max_55_list
        df['min_55_R'] = (df['Close'] - min_55_list) / min_55_list
        df['max_60_R'] = (df['Close'] - max_60_list) / max_60_list
        df['min_60_R'] = (df['Close'] - min_60_list) / min_60_list

        df['max_10_R'] = df['max_10_R'].round(5)
        df['min_10_R'] = df['min_10_R'].round(5)
        df['max_20_R'] = df['max_20_R'].round(5)
        df['min_20_R'] = df['min_20_R'].round(5)
        df['max_55_R'] = df['max_55_R'].round(5)
        df['min_55_R'] = df['min_55_R'].round(5)
        df['max_60_R'] = df['max_60_R'].round(5)
        df['min_60_R'] = df['min_60_R'].round(5)

        # # 열순서 정렬
        # df = df[['name', 'Open', 'High', 'Low', 'Close', 'Volume', 'Change', 'change_ratio', 'Open_R',
        #          'High_R', 'Low_R', 'Vol_R', 'atr20(N)', 'vma20_R', 'dp_ma3', 'dp_ma5', 'dp_ma10', 'dp_ma20',
        #          'dp_ma60', 'bbands_width', 'bong', 'N20_R', 'd1_open', 'd1_high', 'd1_low', 'd1_close',
        #          'd1_open_R', 'd1_high_R', 'd1_low_R', 'd1_close_R',
        #          'd2_open', 'd2_high', 'd2_low', 'd2_close', 'd2_open_R', 'd2_high_R', 'd2_low_R', 'd2_close_R',
        #          'max_10_R', 'min_10_R', 'max_20_R', 'min_20_R', 'max_55_R', 'min_55_R', 'max_60_R', 'min_60_R']]
        return df

    def DataToSqlite(self, df, name='삼성전자', file_name='Calculate_tech_test_data'):  # Sqlite로 데이터 저장 이름 = 삼성전자

        file_name = file_name + '.db'
        con = sqlite3.connect("c:/users/백/%s" % file_name)  # 키움증권 다운로드 종목 데이터 베이스
        df.to_sql(name, con, if_exists="replace")
        con.close()
        print(name, "Sqlite3로 저장완료")

    def LoadDataFromSqlite(self, file_name='Calculate_tech_test_data'):  # Sqlite로부터 데이터 불러오기
        file_name = file_name + '.db'
        name = '삼성전자'
        con = sqlite3.connect("c:/users/백/%s" % file_name)  # 키움증권 다운로드 종목 데이터 베이스
        df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')  # 인덱스 칼럼을 Date로 지정
        con.close()
        print(name, "Sqlite3로 불러오기 완료")
        return df

    def UpdateBasicDateAndTechnicalIndex(self):  # code = 삼성전자,
        today_weekday = datetime.today().weekday()  # 오늘 요일, 0 ~ 월요일, 6 ~ 일요일
        delta_date = 0
        if today_weekday == 5:  # 오늘 토요일이면
            delta_date = 1
        elif today_weekday == 6:  # 오늘이 일요일이면
            delta_date = 1

        last_date = datetime.today() - timedelta(delta_date)  # 마지막날 토일 제외
        last_date = last_date.strftime("%Y%m%d")

        df = LoadDataFromSqlite()
        last_date_from_data = df.index[-1]  # index는 리스트 형식으므로 iloc를 쓰면 안된다는 걸 기억할 것

        if last_date_from_data == last_date:
            print("데이터가 이미 최신입니다.")
            exit()  # 데이터가 최신이면 프로그램을 종료한다.
        else:
            date_100day_ago = (datetime.today() - timedelta(120)).strftime("%Y%m%d")  # 100일전 날짜
            basic_data = download_basic_data('005930', start_date=date_100day_ago)
            data_calculated_techniclal = calculate_technical_index(basic_data)
            df = df.append(data_calculated_techniclal).drop_duplicates(subset=['Date'])
            df.reset_index(inplace=True)
            # new_data = df.append(data_calculated_techniclal).sort_values(by='Date')
            return df  # 최신화된 데이터를 리턴한다.

    def run_tutle_strategy(self, df):  # 터틀전략 실행
        close_list = df.Close
        event = 0
        event_list = []
        position = 0
        position_list = []
        order_list = []
        buy_type = 0
        buy_type_list = []
        buy_price = 0
        buy_price_list = []
        sell_price = 0
        sell_price_list = []
        hold_day = 0
        hold_day_list = []  # 보유기간
        open_profit = 0
        open_profit_list = []
        close_profit = 0
        close_profit_list = []
        winnloss = 0
        winnloss_list = []

        # stop_loss_price_list = np.zeros(len(close_list))  # 나중에 추가할 것

        max_20_list = df.max_20_R
        max_55_list = df.max_55_R
        min_10_list = df.min_10_R
        min_20_list = df.min_20_R

        for i, close in enumerate(close_list):
            max_20 = max_20_list.iloc[i]
            max_55 = max_55_list.iloc[i]
            min_10 = min_10_list.iloc[i]
            min_20 = min_20_list.iloc[i]

            if max_20 == 0 and winnloss == 0 and position == 0:  # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0)
                event = 1
                event_list.append(event)
                order = 1
                order_list.append(order)
                position = 1
                position_list.append(position)
                buy_type = "S1"
                buy_type_list.append(buy_type)

                buy_price = close
                buy_price_list.append(buy_price)
                sell_price = 0
                sell_price_list.append(sell_price)
                hold_day = 0
                hold_day_list.append(hold_day)
                open_profit = 0
                open_profit_list.append(open_profit)
                close_profit = 0
                close_profit_list.append(close_profit)
                winnloss = 0
                winnloss_list.append(winnloss)

            # if max_20 == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 0 and position == 0:
            # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
            # if max_20 == 0 and N != None and N_ratio > 0.01 and winnloss == 0 and position == 0:
            #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거

            elif max_55 == 0 and winnloss == 1 and position == 0:  # S2 신규매입 조건 : 55일 신고가, 직전거래에서 수익시(winnloss == 1)

                # elif max_55 == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 1 and position == 0:
                # S2 매수조건 진입, 60이평 위 조건 추가
                # elif max_55 == 0 and N != None and N_ratio > 0.01 and winnloss == 1 and position == 0:
                #     # S2 매수조건 진입, 60이평 위 조건 제거
                event = 1
                event_list.append(event)
                order = 1
                order_list.append(order)
                position = 1
                position_list.append(position)
                buy_type = "S2"
                buy_type_list.append(buy_type)
                buy_price = close
                buy_price_list.append(buy_price)
                sell_price = 0
                sell_price_list.append(sell_price)
                hold_day = 0
                hold_day_list.append(hold_day)
                open_profit = 0
                open_profit_list.append(open_profit)
                close_profit = 0
                close_profit_list.append(close_profit)
                winnloss = 1
                winnloss_list.append(winnloss)

            elif min_10 == 0 and buy_type == "S1" and position == 1:  # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파
                event = 1
                event_list.append(event)
                order = -1
                order_list.append(order)
                position = 0
                position_list.append(position)

                buy_type_list.append(buy_type)
                buy_type = 0

                close_profit = close - buy_price

                buy_price = 0
                buy_price_list.append(buy_price)
                sell_price = close
                sell_price_list.append(sell_price)
                hold_day += 1
                hold_day_list.append(hold_day)
                open_profit = 0
                open_profit_list.append(open_profit)

                close_profit_list.append(close_profit)
                if close_profit > 0:
                    winnloss = 1
                else:
                    winnloss = 0
                winnloss_list.append(winnloss)

            elif min_20 == 0 and buy_type == "S2" and position == 1:  # S2 청산 조건 : 종가기준 20일 최저가 하향 돌파
                event = 1
                event_list.append(event)
                order = -1
                order_list.append(order)
                position = 0
                position_list.append(position)
                buy_type_list.append(buy_type)
                buy_type = 0

                close_profit = close - buy_price

                buy_price = 0
                buy_price_list.append(buy_price)
                sell_price = close
                sell_price_list.append(sell_price)
                hold_day += 1
                hold_day_list.append(hold_day)
                open_profit = 0
                open_profit_list.append(open_profit)

                close_profit_list.append(close_profit)
                if close_profit > 0:
                    winnloss = 1
                else:
                    winnloss = 0
                winnloss_list.append(winnloss)

            else:
                event = 0
                event_list.append(event)
                order = 0
                order_list.append(order)
                position_list.append(position)
                buy_type_list.append(buy_type)
                buy_price_list.append(buy_price)
                sell_price = 0
                sell_price_list.append(sell_price)
                if position == 1:
                    hold_day += 1
                    hold_day_list.append(hold_day)
                    open_profit = close - buy_price
                    open_profit_list.append(open_profit)
                else:
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    open_profit = 0
                    open_profit_list.append(open_profit)
                close_profit = 0
                close_profit_list.append(close_profit)
                winnloss_list.append(winnloss)
        df['event'] = event_list
        df['order'] = order_list
        df['position'] = position_list
        df['buy_type'] = buy_type_list
        df['buy_price'] = buy_price_list
        df['hold_day'] = hold_day_list
        df['stock_open_profit'] = open_profit_list
        df['stock_close_profit'] = close_profit_list
        df['winnloss'] = winnloss_list
        df['stock_cum_profit'] = df['stock_close_profit'].cumsum() + df['stock_open_profit']  # 누적 수익
        df['stock_day20_profit'] = df['stock_close_profit'].rolling(window=20).sum() + df[
            'stock_open_profit']  # 최근 20일 수익
        return df

    def DataToExcel(self, data, save_name="data_result"):  # contine : 이어쓰기 여부 1이면 이어쓰기임   @@@@@@@@@@@@ 수정필요함

        path = "c:/users/백/save_excel.xlsx"
        wb = load_workbook(path, data_only=True)
        worksheet_1 = wb.active
        worksheet_1.delete_cols(1, 100)  # A열 이후 100열 삭제

        # Dataframe을 엑셀로 뿌린다.
        maxrow = worksheet_1.max_row
        if maxrow == 1:
            header = True
        else:
            header = False

        for i, row in enumerate(dataframe_to_rows(data, index=True, header=header)):  # 한줄씩 엑셀로
            if len(row) > 1:
                worksheet_1.append(row)
        maxrow = worksheet_1.max_row
        print("최대행은 %s행 입니다." % maxrow)
        save_name = str(save_name)
        save_path = "c:/users/백/" + save_name + ".xlsx"
        wb.save(save_path)
        print("saved to file name %s.xlsx!!" % save_name)

    def Download_Jisu(self, start='20130102'):  # 지수 다운로드
        kospi_ticker = '^KS11'
        kosdaq_ticker = '^KQ11'

        data_kospi = pdr.get_data_yahoo(kospi_ticker, start=start)
        data_kospi.reset_index(inplace=True)  # inplace=True 원본을 놔두지않고 바로 바꾼다.

        return_data = pd.DataFrame()
        return_data['Date'] = data_kospi['Date'].dt.strftime("%Y%m%d")  # 20220128 이런 형식으로 변경
        return_data['Kospi_Close'] = data_kospi['Adj Close']
        return_data['Kospi_Change_R'] = return_data['Kospi_Close'].pct_change()
        return_data['Kospi_Change_R'] = return_data['Kospi_Change_R'].round(5)
        data_kosdaq = pdr.get_data_yahoo(kosdaq_ticker, start=start)
        data_kosdaq.reset_index(inplace=True)
        return_data['Kosdaq_Close'] = data_kosdaq['Adj Close']
        return_data['Kosdaq_Change_R'] = return_data['Kosdaq_Close'].pct_change()
        return_data['Kosdaq_Change_R'] = return_data['Kosdaq_Change_R'].round(5)
        return return_data

    def test_data(self):
        try:
            data = pd.read_csv("test_data.csv")
        except FileNotFoundError:
            data = download_basic_data('005930')
            data = add_tech(data)
            DataToCSV(data)
        df = run_tutle_strategy(data)
        DataToExcel(df)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    prototype = prototype_trader()
    prototype.show()
    app.exec_()