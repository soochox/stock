import win32com.client  #엑셀을 쓰기위한

excel = win32com.client.Dispatch("Execl.Application")
print("111")
invest_ratio_file = 'C:\\Users\\백\\PycharmProjects\\gaebal\\투자비중(모멘텀,연승).xlsx'
print("여기")
wb = excel.Workbooks.Open(invest_ratio_file)
print("여기2")
ws = wb.Worksheets("모멘텀+연승적용 투자비중")
invest_ratio = ws.Cells(3,2).Value
print(invest_ratio)
excel.Quit()


# from openpyxl import Workbook
# from openpyxl import load_workbook
# import os
#
# base = 'C:\\Users\\백\\PycharmProjects\\gaebal'
# file = '지수 모멘텀.xlsx'
# file_path = os.path.join(base, file)
# wb = load_workbook(file_path, data_only=True)
#
# ws_kospi = wb['kospi']
# ws_kosdaq = wb['kosdaq']
#
# kospi_momentum = ws_kospi['T3'].value
# kosdaq_momentum = ws_kosdaq['T3'].value
# print(kosdaq_momentum, kospi_momentum)





# wb = Workbook()
# ws = wb.active   #기본 시트를 변수 ws에 할당
# ws.title = "test"  #시트이름을 test로 설정
# ws3 = wb["test"]   #시트이름이 설정됐으면 이러게도 접근 가능함
#
# ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# # or
# ws2 = wb.create_sheet("Mysheet2", 0) # insert at first position
# #ws3 = wb.create_sheet("Mysheet", -1) # insert at the last position
#
# print(wb.sheetnames)  #sheetnames로 모든시트이름을 리스트로 저장가능
#
# c = ws['a4']
# ws['a4'] = 4
# d = ws.cell(row=4, column=2, value=10)   #셀을 변수할당과 동시에 셀값도 할당
#
# c.value = "hellow, world"   #값을 할당
# print(c.value)
#
# ### When a worksheet is created in memory, it contains no cells. They are created when first accessed.
#
# #Accessing many cells
# cell_range = ws['A1':'C2']
# colC = ws['C']
# col_range = ws['C:D']
# row10 = ws[10]
# row_range = ws[5:10]

# 배열형태
# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):   #한행씩 row에 들어간다. 데이터형식은 set
#     print(row)
#     for cell in row:
#         print(cell)
#
# for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):   #한열씩
#     for cell in col:
#         print(cell)

##주의 For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.


##################################################################################################

#If you need to iterate through all the rows or columns of a file, you can instead use the Worksheet.rows property
 #파일의 모든 행 또는 열을 반복해야하는 경우 Worksheet.rows 속성을 대신 사용할 수 있습니다.

#print(ws.rows)

##################################################################################################
#Values only
#If you just want the values from a worksheet you can use the Worksheet.values property. This iterates over all the rows in a worksheet but returns just the cell values:

# for row in ws.values:
#    for value in row:
#      print(value)

# ## values_only=True
# for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
#     print(row)

##################################################################################################

# #Saving to a file
# wb.save('test201101.xlsx')   #파일저장   주의!! 파일을 바로 덮어씀
#
# #load_workbook
# wb2 = load_workbook('test201101.xlsx')
# print(wb2.sheetnames)


##################################################################################################

# 1. Write a workbook

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# wb = Workbook()
# filename = 'empty_book.xlsx'
# ws1 = wb.active
# ws1.title = "range names"
#
# for row in range(1, 40):
#     ws1.append(range(600))   #append 다음행에 이어서 쓴다.
# ws2 = wb.create_sheet(title="Pi")
# ws2['F5'] = 3.14
# ws3 = wb.create_sheet(title="Data")
#
# for row in range(10, 20):
#     for col in range(27, 54):
#         # _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))    #get_column_letter 열이름 갖고 오기
#         ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))    # "{0}.format(get_column_letter(col) 이부분은 print문과 같은형식
#
#         #print문   {순서1} {순서2}    문자열.format(출력값1,출력값2)
#
# print("{0}" .format(ws3['AA10'].value))
# wb.save(filename = filename)
#
#
# ##################################################################################################
# #행열 삽입
# ws.insert_rows(7)
#
# #행열 삭제
# ws.
# ws.delete_cols(6, 3)
#
# wb = load_workbook('test201101.xlsx')
# ws = wb['Mysheet2']
# max_row = ws.max_row
# print(max_row)
