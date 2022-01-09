import sys
import time
import datetime
import telegram

from pywinauto import application
import win32com.client

from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from PyQt5.QtCore import *
from PyQt5 import uic
import os

import pandas as pd
import pandas_datareader as pdr
from pandas import Series, DataFrame

import pandas.io.sql as pd_sql
import sqlite3
import talib as ta
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from talib.abstract import *
import FinanceDataReader as fdr
import numpy as np
import winsound

form_class = uic.loadUiType("Serch_UI.ui")[0]

class baek_data_analysis2(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.timer = QTimer(self)
        self.timer.start(1000)
        self.timer.timeout.connect(self.timeout)
        self.recommanded_done = False

        self.pushButton.clicked.connect(self.execute_query)
        self.pushButton_2.clicked.connect(self.add_tech)
        self.pushButton_3.clicked.connect(self.add_tech_manual)
        self.pushButton_5.clicked.connect(self.merge___)
        self.pushButton_6.clicked.connect(self.del_column)
        self.pushButton_8.clicked.connect(self.test___)
        self.pushButton_10.clicked.connect(self.add_day_max_min)
        self.pushButton_12.clicked.connect(self.add_month_change)
        self.pushButton_13.clicked.connect(self.add_month_index_all)
        self.pushButton_14.clicked.connect(self.add_min_tech_button)
        self.pushButton_15.clicked.connect(self.get_minute_info_form_day_info)  # 선택된 일봉정보의 분봉정보로 가져오기
        self.pushButton_16.clicked.connect(self.execute_query_for_all_stock)  # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행
        self.pushButton_7.clicked.connect(self.execute_query_for_all_stock_plus_past_future)  # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행 + 과거/미래
        self.pushButton_4.clicked.connect(self.calc_today_profit)    # 오늘까지 수익을 시뮬레이션하여 엑셀로 전달

        # self.pushButton_9.clicked.connect(self.condition_filter)  # 조건 필터 추가
        self.pushButton_9.clicked.connect(self.condition_filter_etf)  # 조건 필터 추가

        self.pushButton_17.clicked.connect(self.data_read)  # 일봉 다운로드
        self.pushButton_18.clicked.connect(self.download_jisu)  # 지수 다운로드
        self.pushButton_19.clicked.connect(self.daily_choochun)  # 지수 다운로드
        self.sound_play()

    def timeout(self):
        recommand_time = QTime(15, 0, 0)        # 종목 추천 시간 (3시 2분)
        end_time = QTime(15, 8, 0)          # 종목 추천 종료 시간 (3시 8분)
        # end_time = QTime(23, 59, 0)
        current_time = QTime.currentTime()
        text_time = current_time.toString("hh:mm:ss")
        time_msg = "현재시간: " + text_time
        self.statusbar.showMessage(time_msg)
        if (end_time > current_time > recommand_time) and self.recommanded_done is False and self.checkBox.isChecked():
            self.daily_choochun()
            self.recommanded_done = True

    def send_msg_telegram(self, msg='TEST'):
        TOKEN = '1968085479:AAEck4QaWFrY9Bl7tNKRaqF3lNe0dIrr3Kg'
        mc = '1956916092'
        bot = telegram.Bot(TOKEN)
        bot.sendMessage(mc, msg)

    def daily_choochun(self):       # 당일 추천 종목을 골라 텔레그램으로 메세지를 보낸다.
        self.data_read2('etf_real')
        time.sleep(1)
        self.add_tech()
        print("자동 기술적 지표 완료")
        time.sleep(1)
        self.add_tech_manual()
        time.sleep(1)
        self.condition_filter_etf()
        time.sleep(1)
        today = datetime.datetime.today().strftime("%Y-%m-%d")

        query = "event > 0 and date = '" + today + " 00:00:00'"
        today_result = self.execute_query_input('etf_real', query)

        name_list = list(today_result['name'])
        N_list = list(today_result['N20(%)'])
        buy_sell_list = list(today_result['buy_sell'])

        final_names = []
        final_N20s = []
        final_trades = []

        for i, today_N20 in enumerate(N_list):
            if today_N20 > 0.02:
                final_name = name_list[i]
                final_names.append(final_name)
                final_N20 = today_N20
                final_N20 = round(final_N20 * 100, 3)
                final_N20s.append(final_N20)
                final_trade = buy_sell_list[i]
                final_trades.append(final_trade)

                print("종목명: %s, N: %s, 매수/매도: %s" % (final_name, final_N20, final_trade))

                msg = "종목명: " + final_name + ", N:" + str(final_N20) + ", 매수/매도: " + final_trade
                self.send_msg_telegram(msg)

        if len(final_names) == 0:
            msg = "금일 ETF 추천 종목 없음."
            self.send_msg_telegram(msg)

        print("eft완료")

        self.data_read2('quant')
        time.sleep(1)
        self.add_tech('quant')
        time.sleep(1)
        self.add_tech_manual('quant')
        time.sleep(1)
        self.condition_filter_quant()
        time.sleep(1)
        today = datetime.datetime.today().strftime("%Y-%m-%d")
        # today = "2021-09-17"
        query = "event > 0 and date = '" + today + " 00:00:00'"
        today_result = self.execute_query_input('quant', query)
        name_list = list(today_result['name'])
        N_list = list(today_result['N20(%)'])
        buy_sell_list = list(today_result['buy_sell'])

        if len(name_list) > 0:
            for i in range(len(name_list)):
                name = name_list[i]
                N = N_list[i]
                N = round(N, 4) * 100
                buy_sell = buy_sell_list[i]
                print("종목명: %s, N: %s, 매수/매도: %s" % (name, N, buy_sell))
                msg = "종목명: " + name + ", N:" + str(N) + ", 매수/매도: " + buy_sell
                self.send_msg_telegram(msg)
        else:
            msg = "금일 퀀트 추천 종목 없음."
            self.send_msg_telegram(msg)
        print("퀀트 완료")

    def execute_query_input(self, kospi_kosdaq="etf_real", query="event > 0"):    # 쿼리 조건값을 입력받아 실행

        code_name = self.codeNname_load2(kospi_kosdaq)  # 종목 코드 로드
        # con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장


        df2 = DataFrame()
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print("%s 진입 %s / %s" % (name, i+1, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query, con, index_col='index')
            df2 = df2.append(df)
        print(df2)
        con.close()
        print("완료")
        return df2

    def download_jisu(self):
        tickers = ['^KS11', '^KQ11']      # 코스피, 코스닥
        names = ['kospi', 'kosdaq']
        file_name = "jisu_data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        for i, code in enumerate(tickers):  # 뺑뺑이 돌리기
            date_list = []
            name = names[i]
            print(name + " 진입")

            data = pdr.get_data_yahoo(code, start='20190101')

            print(i + 1, "/", len(names), name, " 완료")
            date_index = data.index
            for date in date_index:
                date = str(date)
                date = date[:4] + date[5:7] + date[8:10]
                date = int(date)
                date_list.append(date)
            df = pd.DataFrame(data)
            df['index'] = date_list
            df['Open'] = pd.to_numeric(df['Open'])
            df['Close'] = pd.to_numeric(df['Close'])
            df['High'] = pd.to_numeric(df['High'])
            df['Low'] = pd.to_numeric(df['Low'])
            df['Change'] = df['Close'].diff()
            df['Volume'] = pd.to_numeric(df['Volume'])
            df.to_sql(name, con, if_exists="replace")

        con.close()

        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()


    def data_read(self):
        startdate = "20190101"  # 데이터 다운로드 시작 기준일
        kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드로드
        codes = code_name['code']  # 코드네임을 리스트로 저장
        names = code_name['name']  # 코드네임을 리스트로 저장
        # codes = ['A058860']  # 코드네임을 리스트로 저장
        # names = ['KTis']  # 코드네임을 리스트로 저장

        file_name = kospi_kosdaq + "data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        for i, code in enumerate(codes):  # 뺑뺑이 돌리기
            date_list = []
            name = names[i]
            print(name + " 진입")
            code = code[1:]
            # data = pdr.get_data_yahoo(code + '.KS', start=startdate)
            data = pdr.naver.NaverDailyReader(code, start=startdate).read()

            date_index = data.index
            for date in date_index:
                date = str(date)
                date = date[:4] + date[5:7] + date[8:10]
                date = int(date)
                date_list.append(date)

            df = pd.DataFrame(data)
            df['index'] = date_list
            df['Open'] = pd.to_numeric(df['Open'])
            df['Close'] = pd.to_numeric(df['Close'])
            df['High'] = pd.to_numeric(df['High'])
            df['Low'] = pd.to_numeric(df['Low'])
            df['Change'] = df['Close'].diff()
            df['change_ratio'] = df['Close'].pct_change()  # 변화량을 퍼센테이지로 구한다.
            df['change_ratio'] = df['change_ratio'].round(5)
            df['Volume'] = pd.to_numeric(df['Volume'])
            df.to_sql(name, con, if_exists="replace")
            print("완료2")
            print(i + 1, "/", len(names), name, " 완료")
        con.close()

        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()

    def data_read2(self, kospi_kosdaq='quant'):       # 입력값을 받아 일봉 다운로드
        startdate = '20190101'
        # kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load2(kospi_kosdaq)  # 종목 코드로드
        codes = code_name['code']  # 코드네임을 리스트로 저장
        names = code_name['name']  # 코드네임을 리스트로 저장
        # codes = ['A058860']  # 코드네임을 리스트로 저장
        # names = ['KTis']  # 코드네임을 리스트로 저장

        file_name = kospi_kosdaq + "data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        for i, code in enumerate(codes):  # 뺑뺑이 돌리기
            date_list = []
            name = names[i]
            print(name + " 진입")
            code = code[1:]
            # data = pdr.get_data_yahoo(code + '.KS', start='20190101')
            data = pdr.naver.NaverDailyReader(code, start=startdate).read()

            date_index = data.index
            for date in date_index:
                date = str(date)
                date = date[:4] + date[5:7] + date[8:10]
                date = int(date)
                date_list.append(date)

            df = pd.DataFrame(data)
            df['index'] = date_list
            df['Open'] = pd.to_numeric(df['Open'])
            df['Close'] = pd.to_numeric(df['Close'])
            df['High'] = pd.to_numeric(df['High'])
            df['Low'] = pd.to_numeric(df['Low'])
            df['Change'] = df['Close'].diff()
            df['change_ratio'] = df['Close'].pct_change()  # 변화량을 퍼센테이지로 구한다.
            df['change_ratio'] = df['change_ratio'].round(5)
            df['Volume'] = pd.to_numeric(df['Volume'])
            df.to_sql(name, con, if_exists="replace")
            print("완료2")
            print(i + 1, "/", len(names), name, " 완료")
        con.close()

        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()


    def simul_data_to_excel(self, data, save_name="data_result"):  # contine : 이어쓰기 여부 1이면 이어쓰기임   @@@@@@@@@@@@ 수정필요함
        # num = 1
        # num = str(num)
        # path = "c:/users/백/save_excel" + num + ".xlsx"

        path = "c:/users/백/save_excel.xlsx"
        # wb = Workbook()
        wb = load_workbook(path, data_only=True)
        worksheet_1 = wb.active

        # Dataframe을 엑셀로 뿌린다.
        maxrow = worksheet_1.max_row
        print("최대행은 %s행 입니다." % maxrow)
        if maxrow == 1:
            header = True
        else:
            header = False

        for i, row in enumerate(dataframe_to_rows(data, index=True, header=header)):  #한줄씩 엑셀로
            if len(row) > 1:
                worksheet_1.append(row)
        save_name = str(save_name)
        save_path = "c:/users/백/" + save_name + ".xlsx"
        wb.save(save_path)
        print("saved to file name %s.xlsx!!" % save_name)
        self.sound_play()

    def codeNname_load2(self, kospi_kosdaq='etf_real'):      # 인자를 받아서 코드 불러오기

        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")
        sheet = file.get_sheet_by_name(kospi_kosdaq)  #인자 받는 부분 나중에 수정할 것, 앞으로 지원하지 않는 기능임
        code_list = []
        name_list = []
        index_list = []

        for r in sheet.rows:
            index_list.append(r[0].value)
            code_list.append(r[1].value)
            name_list.append(r[2].value)

        codeNname = {"index": index_list, "code": code_list, "name": name_list}

        df = DataFrame(codeNname)
        return df


    def codeNname_load(self, category='Default'):      # GUI에 적혀있는 코드 불러오기

        if category == 'Default':
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            kospi_kosdaq = category

        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")

        sheet = file.get_sheet_by_name(kospi_kosdaq)  #인자 받는 부분 나중에 수정할 것, 앞으로 지원하지 않는 기능임
        code_list = []
        name_list = []
        index_list = []

        for r in sheet.rows:
            index_list.append(r[0].value)
            code_list.append(r[1].value)
            name_list.append(r[2].value)

        codeNname = {"index": index_list, "code": code_list, "name": name_list}

        df = DataFrame(codeNname)
        return df


    def tech__(self, name, kospi_kosdaq='etf_real'):    #한종목의 기술적 지표 계산
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            pass
        # con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        # con = sqlite3.connect("c:/users/백/%s_data_month_index.db" % self.kospi_kosdaq)

        file_name = kospi_kosdaq + "data_download.db"
        con = sqlite3.connect("c:/users/백/" + file_name)

        try:
            df = pd.read_sql("SELECT * FROM "+ "'"+ name +"' ", con, index_col='index') #df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        except:
            return

        # df.sort_index(inplace=True, ascending=True)  # 인덱스 기준으로 역으로 정렬

        # if len(df['open']) > 0:  # 값이 없는것을 제외시킨다
        #
        #     op = df['open'] * 0.1 * 10
        #     cl = df['close'] * 0.1 * 10
        #     hi = df['high'] * 0.1 * 10
        #     lo = df['low'] * 0.1 * 10
        #     vo = df['volume'] * 0.1 * 10

        if len(df['Open']) > 0:  # 값이 없는것을 제외시킨다

            # op = pd.to_numeric(df['Open'])
            # cl = pd.to_numeric(df['Close'])
            # hi = pd.to_numeric(df['High'])
            # lo = pd.to_numeric(df['Low'])
            # vo = pd.to_numeric(df['Volume'])

            op = df['Open']
            cl = df['Close']
            hi = df['High']
            lo = df['Low']
            vo = df['Volume']

            dfma3 = ta.SMA(cl, 3)
            dfma5 = ta.SMA(cl, 5)
            dfma10 = ta.SMA(cl, 10)
            dfma20 = ta.SMA(cl, 20)            
            dfma60 = ta.SMA(cl, 60)
            # dfma120 = ta.SMA(cl, 120)
            # # dfma480 = ta.SMA(cl, 480)
            # df['atr20(N)'] = ta.ATR(hi, lo, cl, timeperiod = 20)
            # df['Trange'] = ta.TRANGE(hi, lo, cl)
            tr = ta.TRANGE(hi, lo, cl)
            df['atr20(N)'] = ta.SMA(tr, 20)

            df['vma20_R'] = vo / ta.SMA(vo, 20)
            df['dp_ma3'] = cl / dfma3
            df['dp_ma5'] = cl / dfma5
            df['dp_ma10'] = cl / dfma10
            df['dp_ma20'] = cl / dfma20
            df['dp_ma60'] = cl / dfma60
            # df['dp_ma120'] = cl / dfma120
            # # df['dp_ma480'] = cl / dfma480

            bbands20 = pd.Series(ta.BBANDS(cl, timeperiod=20, nbdevup=2, nbdevdn=2))  # 밴드는 또 Series를 만들었다가
            # df['bbands20_upR'] = cl/bbands20[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # df['bbands30_mov'] = bbands30[1]
            # df['bbands20_downR'] = cl/bbands20[2]
            df['bbands_width'] = (bbands20[0] - bbands20[2]) / cl


            # bbands30 = pd.Series(ta.BBANDS(cl, timeperiod=30, nbdevup=1.8, nbdevdn=1.8))  # 밴드는 또 Series를 만들었다가
            # df['bbands30_upR'] = cl/bbands30[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # #df['bbands30_mov'] = bbands30[1]
            # df['bbands30_downR'] = cl/bbands30[2]
            # df['bbands_width'] = (bbands30[0] - bbands30[2]) / cl
            # print("완료")
            #
            # macd = pd.Series(ta.MACD(cl, 5, 34, 7))  # 밴드는 또 Series를 만들었다가
            # #df['macd_line5,34,7'] = macd[0]  # 하나씩 일일이 인덱싱해줘야 함.
            # #df['macd_sig5,34,7'] = macd[1]
            # #df['macd_histo5,34,7'] = macd[2]
            #
            # #df['macd_line5,34,7'] = macd[0] / dfma5  # 5일 이평기준 어느정도에 있는가?
            # #df['macd_sig5,34,7'] = macd[1] / dfma5
            # #df['macd_histo5,34,7'] = macd[2] / dfma5  #히스토그램 구하는 공식은 단기이평-장기 이평임
            #
            # #df['rsi5'] = ta.RSI(cl, 5)
            # #df['rsi10'] = ta.RSI(cl, 10)
            # #df['rsi20'] = ta.RSI(cl, 20)

        return df


    def add_tech(self, kospi_kosdaq='etf_real'):         # 전체 종목 대상 자동 기술적지표 구하기
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            print(kospi_kosdaq)
        print(kospi_kosdaq)
        file_name = kospi_kosdaq + "_data_add_auto.db"
        con2 = sqlite3.connect("c:/users/백/" + file_name)  # 키움증권 다운로드 종목 데이터 베이스

        code_name = self.codeNname_load(kospi_kosdaq)   # 종목 코드로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.tech__(name, kospi_kosdaq)
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("최종완료, %s 로 저장 완료" % file_name)
        self.sound_play()

    def tech_manual(self, name, kospi_kosdaq='etf_real'):
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            pass

        con = sqlite3.connect("c:/users/백/" + kospi_kosdaq + "_data_add_auto.db")
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col='index')

        if len(df['Open']) > 0:  # 값이 없는것을 제외시킨다
            op = list(df['Open'])
            cl = list(df['Close'])
            hi = list(df['High'])
            lo = list(df['Low'])
            vo = list(df['Volume'])
            bong_list = (df['Close'] - df['Open']) / df['Open']  # 봉크기
            N20 = list(df['atr20(N)'])

            # dp_ma20 = df['dp_ma20'] * 0.1 * 10
            # bband__ = df['bbands30_upR']
            open_R_list = []
            high_R_list = []
            low_R_list = []

            gap_list = []  # 갭 비율
            N20_ratio_list = []   # N 의 비율(20일기준)

            max_10_R_list = []  # 10일 최고가대비 현재가 비율
            min_10_R_list = []  # 10일 최저가대비 현재가 비율

            max_20_R_list = []  # 20일 최고가대비 현재가 비율
            min_20_R_list = []  # 20일 최저가대비 현재가 비율

            max_55_R_list = []  # 60일 최고가대비 현재가 비율
            min_55_R_list = []  # 60일 최저가대비 현재가 비율

            max_60_R_list = []  # 60일 최고가대비 현재가 비율
            min_60_R_list = []  # 60일 최저가대비 현재가 비율

            # max_240_R_list = []  # 240일 최고가대비 현재가 비율
            # min_240_R_list = []  # 240일 최저가대비 현재가 비율
            # ma20_grd_5_list = [] #20일 이평선의 기울기(5일기준)
            future_120_max_list = []
            future_120_min_list = []


            u_tail_R_list = []  # 윗꼬리 비율
            l_tail_R_list = []   # 아랫꼬리 비율

            vol_d1_R = []  # 전일대비 거래량 비율 구하기
            D1_open_list = []   # 익일 시가
            D1_open_ratio_list = [] # 전일 종가대비 익일 시가 비율
            D1_high_list = []  # 익일 고가
            D1_high_ratio_list = []  # 전일 종가대비 익일 고가 비율
            D1_low_list = []   # 익일 저가
            D1_low_ratio_list = []  # 전일 종가대비 익일 저가 비율
            D1_close_list = []  # 익일 종가
            D1_close_ratio_list = []   # 전일 종가대비 익일 종가 비율
            # D1_profit_list = []  # 익일 종가 - 당일종가 비율
            # D1_bband_list = [] # 익일 볼린저밴드

            for i in range(len(op)):
                if i == 0:  # 첫번째 데이터는 전일 정보가 없으므로 0추가

                    vol_d1_R.append(0)    # 전일대비 거래량 비율 구하기
                    gap_list.append(0)

                    N20_ratio_list.append(0)

                    max_10_R_list.append(0)
                    min_10_R_list.append(0)

                    max_20_R_list.append(0)
                    min_20_R_list.append(0)

                    max_55_R_list.append(0)
                    min_55_R_list.append(0)

                    max_60_R_list.append(0)
                    min_60_R_list.append(0)
                    # max_240_R_list.append(0)
                    # min_240_R_list.append(0)
                    open_R_list.append(0)
                    high_R_list.append(0)
                    low_R_list.append(0)

                else:
                    try:
                        vol_d1_Ratio = vo[i] / vo[i - 1]  # i=오늘, i-1 = 어제 i가 커질수록 과거에서 미래로 가는것임. 전일대비 거래량
                        vol_d1_R.append(vol_d1_Ratio)
                    except ZeroDivisionError:
                        vol_d1_R.append(0)

                    cal10 = i - 10
                    cal20 = i - 20
                    cal55 = i - 55
                    cal60 = i - 60
                    # cal240 = i - 240
                    d1_cl = cl[i - 1]
                    open_R = (op[i] - d1_cl) / d1_cl
                    high_R = (hi[i] - d1_cl) / d1_cl
                    low_R = (lo[i] - d1_cl) / d1_cl

                    open_R_list.append(open_R)
                    high_R_list.append(high_R)
                    low_R_list.append(low_R)

                    gap = round((cl[i - 1] - op[i]) / cl[i - 1], 4)
                    gap_list.append(gap)

                    if cal10 < 0:
                        max_10_R_list.append(0)
                        min_10_R_list.append(0)
                    else:
                        max_10 = max(cl[(cal10):i + 1])
                        max_10_R = (cl[i] - max_10) / max_10  # 고점대비 얼마나 떨어졌나
                        max_10_R_list.append(max_10_R)
                        min_10 = min(cl[(cal10):i + 1])
                        min_10_R = (cl[i] - min_10) / min_10  # 저점대비 얼마나 올랐나
                        min_10_R_list.append(min_10_R)

                    if cal20 < 0:
                        max_20_R_list.append(0)
                        min_20_R_list.append(0)

                        N20_ratio_list.append(0)
                    else:
                        max_20 = max(cl[(cal20):i + 1])
                        max_20_R = (cl[i] - max_20) / max_20  # 고점대비 얼마나 떨어졌나
                        max_20_R_list.append(max_20_R)
                        min_20 = min(cl[(cal20):i + 1])
                        min_20_R = (cl[i] - min_20) / min_20  # 저점대비 얼마나 올랐나
                        min_20_R_list.append(min_20_R)

                        N20_ratio = N20[i] / cl[i]
                        N20_ratio_list.append(N20_ratio)
                        
                    if cal55 < 0:
                        max_55_R_list.append(0)
                        min_55_R_list.append(0)
                    else:
                        max_55 = max(cl[(cal55):i + 1])
                        max_55_R = (cl[i] - max_55) / max_55  # 고점대비 얼마나 떨어졌나
                        max_55_R_list.append(max_55_R)
                        min_55 = min(cl[(cal55):i + 1])
                        min_55_R = (cl[i] - min_55) / min_55  # 저점대비 얼마나 올랐나
                        min_55_R_list.append(min_55_R)

                    if cal60 < 0:
                        max_60_R_list.append(0)
                        min_60_R_list.append(0)
                    else:
                        max_60 = max(cl[(cal60):i + 1])
                        max_60_R = (cl[i] - max_60) / max_60  # 고점대비 얼마나 떨어졌나
                        max_60_R_list.append(max_60_R)
                        min_60 = min(cl[(cal60):i + 1])
                        min_60_R = (cl[i] - min_60) / min_60  # 저점대비 얼마나 올랐나
                        min_60_R_list.append(min_60_R)

                    # if cal240 < 0:
                    #     max_240_R_list.append(0)
                    #     min_240_R_list.append(0)
                    # else:
                    #     max_240 = max(cl[(cal240):i+1])
                    #     max_240_R = (cl[i] - max_240) / max_240  # 고점대비 얼마나 떨어졌나
                    #     max_240_R_list.append(max_240_R)
                    #     min_240 = min(cl[(cal240):i+1])
                    #     min_240_R = (cl[i] - min_240) / min_240  # 저점대비 얼마나 올랐나
                    #     min_240_R_list.append(min_240_R)


            for i in range(len(op)):
                try:
                    u_tail_R = (hi[i] - max(op[i], cl[i]))/max(op[i], cl[i])

                except ZeroDivisionError:
                    u_tail_R = 0

                try:
                    l_tail_R = (min(op[i], cl[i]) - lo[i]) / min(op[i], cl[i])
                except ZeroDivisionError:
                    l_tail_R = 0
                u_tail_R_list.append(u_tail_R)
                l_tail_R_list.append(l_tail_R)

                if i == len(op) - 1:  # 가장 최근 데이터는 익일 정보가 없으므로 0추가
                #
                    D1_open_list.append(0)
                    D1_high_list.append(0)
                    D1_low_list.append(0)
                    D1_close_list.append(0)
                    D1_open_ratio_list.append(0)
                    D1_high_ratio_list.append(0)
                    D1_low_ratio_list.append(0)
                    D1_close_ratio_list.append(0)
                #     index_A_list.append(i)    # 인덱스 추가
                #     future_120_max = max(cl[i:i+120])
                #     future_120_max_list.append(future_120_max)
                #     future_120_min = min(cl[i:i + 120])
                #     future_120_min_list.append(future_120_min)
                #     # D1_bband_list.append(0)
                #
                elif cl[i] == 0:  # 0으로 나누는 경우 error방지

                    D1_open_ratio_list.append(0)
                    D1_high_ratio_list.append(0)
                    D1_low_ratio_list.append(0)
                    D1_close_ratio_list.append(0)
                #     index_A_list.append(i)   # 인덱스 추가
                #     future_120_max = max(cl[i:i + 120])
                #     future_120_max_list.append(future_120_max)
                #     future_120_min = min(cl[i:i + 120])
                #     future_120_min_list.append(future_120_min)
                #
                else:
                #
                    D1_open_list.append(op[i + 1])
                    D1_high_list.append(hi[i + 1])
                    D1_low_list.append(lo[i + 1])
                    D1_close_list.append(cl[i + 1])
                #     future_120_max = max(cl[i:i + 120])
                #     future_120_max_ratio = (future_120_max - cl[i])/cl[i]   # 현재가 기준 미래 120일 최고가 기준 상승률
                #     future_120_max_list.append(future_120_max_ratio)
                #
                #     future_120_min = min(cl[i:i + 120])
                #     future_120_min_ratio = (future_120_min - cl[i]) / cl[i]  # 현재가 기준 미래 120일 최저가 하락률
                #     future_120_min_list.append(future_120_min_ratio)

                    D1_open_ratio = (op[i + 1] - cl[i]) / cl[i]
                    D1_open_ratio = round(D1_open_ratio, 5)
                    D1_high_ratio = (hi[i + 1] - cl[i]) / cl[i]
                    D1_high_ratio = round(D1_high_ratio, 5)
                    D1_low_ratio = (lo[i + 1] - cl[i]) / cl[i]
                    D1_low_ratio = round(D1_low_ratio, 5)
                    D1_close_ratio = (cl[i + 1] - cl[i]) / cl[i]
                    D1_close_ratio = round(D1_close_ratio, 5)
                    #
                    D1_open_ratio_list.append(D1_open_ratio)
                    D1_high_ratio_list.append(D1_high_ratio)
                    D1_low_ratio_list.append(D1_low_ratio)
                    D1_close_ratio_list.append(D1_close_ratio)
                    # index_A_list.append(i)    # 인덱스 추가

                    # D1_bband_list.append(bband__[i+1])

            df['bong_ratio'] = bong_list
            df['u_tail_R'] = u_tail_R_list
            df['l_tail_R'] = l_tail_R_list
            df['vol_d1_R'] = vol_d1_R
            # df['D1_open'] = D1_open_list
            # df['D1_high'] = D1_high_list
            # df['D1_low'] = D1_low_list

            df['D1_open_R'] = D1_open_ratio_list  # 종가거래 하루 수익
            df['D1_high_R'] = D1_high_ratio_list  # 종가거래 하루 수익
            df['D1_low_R'] = D1_low_ratio_list  # 종가거래 하루 수익
            df['D1_close'] = D1_close_list
            df['D1_close_R'] = D1_close_ratio_list  # 종가거래 하루 수익
            # df['index_A'] = index_A_list   # 인덱스 추가
            df['gap'] = gap_list   # 갭 : 전일종가 - 금일 시가
            df['N20_R'] = N20_ratio_list

            df['max_10_R'] = max_10_R_list
            df['min_10_R'] = min_10_R_list

            df['max_20_R'] = max_20_R_list
            df['min_20_R'] = min_20_R_list

            df['max_55_R'] = max_55_R_list
            df['min_55_R'] = min_55_R_list
            
            df['max_60_R'] = max_60_R_list
            df['min_60_R'] = min_60_R_list

            # df['max_240_R'] = max_240_R_list
            # df['min_240_R'] = min_240_R_list

            # df['mirae_120_max_R'] = future_120_max_list
            # df['mirae_120_min_R'] = future_120_min_list
            # df['D1_bbands30_upR'] = D1_bband_list

        return df

    def add_tech_manual(self, kospi_kosdaq='etf_real'):
        if kospi_kosdaq is False:
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            print(kospi_kosdaq)
        code_name = self.codeNname_load(kospi_kosdaq)  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ['셀트리온제약', '천보']
        # name_list = ['한일철강']

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.tech_manual(name, kospi_kosdaq)
            df['name'] = name   # 종목명 삽입
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        print("tech_manual 최종완료")
        self.sound_play()



    def condition_filter_soup(self):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 터틀수프 플러스원 전략
        ##### https://stock79.tistory.com/entry/%EC%8B%A4%EC%A0%84-%ED%88%AC%EC%9E%90-%EC%A0%84%EB%9E%B5-83-Street-smart-4-Turtle-soup-plus-one-%EC%A0%84%EB%9E%B5?category=457287

        self.sound_play()
        max_hold_day = 2        # 보유기간
        stopgain = 1.10          # 익절값 1.20
        stoploss_R = 0.05       # 손절값
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가
        stopgain_price = 0

        save_file_name = "tutle_soup_" + "max_hold_day" + str(max_hold_day) + "_stopgain" + str(stopgain)

        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % self.kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % self.kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["셀트리온제약","천보"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['open'])
            high_list = list(df['high'])
            low_list = list(df['low'])          # 저가
            close_list = list(df['close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])
            min_close_day3_list = []
            d1_20_low_list = []
            d2_20_low_list = []
            d3_20_low_list = []
            d4_20_low_list = []

            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            # max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])
            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            # buy_type_list = []
            setup_condition_list = []

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                # max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                low = low_list[j]
                high = high_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                hold_unit = 0
                stoploss_price = 0
                # buy_type = 0
                setup_condition = False

                if j < 4:
                    min_close_day3 = 0
                    d1_20_low = 0           # 4일째부터 계산 그전은 신경안쓴다.
                    d2_20_low = 0
                    d3_20_low = 0
                    d4_20_low = 0

                    d1_20_low_list.append(d1_20_low)
                    d2_20_low_list.append(d2_20_low)
                    d3_20_low_list.append(d3_20_low)
                    d4_20_low_list.append(d4_20_low)

                else:
                    min_close_day3 = min(close_list[j-4:j-1])                      # 전전일기준 3일 종가 중 최소값
                    d1_20_low = min_20_R_list[j - 1]
                    d2_20_low = min_20_R_list[j - 2]
                    d3_20_low = min_20_R_list[j - 3]
                    d4_20_low = min_20_R_list[j - 4]

                    d1_20_low_list.append(d1_20_low)
                    d2_20_low_list.append(d2_20_low)
                    d3_20_low_list.append(d3_20_low)
                    d4_20_low_list.append(d4_20_low)

                if j == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    # buy_type_list.append(buy_type)

                    min_close_day3_list.append(min_close_day3)
                    setup_condition_list.append(False)

                elif max_20_R == 0 and min_10_R == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)
                    min_close_day3_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    # buy_type_list.append(buy_type)

                    setup_condition_list.append(False)

                else:
                    hold_unit = hold_unit_list[j-1]
                    min_close_day3_list.append(min_close_day3)
                    if hold_unit == 1:
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                        # buy_type = buy_type_list[j-1]
                    else:
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j-1]

                    setup_condition = False

                    if (d2_20_low * d3_20_low * d4_20_low) == 0 and j > 25:

                        setup_condition = True
                        # 셋업 조건 : 2일전부터 4일전까지 한번 이상 20신저가 발생

                    if hold_unit == 0 and setup_condition is True and d1_20_low == 0 and close > min_close_day3:
                        # 신규매입 조건 : 전일 20일 신저가 발생, 당일 종가가 전전일부터 3일간 저가 돌파시 매수

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if low < d1_low:
                            stoploss_price = low
                        else:
                            stoploss_price = d1_low
                        stoploss_price_list.append(stoploss_price)

                        stopgain_price = buy_price * stopgain

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        # buy_type = "S1"
                        # buy_type_list.append(buy_type)
                        setup_condition_list.append(setup_condition)

                    elif hold_unit == 1 and low < stoploss_price:      # 손절

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)

                        if open_list[j] < stoploss_price:
                            sell_price = open_list[j]
                        else:
                            sell_price = stoploss_price
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        # buy_type = buy_type_list[j-1]
                        # buy_type_list.append(buy_type)

                        setup_condition_list.append(setup_condition)

                    elif hold_unit > 0 and (hold_day == max_hold_day or high > stopgain_price):   # 청산 조건 :

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        if open_list[j] > stopgain_price:
                            sell_price = open_list[j]
                        elif high > stopgain_price:
                            sell_price = stopgain_price
                        else:
                            sell_price = close                              # 종가기준 청산

                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        # buy_type = buy_type_list[j - 1]
                        # buy_type_list.append(buy_type)

                        setup_condition_list.append(setup_condition)

                    elif hold_unit == 1:    # 홀드

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        # buy_type = buy_type_list[j - 1]
                        # buy_type_list.append(buy_type)

                        setup_condition_list.append(setup_condition)

                    else:                               # 그외는 관망

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        # buy_type = 0
                        # buy_type_list.append(buy_type)

                        setup_condition_list.append(setup_condition)

            df_result['date'] = df['date']
            df_result['name'] = df['name']
            df_result['open'] = df['open']
            df_result['high'] = df['high']
            df_result['low'] = df['low']
            df_result['close'] = df['close']
            df_result['change'] = df['change']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = ma60_list
            # df_result['open_ratio'] = open_ratio_list
            # df_result['volume'] = df['volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            # df_result['max_20R'] = df['max_20_R']
            # df_result['min_10R'] = df['min_10_R']
            # df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            df_result['d1_d20_low'] = d1_20_low_list
            df_result['d2_d20_low'] = d2_20_low_list
            df_result['d3_d20_low'] = d3_20_low_list
            df_result['d4_d20_low'] = d4_20_low_list
            df_result['min_cl_3day'] = min_close_day3_list

            df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']

            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list

            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['winNlose'] = winnloss_list
            df_result['set_up'] = setup_condition_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()           # 터틀수프 플러스원 전략 끝.
        self.execute_query_for_all_stock(save_file_name)


    def condition_filter_low_buy(self):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 터틀트레이딩--익일에 조금 싸게 사기

        #####
        self.sound_play()
        # max_hold_day = 1        # 보유기간
        stopgain = 1.5          # 익절값
        constant_stoploss = 0.1       # 절대 손절값, 2N과 비교하여 더 작은값을 손절값으로
        buy_conditon = -0.005       # 매수 조건(전일 종가 대비 하락비율)
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가

        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % self.kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % self.kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["셀트리온제약","천보"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['open'])
            high_list = list(df['high'])
            low_list = list(df['low'])          # 저가
            close_list = list(df['close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])

            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            set_up_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            buy_type_list = []
            set_up = 0

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                low = low_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                hold_unit = 0
                stoploss_price = 0
                buy_type = 0


                if j == 0:                     # 첫번째 값은 무조건 0
                    set_up = 0
                    set_up_list.append(set_up)
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                elif max_20_R == 0 and min_10_R == 0:           # 20일 이전값만 취급
                    set_up = 0
                    set_up_list.append(set_up)
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                else:
                    hold_unit = hold_unit_list[j-1]
                    if hold_unit == 1:

                        stopgain_price = buy_price_list[j-1] * stopgain         # 익절값 설정
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                        buy_type = buy_type_list[j-1]
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j-1]

                    # try:
                    #     d1_u_tail_ratio = (d1_open - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 윗꼬리 가격상 비율
                    #     d1_l_tail_ratio = (d1_close - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 아랫꼬리 가격상 비율
                    # except ZeroDivisionError:
                    #     d1_u_tail_ratio = 0
                    #     d1_l_tail_ratio = 0

                    # if (open_list[j] < d1_low and close > open_list[j] and close > (low + high) / 2 and 0.1 > change > 0.04
                    #         and hold_unit == 0):

                    if max_20_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 0 and hold_unit == 0:
                        # S1 셋업 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
                    # if max_20_R == 0 and N != None and N_ratio > 0.01 and winnloss == 0 and hold_unit == 0:
                    #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거

                        set_up = "s1"
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "setup_s1"
                        condition_fliter_list.append(condition)
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)
                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        open_ratio_list.append(0)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type_list.append(buy_type)

                    elif set_up == "s1" and hold_unit == 0 and low < (close_list[j-1] * (1 + buy_conditon)):       # 신규진입S1 : 셋업 조건 만족하고 전일종가보다 최저값이 설정한 값보나 낮아야함

                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s1"
                        condition_fliter_list.append(condition)
                        buy_price = close_list[j-1] * (1 + buy_conditon)
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)


                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S1"
                        buy_type_list.append(buy_type)


                    elif max_55_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 1 and hold_unit == 0:
                    # S2 매수조건 진입, 60이평 위 조건 추가
                    # elif max_55_R == 0 and N != None and N_ratio > 0.01 and winnloss == 1 and hold_unit == 0:
                    #     # S2 매수조건 진입, 60이평 위 조건 제거

                        set_up = "s2"
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "setup_s2"
                        condition_fliter_list.append(condition)
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)
                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        open_ratio_list.append(0)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type_list.append(buy_type)


                    elif set_up == "s2" and hold_unit == 0 and low < (close_list[j - 1] * (1 + buy_conditon)):  # 신규진입 : 셋업 조건 만족하고 전일종가보다 최저값이 설정한 값보나 낮아야함

                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s2"
                        condition_fliter_list.append(condition)
                        buy_price = close_list[j - 1] * (1 + buy_conditon)
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)  # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S2"
                        buy_type_list.append(buy_type)


                    # elif hold_unit == 1 and low < stoploss_price or open_list[j] < stoploss_price:      # 손절(즉시)
                    elif hold_unit == 1 and close < stoploss_price:                                       # 손절(종가기준)
                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)

                        # if open_list[j] < stoploss_price:       # 즉시 손절
                        #     sell_price = open_list[j]
                        # else:
                        #     sell_price = stoploss_price
                        # sell_price_list.append(sell_price)

                        sell_price = close                      # 종가 기준 손절
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j-1]
                        buy_type_list.append(buy_type)

                    elif min_10_R == 0 and buy_type == "S1" and hold_unit == 1:   # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_20_R == 0 and buy_type == "S2" and hold_unit == 1:   # S2 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1:    # 홀드
                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    else:                               # 그외는 관망
                        set_up = 0
                        set_up_list.append(set_up)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        buy_type = 0
                        buy_type_list.append(buy_type)

            df_result['date'] = df['date']
            df_result['name'] = df['name']
            df_result['open'] = df['open']
            df_result['high'] = df['high']
            df_result['low'] = df['low']
            df_result['close'] = df['close']
            df_result['change'] = df['change']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = ma60_list
            # df_result['open_ratio'] = open_ratio_list
            df_result['volume'] = df['volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            df_result['max_20R'] = df['max_20_R']
            df_result['min_10R'] = df['min_10_R']
            df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']
            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list
            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['setup'] = set_up_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['buy_type'] = buy_type_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()           #터틀 트레이딩 전략 끝

    def condition_filter(self, category="Default"):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 터틀트레이딩

        #####
        self.sound_play()
        # max_hold_day = 1        # 보유기간
        stopgain = 1.5          # 익절값
        constant_stoploss = 0.1       # 절대 손절값, 2N과 비교하여 더 작은값을 손절값으로
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가
        print(category)
        if category is False:       # 나중에 왜 이렇게 되는지 확인할 것
            kospi_kosdaq = self.comboBox_2.currentText()
        else:
            kospi_kosdaq = category
        print(kospi_kosdaq)
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["KODEX 200", "KODEX 200TR"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['Open'])
            high_list = list(df['High'])
            low_list = list(df['Low'])          # 저가
            close_list = list(df['Close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])
            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            buy_type_list = []

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                low = low_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                hold_unit = 0
                stoploss_price = 0
                buy_type = 0

                if j == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                elif max_20_R == 0 and min_10_R == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                else:

                    hold_unit = hold_unit_list[j-1]
                    if hold_unit == 1:

                        stopgain_price = buy_price_list[j-1] * stopgain         # 익절값 설정
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                        buy_type = buy_type_list[j-1]
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j-1]

                    # try:
                    #     d1_u_tail_ratio = (d1_open - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 윗꼬리 가격상 비율
                    #     d1_l_tail_ratio = (d1_close - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 아랫꼬리 가격상 비율
                    # except ZeroDivisionError:
                    #     d1_u_tail_ratio = 0
                    #     d1_l_tail_ratio = 0

                    # if (open_list[j] < d1_low and close > open_list[j] and close > (low + high) / 2 and 0.1 > change > 0.04
                    #         and hold_unit == 0):

                    if max_20_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 0 and hold_unit == 0:
                        # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
                    # if max_20_R == 0 and N != None and N_ratio > 0.01 and winnloss == 0 and hold_unit == 0:
                    #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s1"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        # stoploss_price = close - (2 * N)
                        # stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S1"
                        buy_type_list.append(buy_type)


                    elif max_55_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 1 and hold_unit == 0:
                    # S2 매수조건 진입, 60이평 위 조건 추가
                    # elif max_55_R == 0 and N != None and N_ratio > 0.01 and winnloss == 1 and hold_unit == 0:
                    #     # S2 매수조건 진입, 60이평 위 조건 제거

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s2"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        # stoploss_price = close - (2 * N)
                        # stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S2"
                        buy_type_list.append(buy_type)

                    # elif hold_unit == 1 and low < stoploss_price or open_list[j] < stoploss_price:      # 손절(즉시)
                    elif hold_unit == 1 and close < stoploss_price:                                       # 손절(종가기준)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)

                        # if open_list[j] < stoploss_price:       # 즉시 손절
                        #     sell_price = open_list[j]
                        # else:
                        #     sell_price = stoploss_price
                        # sell_price_list.append(sell_price)

                        sell_price = close                      # 종가 기준 손절
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j-1]
                        buy_type_list.append(buy_type)

                    elif min_10_R == 0 and buy_type == "S1" and hold_unit == 1:   # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_20_R == 0 and buy_type == "S2" and hold_unit == 1:   # S2 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1:    # 홀드

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    else:                               # 그외는 관망

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        buy_type = 0
                        buy_type_list.append(buy_type)

            df_result['date'] = df['Date']
            df_result['name'] = df['name']
            df_result['open'] = df['Open']

            df_result['high'] = df['High']
            df_result['low'] = df['Low']
            df_result['close'] = df['Close']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = df['dp_ma60']
            # df_result['open_ratio'] = open_ratio_list
            df_result['volume'] = df['Volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            df_result['max_20R'] = df['max_20_R']
            df_result['min_10R'] = df['min_10_R']
            df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            # df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']

            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list
            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['buy_type'] = buy_type_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()           #터틀 트레이딩 전략 끝

    def condition_filter_etf(self):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), etf 전용(60이평조건 제거)터틀트레이딩

        #####
        self.sound_play()

        stopgain = 1.5          # 익절값
        constant_stoploss = 0.1       # 절대 손절값, 2N과 비교하여 더 작은값을 손절값으로
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가

        kospi_kosdaq = 'etf_real'
        code_name = self.codeNname_load2(kospi_kosdaq)  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["KODEX 200", "KODEX 200TR"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['Open'])
            high_list = list(df['High'])
            low_list = list(df['Low'])          # 저가
            close_list = list(df['Close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])
            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            buy_type_list = []

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                low = low_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                hold_unit = 0
                stoploss_price = 0
                buy_type = 0

                if j == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                elif max_20_R == 0 and min_10_R == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                else:

                    hold_unit = hold_unit_list[j-1]
                    if hold_unit == 1:

                        stopgain_price = buy_price_list[j-1] * stopgain         # 익절값 설정
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                        buy_type = buy_type_list[j-1]
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j-1]

                    # if max_20_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 0 and hold_unit == 0:
                    #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
                    # if max_20_R == 0 and N != None and N_ratio > 0.02 and winnloss == 0 and hold_unit == 0:
                    #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거
                    if max_20_R == 0 and winnloss == 0 and hold_unit == 0:
                        # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0)

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s1"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)


                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S1"
                        buy_type_list.append(buy_type)


                    # elif max_55_R == 0 and N != None and N_ratio > 0.01 and ma60 != None and ma60 > 1 and winnloss == 1 and hold_unit == 0:
                    # # S2 매수조건 진입, 60이평 위 조건 추가
                    elif max_55_R == 0 and N != None and N_ratio > 0.02 and winnloss == 1 and hold_unit == 0:
                        # S2 매수조건 진입, 60이평 위 조건 제거

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s2"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S2"
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1 and close < stoploss_price:                                       # 손절(종가기준)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)


                        sell_price = close                      # 종가 기준 손절
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j-1]
                        buy_type_list.append(buy_type)

                    elif min_10_R == 0 and buy_type == "S1" and hold_unit == 1:   # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_20_R == 0 and buy_type == "S2" and hold_unit == 1:   # S2 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1:    # 홀드

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    else:                               # 그외는 관망

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        buy_type = 0
                        buy_type_list.append(buy_type)

            df_result['date'] = df['Date']
            df_result['name'] = df['name']
            df_result['open'] = df['Open']

            df_result['high'] = df['High']
            df_result['low'] = df['Low']
            df_result['close'] = df['Close']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = df['dp_ma60']
            # df_result['open_ratio'] = open_ratio_list
            df_result['volume'] = df['Volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            df_result['max_20R'] = df['max_20_R']
            df_result['min_10R'] = df['min_10_R']
            df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            # df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']

            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list
            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['buy_type'] = buy_type_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()           #터틀 트레이딩 전략 끝

    def condition_filter_quant(self):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 퀀트 전용 터틀트레이딩(60이평 조건, 10% 이하 상승조건 추가)

        #####
        self.sound_play()

        stopgain = 1.5          # 익절값
        constant_stoploss = 0.1       # 절대 손절값, 2N과 비교하여 더 작은값을 손절값으로
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가

        kospi_kosdaq = 'quant'
        code_name = self.codeNname_load2(kospi_kosdaq)  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용
        # name_list = ["오비고"]  # 실험용
        # name_list = ["KODEX 200", "KODEX 200TR"]  # 실험용

        for i, name in enumerate(name_list):
            print(name, "진입")

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['Open'])
            high_list = list(df['High'])
            low_list = list(df['Low'])          # 저가
            close_list = list(df['Close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []
            ma60_list = list(df['dp_ma60'])
            # u_tail_R_list = list(df['u_tail_R'])
            # l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            N_R_list = list(df['N20_R'])
            bband_width_R_list = list(df['bbands_width'])
            max_20_R_list = list(df['max_20_R'])
            min_10_R_list = list(df['min_10_R'])
            max_55_R_list = list(df['max_55_R'])
            min_20_R_list = list(df['min_20_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []
            winnloss_list = []
            buy_type_list = []

            for j in range(len(close_list)):

                max_20_R = max_20_R_list[j]
                min_10_R = min_10_R_list[j]
                max_55_R = max_55_R_list[j]
                min_20_R = min_20_R_list[j]
                ma60 = ma60_list[j]
                close = close_list[j]
                change = change_list[j]
                low = low_list[j]
                N = N_list[j]
                N_ratio = N_R_list[j]
                bband_width_R = bband_width_R_list[j]
                hold_unit = 0
                stoploss_price = 0
                buy_type = 0

                if j == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                elif max_20_R == 0 and min_10_R == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)
                    winnloss = 0
                    winnloss_list.append(winnloss)
                    buy_type_list.append(buy_type)

                else:

                    hold_unit = hold_unit_list[j-1]
                    if hold_unit == 1:

                        stopgain_price = buy_price_list[j-1] * stopgain         # 익절값 설정
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                        buy_type = buy_type_list[j-1]
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0
                        buy_type = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)
                    winnloss = winnloss_list[j-1]

                    if max_20_R == 0 and N != None and N_ratio > 0.03 and change < 0.1 and ma60 != None and ma60 > 1 and winnloss == 0 and hold_unit == 0:
                        # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 위
                    # if max_20_R == 0 and N != None and N_ratio > 0.02 and winnloss == 0 and hold_unit == 0:
                    #     # S1 신규매입 조건 : 20일 신고가, 직전거래에서 손해시(winnloss == 0), N은 최소 2%이상 --- S1 적용, 60이평 조건 제거

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s1"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)


                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S1"
                        buy_type_list.append(buy_type)


                    elif max_55_R == 0 and N != None and N_ratio > 0.03 and change < 0.1 and ma60 != None and ma60 > 1 and winnloss == 1 and hold_unit == 0:
                    # S2 매수조건 진입, 60이평 위 조건 추가
                    # elif max_55_R == 0 and N != None and N_ratio > 0.02 and winnloss == 1 and hold_unit == 0:
                    #     # S2 매수조건 진입, 60이평 위 조건 제거

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy_s2"
                        condition_fliter_list.append(condition)
                        buy_price = close
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        if close * (1 - constant_stoploss) < close - (2 * N):
                            stoploss_price = close * (1 - constant_stoploss)           # 손절값을 x%보다는 더작게
                        else:
                            stoploss_price = close - (2 * N)
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss_list.append(winnloss)
                        buy_type = "S2"
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1 and close < stoploss_price:                                       # 손절(종가기준)

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)


                        sell_price = close                      # 종가 기준 손절
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)
                        winnloss = 0
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j-1]
                        buy_type_list.append(buy_type)

                    elif min_10_R == 0 and buy_type == "S1" and hold_unit == 1:   # S1 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif min_20_R == 0 and buy_type == "S2" and hold_unit == 1:   # S2 청산 조건 : 종가기준 10일 최저가 하향 돌파

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)
                        sell_price = close                              # 종가기준 청산
                        condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                        if profit > 0:
                            winnloss = 1        # 이득
                        else:
                            winnloss = 0        # 손해
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    elif hold_unit == 1:    # 홀드

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        # event_list.append(1)          # 홀드는 이벤트에 포함/불포함 선택한다.
                        winnloss_list.append(winnloss)
                        buy_type = buy_type_list[j - 1]
                        buy_type_list.append(buy_type)

                    else:                               # 그외는 관망

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)
                        winnloss_list.append(winnloss)
                        buy_type = 0
                        buy_type_list.append(buy_type)

            df_result['date'] = df['Date']
            df_result['name'] = df['name']
            df_result['open'] = df['Open']

            df_result['high'] = df['High']
            df_result['low'] = df['Low']
            df_result['close'] = df['Close']
            df_result['change_ratio'] = df['change_ratio']
            df_result['dp_ma5'] = df['dp_ma5']
            df_result['dp_ma10'] = df['dp_ma10']
            df_result['dp_ma20'] = df['dp_ma20']
            df_result['dp_ma60'] = df['dp_ma60']
            # df_result['open_ratio'] = open_ratio_list
            df_result['volume'] = df['Volume']
            df_result['atr20(N)'] = df['atr20(N)']
            df_result['N20(%)'] = df['N20_R']
            df_result['bbands_width_R'] = bband_width_R_list
            df_result['max_20R'] = df['max_20_R']
            df_result['min_10R'] = df['min_10_R']
            df_result['max_55R'] = df['max_55_R']
            df_result['min_20R'] = df['min_20_R']
            # df_result['sichong'] = df['si_chong']
            # df_result['vol_d1r'] = df['vol_d1_R']

            # df_result['up_ggori'] = u_tail_R_list
            # df_result['btm_ggori'] = l_tail_R_list
            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list
            df_result['buy_type'] = buy_type_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.sound_play()           #터틀 트레이딩 전략 끝

    def condition_filter_8020(self):   # buy, hold, sell, stoploss   조건 필터열 삽입(전략을 시뮬레이션 한다), 8020전략
        # https://stock79.tistory.com/entry/%EC%8B%A4%EC%A0%84-%ED%88%AC%EC%9E%90-%EC%A0%84%EB%9E%B5-84-Street-smart-5-80-20-%EC%A0%84%EB%9E%B5?category=457287
        ####
        self.sound_play()
        max_hold_day = 5        # 보유기간
        stopgain = 1.20          # 익절값
        stoploss_R = 0.05       # 손절값
        condition_low = -0.03       # 매수조건 전일 하락값
        condition_open = -0.01     # 매수조건 당일 시초가

        save_file_name = "st_8020_" + "hold_day_" + str(max_hold_day) + "stopgain_" + str(stopgain) + "stoploss_" + str(stoploss_R)

        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con2 = sqlite3.connect("c:/users/백/%s_condition_filter.db" % self.kospi_kosdaq)  # 새로저장할 파일
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % self.kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "GS건설"]  # 실험용
        # name_list = ["셀트리온제약", "천보"]  # 실험용

        self.sound_play()
        for i, name in enumerate(name_list):

            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df_result = DataFrame()

            open_list = list(df['open'])
            high_list = list(df['high'])
            low_list = list(df['low'])          # 저가
            close_list = list(df['close'])  # 종가
            change_list = list(df['change_ratio'])
            open_ratio_list = []

            u_tail_R_list = list(df['u_tail_R'])
            l_tail_R_list = list(df['l_tail_R'])
            N_list = list(df['atr20(N)'])
            # max_20_R_list = list(df['max_20_R'])
            # min_10_R_list = list(df['min_10_R'])

            buy_price_list = []
            sell_price_list = []
            condition_fliter_list = []
            event_list = []            # 관망이면 0, 나머지는 전부 1

            stoploss_price_list = []
            stoploss_list = []
            hold_unit_list = []
            hold_day_list = []    # 보유일수
            profit_list = []
            profit_ratio_list = []

            for j in range(len(close_list)):

                # max_20_R = max_20_R_list[j]
                # min_10_R = min_10_R_list[j]
                close = close_list[j]
                low = low_list[j]
                high = high_list[j]
                N = N_list[j]
                hold_unit = 0
                stoploss_price = 0

                if j == 0:
                    hold_unit_list.append(hold_unit)
                    hold_day = 0
                    hold_day_list.append(hold_day)
                    buy_price_list.append(0)
                    sell_price_list.append(0)
                    condition = "non"
                    condition_fliter_list.append(condition)
                    stoploss_price_list.append(stoploss_price)
                    stoploss_list.append("x")
                    profit = 0
                    profit_list.append(profit)
                    profit_ratio = 0
                    profit_ratio_list.append(profit_ratio)
                    event_list.append(0)

                    open_ratio_list.append(0)

                else:
                    hold_unit = hold_unit_list[j-1]
                    if hold_unit == 1:

                        stopgain_price = buy_price_list[j-1] * stopgain         # 익절값 설정
                        stoploss_price = stoploss_price_list[j-1]
                        hold_day = hold_day_list[j-1] + 1
                    else:
                        stopgain_price = 0
                        stoploss_price = 0
                        hold_day = 0

                    d1_change = change_list[j-1]
                    d1_open = open_list[j-1]
                    d1_close = close_list[j-1]
                    d1_high = high_list[j-1]
                    d1_low = low_list[j-1]
                    d1_truerange = d1_high - d1_low
                    try:
                        d1_u_tail_ratio = (d1_open - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 윗꼬리 가격상 비율
                        d1_l_tail_ratio = (d1_close - d1_low) / d1_truerange  # 전일 변동분 중 음봉일때 아랫꼬리 가격상 비율
                    except ZeroDivisionError:
                        d1_u_tail_ratio = 0
                        d1_l_tail_ratio = 0

                    open_ratio = (open_list[j] - close_list[j - 1]) / close_list[j - 1]
                    open_ratio_list.append(open_ratio)

                    if (d1_u_tail_ratio > 0.8 and d1_l_tail_ratio < 0.2 and d1_change < condition_low and
                            open_ratio < condition_open and close > d1_low and close > open_list[j] and
                            hold_unit == 0 and N > 0):    # 신규매수 조건, 종가기준"
                        # 신규 진입 조건 : 윗꼬리, 아랫꼬리가 2% 이하 전일 종가 5%이상 하락, 당일 시가 1%이상 하락, 당일 종가가 전일 저가보다 높아야야
                    # if d1_u_tail_ratio > 0.8 and d1_l_tail_ratio < 0.2 and d1_change < condition_low and open_ratio < condition_open and high > d1_close and hold_unit == 0 and N > 0:    # 신규매수 조건, 돌파기준"
                        # 신규 진입 조건 : 윗꼬리 , 아랫꼬리가 20% 이하 전일 종가 5%이상 하락, 당일 시가 1%이상 하락, 전일 종가 돌파시

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        condition = "buy"
                        condition_fliter_list.append(condition)
                        buy_price = close                             # 매수 가격
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)

                        stoploss_price = buy_price * (1 - stoploss_R)  # 기본손절가 4% 설정

                        if low > stoploss_price:
                            stoploss_price = low
                        else:
                            stoploss_price = stoploss_price

                        stoploss_price_list.append(stoploss_price)
                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                    elif low < stoploss_price or open_list[j] < stoploss_price:      # 손절

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = hold_day_list[j - 1]
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("stoploss")
                        buy_price_list.append(0)

                        stoploss = "stoploss"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j-1]
                        stoploss_price_list.append(stoploss_price)

                        if open_list[j] < stoploss_price:
                            sell_price = open_list[j]
                        else:
                            sell_price = stoploss_price
                        sell_price_list.append(sell_price)

                        profit = sell_price - buy_price_list[j-1]
                        profit_list.append(profit)

                        profit_ratio = (profit / buy_price_list[j-1]) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                    elif hold_unit > 0 and (hold_day == max_hold_day or high > stopgain_price):   # 청산 조건 : 최대보유기간(max_hold_day) 이후 또는 익절값 이상

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)

                        buy_price_list.append(0)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        if open_list[j] > stopgain_price:
                            sell_price = open_list[j]
                            condition_fliter_list.append("stopgain")

                        elif high > stopgain_price:
                            sell_price = stopgain_price
                            condition_fliter_list.append("stopgain")
                        else:
                            sell_price = close
                            condition_fliter_list.append("sell")

                        sell_price_list.append(sell_price)
                        profit = sell_price - buy_price_list[j - 1]
                        profit_list.append(profit)
                        profit_ratio = (profit / buy_price_list[j - 1]) * 100
                        profit_ratio = round(profit_ratio, 2)
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                    elif hold_unit == 1:    # 홀드

                        hold_unit = 1
                        hold_unit_list.append(hold_unit)
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("hold")
                        buy_price = buy_price_list[j-1]
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = stoploss_price_list[j - 1]
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(1)

                    else:

                        hold_unit = 0
                        hold_unit_list.append(hold_unit)
                        hold_day = 0
                        hold_day_list.append(hold_day)
                        condition_fliter_list.append("non")         # 관망
                        buy_price = 0
                        buy_price_list.append(buy_price)

                        stoploss = "x"
                        stoploss_list.append(stoploss)
                        stoploss_price = 0
                        stoploss_price_list.append(stoploss_price)

                        sell_price = 0
                        sell_price_list.append(sell_price)

                        profit = 0
                        profit_list.append(profit)
                        profit_ratio = 0
                        profit_ratio_list.append(profit_ratio)
                        event_list.append(0)

            df_result['date'] = df['date']
            df_result['name'] = df['name']
            df_result['open'] = df['open']
            df_result['high'] = df['high']
            df_result['low'] = df['low']
            df_result['close'] = df['close']
            df_result['change'] = df['change']
            df_result['change_ratio'] = df['change_ratio']
            df_result['open_ratio'] = open_ratio_list
            # df_result['volume'] = df['volume']
            df_result['atr20'] = df['atr20(N)']
            df_result['N20_R'] = df['N20_R']
            # df_result['max_20R'] = df['max_20_R']
            # df_result['min_10R'] = df['min_10_R']
            df_result['sichong'] = df['si_chong']
            df_result['vol_d1r'] = df['vol_d1_R']

            df_result['up_ggori'] = u_tail_R_list
            df_result['btm_ggori'] = l_tail_R_list

            df_result['buy_sell'] = condition_fliter_list
            df_result['buy_price'] = buy_price_list
            df_result['sell_price'] = sell_price_list
            df_result['stoploss_price'] = stoploss_price_list
            df_result['stoploss'] = stoploss_list
            df_result['hold_unit'] = hold_unit_list
            df_result['hold_day'] = hold_day_list
            df_result['profit'] = profit_list
            df_result['profit_ratio'] = profit_ratio_list
            df_result['event'] = event_list

            # self.simul_data_to_excel(df_result)
            df_result.to_sql(name, con2, if_exists="replace")
            print("%s / %s %s 완료" % (i+1, len(name_list), name))
        print("최종완료")
        con.close()
        con2.close()
        self.execute_query_for_all_stock(save_file_name)
        self.sound_play()               #### 8020 전략 끝.


    def sound_play(self):
        # 도레미파솔라시 Hz
        so1 = {'do': 261, 're': 293, 'mi': 329, 'pa': 349, 'sol': 391, 'ra': 440, 'si': 493}
        # mel = ['do', 're', 'mi', 'pa', 'sol', 'ra', 'si']
        mel = ['do', 'mi', 'sol']
        dur = [2, 2, 4]

        music = zip(mel, dur)

        for melody, duration in music:
            winsound.Beep(so1[melody], 100 * duration)


    def del_column(self):    # 열 삭제
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % self.kospi_kosdaq)

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # name_list = ["셀트리온", "현대차", "삼성전자", "두산중공업", "동원산업", "GS건설"]  # 실험용

        for i, name in enumerate(name_list):
            print("현재 %s 진행중 %s / %s" % (name, i+1, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            # df.drop(columns=['mirae_120_max_R', 'mirae_120_min_R'], inplace=True)   # 삭제할 열이름, inplace = True(원본에 바로적용)
            df.drop(columns=['D1_open', 'D1_high', 'D1_low', 'D1_close', 'D1_open_R', 'D1_high_R', 'D1_low_R', 'D1_close_R'], inplace=True)  # 삭제할 열이름, inplace = True(원본에 바로적용)
            df.to_sql(name, con, if_exists="replace")
            print(name, "완료")

        con.close()


    def merge___(self):   # 모든 데이터 한테이블 통합
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        # file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"
        file = "c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db"

        con2 = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        # con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_add_manual.db")  # 키움증권 다운로드 종목 데이터 베이스
        # con2 = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_month_index.db")  # 키움증권 다운로드 종목 데이터 베이스

        c = con2.cursor()

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            if i > 0:
                df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con2, index_col="index")

                if len(df['open']) > 0:   # 빈데이터를 패스하기 위하여
                    print("%d / %d %s 진입" %(i, len(name_list), name))
                    if self.kospi_kosdaq == "kospi":
                        c.execute("INSERT INTO 동화약품 SELECT * FROM " + "'" + name + "'")   #코스피 합치기
                        con2.commit()
                    else:
                        c.execute("INSERT INTO 삼천당제약 SELECT * FROM " + "'" + name + "'")
                        con2.commit()

                    print(name+"추가 완료")
                c.execute("DROP TABLE " + "'" + name + "'")  # 테이블 삭제
                print(name + "삭제 완료")
        c.close()
        con2.close()
        #os.rename("kosdaq_data_add_auto.db", "kosdaq_data_one_tabel.db")
        print("한테이블 통합 최종완료, %s로 저장" % file)

    def add_minute_tech(self, name):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_minu_data.db")  #분봉데이터 로드
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)
        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다

            op = df['open'] * 0.1 * 10
            cl = df['close'] * 0.1 * 10
            hi = df['high'] * 0.1 * 10
            lo = df['low'] * 0.1 * 10
            vo = df['volume'] * 0.1 * 10
            index_list = []

            df['date'] = df['index']
            df.drop("index", axis=1, inplace=True)   # axis=1 은 열을 의미한다.

            for i in range(len(df['date'])):
                date = df['date'][i]
                time = df['time'][i]
                index = str(date) + ";" + str(time)
                index_list.append(index)
            df['index'] = index_list
            cl_list = list(cl)
            cl_list2 = cl_list[1:]
            cl_list2.append(cl_list2[-1])
            df['cl2'] = cl_list2
            df['change_2'] = df['close']-df['cl2']
            df = DataFrame(df, columns= ['index', 'date', 'time', 'name', 'open', 'high', 'low', 'close', 'change', 'volume', 'change_ratio'])
        return df

    def add_min_tech_button(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        file_name = self.kospi_kosdaq + "_minu_data.db"
        con = sqlite3.connect("c:/users/백/" + file_name)  # 분봉데이터 로드

        code_name = self.codeNname_load()  # 종목 코드로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.add_minute_tech(name)
            df.to_sql(name, con, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con.close()
        print("최종완료, %s 로 저장 완료" % file_name)


    def add_month_change(self):   #월봉 등락률 열 내용 추가
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        for j, name in enumerate(name_list):

            print("현재 작성중인 종목 : %s (%s / %s)" %(name, j, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col='index')  # 데이터 저장순서는 현재부터 과거로다
            cl_list = list(df["close"])
            change_list = []
            change_ratio_list = []
            data_lenth = len(df["change"])
            # df.drop("change", axis=1, inplace=True)
            # df.drop("change_ratio", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            # dataframe.drop("컬럼 이름", axis=1, inplace=True)
            # 여기서 axis는 축을 의미하고 0은 row를 1은 column을 의미하게 됩니다.
            # inplace=True라고 설정을 해야지 바로 drop의 동작을 실행합니다.

            # df.rename(index={"index", "date"}, inplace=True)
            # df.rename(columns={"index": "date"}, inplace=True)  #칼럼명 변경
            # print(df.head())

            for i in range(len(cl_list)):
                if i == data_lenth-1:
                    change_list.append(0)
                    change_ratio_list.append(0)

                else:
                    new_change = cl_list[i] - cl_list[i+1]
                    change_list.append(new_change)
                    if cl_list[i+1] == 0:               #종가가 0인경우 대비
                        change_ratio_list.append(0)
                    else:
                        new_change_ratio = new_change / cl_list[i+1]
                        change_ratio_list.append(new_change_ratio)
            df["change"] = change_list
            df["change_ratio"] = change_ratio_list
            df.to_sql(name, con, if_exists="replace")
        con.close()


    def execute_query(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        query_ = self.lineEdit.text()   #입력된 값을 저장
        con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db")  # 키움증권 다운로드 종목 데이터 베이스

        if self.kospi_kosdaq == "kospi":
            df = pd.read_sql("SELECT * FROM 동화약품 WHERE " + query_, con, index_col=None)   #코스피

        else:
            df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)

        con.close()
        print("완료2")
        self.simul_data_to_excel(df)

    def execute_query_for_all_stock(self, save_file_name="query_result"):    # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        # con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        con = sqlite3.connect("c:/users/백/%s_condition_filter.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        query_ = self.lineEdit.text()  # 입력된 값을 저장

        df2 = DataFrame()
        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print("%s 진입 %s / %s" % (name, i+1, len(name_list)))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query_, con, index_col='index')
            df2 = df2.append(df)
        print(df2)
        con.close()
        print("완료2")
        self.simul_data_to_excel(df2, save_file_name)

    def add_past_future_data(self, index_a_list, maximum):   # data는 쿼리를 실행한 이후 결과 데이터 프레임, past/future는 과거/미래 며칠까지 받을것인가를 정한 수, maximum는 기본데이터에서 index_A의 마지막값
        result_list = []
        past = self.spinBox_2.value()
        future = self.spinBox_3.value()
        for index_a in index_a_list:
            result_list.append(index_a)
            for i in range(past):    # 과거 며칠 데이터를 확장할 것인가
                past_index = index_a - (i + 1)
                if past_index > 0:   # 최초시작보다 커야함
                    result_list.append(past_index)

            for j in range(future):
                future_index = index_a + (j + 1)
                if future_index < maximum:   # 마지막보다 작아야함
                    result_list.append(future_index)
        final_result = set(result_list)   # 중복제거
        final_result = list(final_result)
        final_result.sort()   # 정렬
        return final_result    # 토해낸 값을 가지고 다시한번 쿼리를 써야함

    def execute_query_for_all_stock_plus_past_future(self):    # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행 + 과거 며칠, 미래 며칠까지 실행
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        print("DB파일을 불러옵니다.(c:/users/백/%s_data_add_manual.db)" % self.kospi_kosdaq)
        # con = sqlite3.connect("c:/users/백/" + self.kospi_kosdaq + "_data_high_low.db")  # 키움증권 다운로드 종목 데이터 베이스

        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        query_ = self.lineEdit.text()  # 입력된 값을 저장

        df3 = DataFrame()

        # name_list = ["비트컴퓨터", "휴먼엔", "원익", "동일기연"]    #실험용

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기, 첫번째 뺑뺑이는 최초 쿼리가 만족하는 데이터의 인덱스 값을 찾는다.
            df2 = DataFrame()   # 매번 종목이 바뀔때마다 기본 쿼리 결과를 리셋함
            print(name + " 진입", i ,"/", len(name_list))
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'", con, index_col='index')
            df1 = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query_, con, index_col='index')
            df2 = df2.append(df1)
            data = df2['index_A']
            maximum = len(df['index_A'])
            result = self.add_past_future_data(data, maximum)
            print("검색크기:", len(result))

            for j, past_future in enumerate(result):   # 과거, 미래 검색 값의 인덱스값을 가져와서 하나하나 쿼리를 묻는다.
                query_1 = "index_A = " + str(past_future)
                df4 = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query_1, con, index_col='index')
                df3 = df3.append(df4)
                if len(df3['index_A']) > 10000:
                    self.simul_data_to_excel(df3)
                    df3 = DataFrame()   # 10000행마다 엑셀에 저장하고 리셋함.
                    print("리셋함.")
        con.close()
        self.simul_data_to_excel(df3)
        print("완료2")


    def add_high_low(self, name, day):
        self.kospi_kosdaq = self.comboBox_2.currentText()

        con = sqlite3.connect("c:/users/백/%s_data_add_manual.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        try:
            df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con,
                             index_col=None)  # df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        except:
            return

        day_max_list = []
        day_min_list = []
        final_close_list = [] #최종 종가

        if len(df['open']) > 0:  # 값이 없는것을 제외시킨다
            hi = list(df['high'] * 0.1 * 10)
            lo = list(df['low'] * 0.1 * 10)
            cl = list(df['close'] * 0.1 * 10)

            for i in range(len(hi)):

                if i == len(hi)-1:
                    day_max_list.append(0)
                    day_min_list.append(0)
                    final_close_list.append(0)

                else:
                    day_high_list = hi[i + 1: i + day+1]  # 다음날 부터 + 향후 X일간의 리스트
                    day_high_list.sort(reverse=True)  # 최대값은 리스트의 첫번째 값
                    day_max_list.append(day_high_list[0])
                    last_close_list = cl[i+1:i+1+day]
                    last_close = last_close_list[-1]
                    final_close_list.append(last_close)

                    day_low_list = lo[i + 1: i + day+1]  # 다음날 부터 + 향후 X일간의 리스트
                    day_low_list.sort()  # 최소값은 리스트의 첫번째 값
                    day_min_list.append(day_low_list[0])

            df['hi_max'] = day_max_list
            df['low_min'] = day_min_list
            df['D+close'] = final_close_list

        con.close()
        return df

    def add_day_max_min(self):
        self.kospi_kosdaq = self.comboBox_2.currentText()
        code_name = self.codeNname_load()  # 종목 코드 로드
        con2 = sqlite3.connect("c:/users/백/%s_data_high_low.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스
        name_list = code_name["name"]  # 코드네임을 리스트로 저장
        day = self.spinBox.value()

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            df = self.add_high_low(name, day)
            # df.drop("level_0", axis=1, inplace=True)  # column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
            df.to_sql(name, con2, if_exists="replace")

            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("high low 최종완료")


    def add_month_index(self, name):   #종목 월검색 열 추가

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con = sqlite3.connect("c:/users/백/stock_%s_vol_ma.db" % self.kospi_kosdaq)
        month_index_list = []
        df = pd.read_sql("SELECT * FROM " + "'" + name + "' ", con, index_col=None)  # 데이터 저장순서는 현재부터 과거로다
        date_list = df['index']
        name_ = df['name'][0]
        print(name_)
        for date in date_list:
            month_index = str(date)[:6]
            name_month = name+";"+month_index
            month_index_list.append(name_month)
        con.close()
        df['month_index'] = month_index_list
        print(name_+"완료")
        return df


    def add_month_index_all(self):
        code_name = self.codeNname_load()  # 종목 코드 로드
        name_list = code_name["name"]  # 코드네임을 리스트로 저장

        self.kospi_kosdaq = self.comboBox_2.currentText()
        con2 = sqlite3.connect("c:/users/백/%s_data_month_index.db" % self.kospi_kosdaq)  # 키움증권 다운로드 종목 데이터 베이스

        for i, name in enumerate(name_list):
            df = self.add_month_index(name)
            df.to_sql(name, con2, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")
        con2.close()
        print("월봉인덱스 삽입 완료")

    def read_lookup_data_from_excel(self, filename):
        wb = load_workbook(filename=filename, data_only=True)
        sheet_range = wb["test1"]

        data1 = sheet_range['B:B']  #종목명
        data2 = sheet_range['C:C']  #날짜

        stock_list = []
        date_list = []

        for i, temp_stock in enumerate(data1):
            stock = temp_stock.value
            date = data2[i].value
            if stock == None:
                break
            stock_list.append(stock)
            date_list.append(date)
        df_dic = {'date': date_list, 'name': stock_list}
        df = DataFrame(df_dic)

        return df


    def read_day_info_from_month(self):   #엑셀의 월봉데이터에 해당하는 일봉데이터를 엑셀로 다시 저장
        self.kospi_kosdaq = self.comboBox_2.currentText()
        compare_list = self.read_excel()  #불러올 종목+월
        file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"

        con = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        size = len(compare_list)
        df2__ = {}
        df2 = DataFrame(df2__)
        for i, compare_ in enumerate(compare_list):
            query_ = "month_index == '" + compare_ + "'"
            print(query_)
            if self.kospi_kosdaq == "kospi":
                full_query = "SELECT * FROM 동화약품 WHERE " + query_
                print(full_query)
                df = pd.read_sql(full_query, con, index_col=None)  # 코스피
            else:
                df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)
            print("진행 %s / %s" % (i, size))
            if i == 0:
                df2 = df
            else:
                df2 = df2.append(df)
        print(df2)
        con.close()
        print("완료2")
        self.simul_data_to_excel(df2)

    def read_day_info_from_month_new(self):   #엑셀의 월봉데이터에 해당하는 일봉데이터를 엑셀로 다시 저장
        self.kospi_kosdaq = self.comboBox_2.currentText()
        date_and_name = self.read_lookup_data_from_excel()  #불러올 종목+월
        date_list = list(date_and_name['date'])
        print(date_list)
        name_list = list(date_and_name['name'])

        # file = "c:/users/백/" + self.kospi_kosdaq + "_data_add_auto.db"
        file = "c:/users/백/" + self.kospi_kosdaq + "_30minute.db"
        print(file)

        con = sqlite3.connect(file)  # 키움증권 다운로드 종목 데이터 베이스
        size = len(date_list)
        df2__ = {}
        df2 = DataFrame(df2__)
        for i, date_ in enumerate(date_list):
            date = str(date_)
            name = name_list[i]
            print(name)
            query_ = "date == '" + date + "'"
            # query_ = "'index' == " + date

            final_query = "SELECT * FROM " + name + " WHERE " + query_

            print(query_)
            print(final_query)

            # df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)

            df = pd.read_sql(final_query, con, index_col=None)
            print("진행 %s / %s" % (i, size))
            if i == 0:
                df2 = df
            else:
                df2 = df2.append(df)
        print(df2)
        con.close()

        self.simul_data_to_excel(df2)


    def get_minute_info_form_day_info(self):   #일봉정보를 분봉정보로 가져오기

        minute_file = "c:/users/백/minute_test.db"    #분봉 파일
        con1 = sqlite3.connect(minute_file)  # 분봉 데이터 베이스

        filename = "C:\\Users\\백\\PycharmProjects\\gaebal\\test.xlsx"
        day_data = self.read_lookup_data_from_excel(filename)    #엑셀 데이터 읽기
        date_list = day_data['date']
        df2 = DataFrame()

        for i in range(len(date_list)):
            date = day_data['date'][i]
            name = day_data['name'][i]
            query = "date = "+ str(date)
            # query = "index = "+ str(date)
            print(name, query)
            df = pd.read_sql("SELECT * FROM " + "'" + name + "'" + " WHERE " + query, con1, index_col=None)  # 코스피
            df2 = df2.append(df)   #df2.append(df)만 하면 안된다.
        print(df2)
        self.simul_data_to_excel(df2)
        con1.close()

    def calc_today_profit(self):   #원본 파일만 있을 때 쿼리에 대한 결과를 바로 엑셀로 뽑아냄
        self.add_tech()
        time.sleep(0.3)
        self.add_tech_manual()
        time.sleep(0.3)
        self.execute_query_for_all_stock()    # 한 테이블로 합치진 않은 파일에 대하여 쿼리 실행

    def test___(self):
        # self.add_month_change()
        # self.tech__("삼천당제약")  # 한종목의 기술적 지표 계산
        print("시작")

        self.daily_choochun()
        # self.add_tech('quant')

        # self.add_month_index_all()
        # df = self.read_lookup_data_from_excel()
        # self.read_day_info_from_month_new()
        # self.read_day_info_from_month()
        # self.condition_filter_whiplash()
        # self.add_minute_tech("삼천당제약")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    baek_data_= baek_data_analysis2()
    baek_data_.show()
    app.exec_()
