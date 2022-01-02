import os
import sys

# import win32com.client  #엑셀을 쓰기위한
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
from pandas import DataFrame, Series
from Kiwoom import *
from openpyxl import load_workbook

form_class = uic.loadUiType("pytrader.ui")[0]


class MyWindow(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.sell_open_done = False  # 익절 매도주문 수행 여부
        self.sell_profit_done = False  # 익절 매도주문 수행 여부
        self.sell_reserved_done = False  #예약 매수 주문 수행 여부
        self.condition_load_done = False  # 조건식 로딩 수행 여부
        self.close_sell_done = False    # 종가 매도 수행 여부
        self.close_buy_done = False    # 종가 매수 수행 여부

        self.dongsi_hoga_check_balance_done = False
        self.jisu_update = False      #지수 업데이트 수행여부, 한종목 투자금액을 정하기 위해 필요
        self.calc_one_money_done = False   #한종목 투자 금액 계산
        print("1. 전역변수 선언 완료")

        self.kiwoom = Kiwoom()
        self.kiwoom.comm_connect()
        print("2. 접속명령 실행 완료")
        self.connection_txt()   #접속시점 텍스트로 기록
        self.timer = QTimer(self)
        self.timer.start(1000)
        self.timer.timeout.connect(self.timeout)
        print("3. 타이머1 실행 완료 ")

        self.check_balance()
        print("4. 계좌 체크 요청 완료 ")

        # Timer2
        self.timer2 = QTimer(self)
        self.timer2.start(1000*10)
        self.timer2.timeout.connect(self.timeout2)
        print("5. 타이머2 실행 완료 ")

        self.lineEdit.textChanged.connect(self.code_changed)  # 여기 맞겠지??

        accounts_num = int(self.kiwoom.get_login_info("ACCOUNT_CNT"))
        accounts = self.kiwoom.get_login_info("ACCNO")
        accounts_list_ = accounts.split(";")[0:accounts_num]
        accounts_list_.append(accounts_list_[3])
        accounts_list_.remove(accounts_list_[3])   # 4번째 계좌를 마지막으로 옮김
        accounts_list = accounts_list_[::-1]
        print(accounts_list)
        print("6. 계좌정보 뿌리기 완료 ")

        self.pushButton.clicked.connect(self.send_order_manual)
        self.pushButton_2.clicked.connect(self.check_balance)
        self.pushButton_3.clicked.connect(self.condition_load)
        self.pushButton_4.clicked.connect(self.test____)
        self.pushButton_5.clicked.connect(self.auto_select_buy)
        self.pushButton_7.clicked.connect(self.kiwoom.comm_connect)
        self.pushButton_8.clicked.connect(self.all_close_sell)
        self.pushButton_9.clicked.connect(app.quit)
        print("7. 버튼 로딩 완료 ")

        self.comboBox.addItems(accounts_list)
        self.load_buy_sell_list()
        self.kiwoom.getconditionload()   #전체 조건식 로드
        self.kiwoom.OnReceiveTrCondition.connect(self._handler_tr_condition)

        # self.kosdaq_name_code_dic()  # 코스닥/코스피 코드-종목명 딕셔너리 만들기
        print(self.now__(), "  초기화 완료")
        print("===========================================================")

    def timeout(self):
        # market_start_time = QTime(22, 56, 30)
        # market_start_sell_end_time = QTime(22, 57, 7)
        # profit_sell_start_time = QTime(22, 58, 0)  # 10시에 익절주문 실행
        # profit_sell_end_time = QTime(22, 59, 0)  # 이시간이후에 다시 실행시 익절 주문을 하지 않는다.
        # quit_time_1 = QTime(23, 0, 0)
        # quit_time_2 = QTime(23, 1, 0)  # 이시간이 넘으면 프로그램을 종료하지 않는다.
        # market_end_time = QTime(23, 5, 7)

        market_start_time = QTime(9, 0, 30)
        market_start_sell_end_time = QTime(9, 4, 7)
        profit_sell_start_time = QTime(9, 8, 0)  # 10시에 익절주문 실행
        profit_sell_end_time = QTime(9, 9, 0)  # 이시간이후에 다시 실행시 익절 주문을 하지 않는다.
        quit_time_1 = QTime(9, 10, 0)
        quit_time_2 = QTime(9, 11, 0)  # 이시간이 넘으면 프로그램을 종료하지 않는다.

        dongsi_hoga_check_balance_time = QTime(15, 24, 10)
        jisu_update_time = QTime(15, 24, 40)
        calc_onestockmoney_time = QTime(15, 25, 40)
        condition_load_time = QTime(15, 26, 40)
        close_sell_time = QTime(15, 27, 40)
        close_buy_time = QTime(15, 28, 20)
        market_end_time = QTime(15, 33, 7)


        current_time = QTime.currentTime()
        text_time = current_time.toString("hh:mm:ss")
        time_msg = "현재시간: " + text_time

        state = self.kiwoom.get_connect_state()
        if state == 1:
            state_msg = "서버 연결 중"
        else:
            state_msg = "서버 미연결 중"
            f = open("connection.txt", "a")
            now = self.now__()
            f.write(now + ";"+ "disconnected")   #접속 끊긴시점 기록
            f.close()

        self.statusbar.showMessage(state_msg + "|" + time_msg)

        if self.checkBox_2.isChecked():
            if current_time < market_end_time:
                if current_time > market_start_time and current_time < market_start_sell_end_time and self.sell_open_done is False:   #시초가 비교후 매도 주문
                    print(self.now__() + ";" + "  시초가 갭상승 종목 매도 실행")
                    self.sell_order_if_gap_up(0.01)

                if current_time > profit_sell_start_time and current_time < profit_sell_end_time and self.sell_profit_done is False:  # 주식 익절매도 자동 실행
                    print(self.now__() + ";" + "  익절매도 실행")
                    self.check_my_stock()
                    data = self.check_my_stock_value
                    code_list = data['code']
                    name_list = data['name']
                    price_list = data['purchase_price']
                    num_list = data['quantity']

                    for i, code in enumerate(code_list):
                        name = name_list[i]
                        price = price_list[i]
                        num = num_list[i]
                        self.auto_profit_sell(code, name, price, num)
                    print(self.now__() + ";" + "   자동익절 주문 실행 완료")
                    print("===============================================================")
                    f = open("connection.txt", "a")
                    now = self.now__()
                    f.write(now + ";" + "자동익절 주문 실행 완료\n")
                    f.close()
                    self.sell_profit_done = True

                    if self.sell_reserved_done is False:
                        time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
                        self.buy_reserved()  # 예약한 종목의 매수
                        self.sell_reserved_done = True  # 예약 매수 주문 수행 여부

                if current_time > dongsi_hoga_check_balance_time and self.dongsi_hoga_check_balance_done is False:
                    print(self.now__() + ";" + "  동시호가 돌입 계좌 체크")
                    self.check_balance()
                    self.dongsi_hoga_check_balance_done = True

                if current_time > jisu_update_time and self.jisu_update is False:  # 지수 및 연승정보 업데이트
                    self.request_jisu()

                if current_time > calc_onestockmoney_time and self.calc_one_money_done is False:  # 한종목 투자 금액 계산
                    self.calc_one_stock_invest_money()

                if current_time > condition_load_time and self.condition_load_done is False:  #조건식 로딩 자동실행
                    self.condition_load()
                    self.condition_load_done = True
                    print(self.now__() + ";" + "   조건식 로딩 요청 완료")
                    print("===============================================================")

                    f = open("connection.txt", "a")
                    now = self.now__()
                    f.write(now + ";" + "조건식 로딩 요청\n")

                    self.cancell_order_button()          #앞선 모든 주문취소
                    print(self.now__() + ";" + "   모든 주문 취소 요청 완료")
                    print("===============================================================")
                    f.write(now + ";" + "주문 취소 요청\n")
                    f.close()

                if current_time > close_sell_time and self.close_sell_done is False:    # sell_list에 있는것 시장가 매도 주문 실행
                    print(self.now__() + ";" + "   종가 매도 주문 실행")

                    f = open("sell_list.txt", "rt")
                    sell_list = f.readlines()
                    f.close()

                    for row_data in sell_list:
                        row_data = row_data.rstrip("\n")
                        split_row_data = row_data.split(";")
                        sell_code = split_row_data[2]
                        sell_name = split_row_data[0]
                        sell_num = split_row_data[5]
                        print(sell_name, sell_code, sell_num)
                        self.sell_stock_sijangga(sell_code, sell_name, sell_num)   # 시장가 매도 주문 실행

                    # self.all_close_sell()    # 보유 중인 모든 종목 매도
                    self.close_sell_done = True
                    print(self.now__() + ";" + "   종가 매도 주문 완료")
                    print("===============================================================")
                if current_time > close_buy_time and self.close_buy_done is False:      #주식 종가 매수 자동실행
                    self.auto_select_buy()
                    self.close_buy_done = True
                    print(self.now__() + ";" + "   종가 매수 주문 완료")
                    print("===============================================================")

                if current_time > quit_time_1 and current_time < quit_time_2:
                    print("프로그램을 종료합니다.")
                    f = open("connection.txt", "a")
                    now = self.now__()
                    f.write(now + ";" + "프로그램 종료")  # 접속 끊긴시점 기록
                    f.close()
                    self.quit_()

                # if current_time > quit_time_3 and current_time < quit_time_4:
                #     print("프로그램을 종료합니다.")
                #     f = open("connection.txt", "a")
                #     now = self.now__()
                #     f.write(now + ";" + "프로그램 종료")  # 접속 끊긴시점 기록
                #     f.close()
                #     self.quit_()

    def calc_one_stock_invest_money(self):        #잔고 읽기
        # balance = self.tableWidget.item(0, 5).text()
        # today_balance = int(self.change_format_3(balance))
        today_balance = 3300000
        print(today_balance)

        # 엑셀로 투자 비중 읽어오기
        base = 'C:\\Users\\백\\PycharmProjects\\gaebal'
        file = '투자비중(모멘텀,연승).xlsx'
        print("여기2")
        file_path = os.path.join(base, file)
        wb = load_workbook(file_path, data_only=True)
        ws_invest_ratio = wb['kosdaq_mmt']
        print("여기1")
        invest_ratio = ws_invest_ratio['U3'].value
        print("투자 비중 : %s" % invest_ratio)

        self.one_stock_invest_money = int(today_balance / 30 * invest_ratio)  # 10만원
        print("한종목 투자금액 : %s" % self.one_stock_invest_money)
        display_one_stock_money = str(self.one_stock_invest_money)
        wb.close()
        self.calc_one_money_done = True
        self.lineEdit_3.setText(display_one_stock_money)
        return display_one_stock_money

    def hoga_function(self, x):   #호가 적용 가격변환 함수 --코스피 50000원이상은 나중에 수정필요함
        if x < 1000:
            ho = 1
        elif x < 5000:
            ho = 5
        elif x < 10000:
            ho = 10
        elif x < 50000:
            ho = 50
        else:
            ho = 100
        apply_hoga = int(round(x/ho, 0) * ho)
        return apply_hoga

    def timeout2(self):
        if self.checkBox.isChecked():
            print("계좌체크")
            self.check_balance()

    def now__(self):
        now = datetime.datetime.now()
        current_time = now.strftime("%H:%M:%S")
        return current_time

    def connection_txt(self):
        f = open("connection.txt", "wt")
        now = self.now__()
        f.write(now +";" + "program excute\n")
        f.close()

    def code_changed(self):
        code = self.lineEdit.text()
        name = self.kiwoom.get_master_code_name(code)
        self.lineEdit_2.setText(name)

    def send_order_manual(self):

        order_type_lookup = {"신규매수": 1, "신규매도": 2, "매수취소": 3, "매도취소": 4}
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        account = self.comboBox.currentText()
        order_type = self.comboBox_2.currentText()
        code = self.lineEdit.text()
        hoga = self.comboBox_3.currentText()
        num = self.spinBox.value()
        price = self.spinBox_2.value()
        print("매수주문 완료")

        self.kiwoom.send_order("send_order_req", "0101", account, order_type_lookup[order_type],code,num,price,
                               hoga_lookup[hoga], "")

    def check_order_info(self):   # 매매주문 조회 또는 주문취소
        # self.kiwoom.order_info_req(1)
        order_type_lookup = {"신규매수": 0, "신규매도": 0, "매수취소": 1, "매도취소": 1}
        trade_type = self.comboBox_2.currentText()
        trade_cont = int(order_type_lookup[trade_type])

        self.kiwoom.order_info_req(0, trade_cont)  # 0: 최초조회, 2: 남은데이터 추가 조회
        # print(self.kiwoom.remained_data)   #남은데이터가 있는가?
        while self.kiwoom.remained_data:
            self.kiwoom.order_info_req(2, 0)

    def cancell_order_button(self):   # 모든 주문취소
        self.kiwoom.order_info_req(0,1)
        i = 1
        while self.kiwoom.remained_data:
            print("재요청")
            i = i + 1
            self.kiwoom.order_info_req(2, 1)    #또다시 데이터가 남았을 때 어떻게 할 지 고쳐야함

    # def take_profit_order(self):
    #     order_type_lookup = {"신규매수": 1, "신규매도": 2, "매수취소": 3, "매도취소": 4}
    #     hoga_lookup = {"지정가": "00", "시장가": "03"}
    #
    #     account = self.comboBox.currentText()
    #     # code = self.lineEdit.text()
    #     hoga = self.comboBox_3.currentText()
    #     num = self.spinBox.value()
    #     price = self.spinBox_2.value()
    #
    #     self.kiwoom.send_order("take_profit_order_req", "0103", account, order_type_lookup["신규매도"], "", num, price,
    #                            hoga_lookup[hoga], order_num)

    def check_balance(self):
        self.kiwoom.reset_opw00018_output()
        account_number = self.kiwoom.get_login_info("ACCNO")
        self.account_number = account_number.split(";")[3]
        # self.account_number = 5916600510
        print(self.account_number)

        self.kiwoom.set_input_value("계좌번호", self.account_number)
        self.kiwoom.comm_rq_data("opw00018_req", "opw00018", 0, "2000")
        # 잔고 및 보유종목 현황 출력에 필요한 대부분의 데이터는 opw00018 에서 얻을 수 있음

        while self.kiwoom.remained_data:
            print("들어감")
            time.sleep(0.2)
            self.kiwoom.set_input_value("계좌번호", self.account_number)
            self.kiwoom.comm_rq_data("opw00018_req", "opw00018", 2, "2000")

        # opw00001
        self.kiwoom.set_input_value("계좌번호", self.account_number)
        self.kiwoom.comm_rq_data("opw00001_req", "opw00001", 0, "2000")

        # 잔고
        item = QTableWidgetItem(self.kiwoom.d2_deposit)  #객체 생성
        item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
        self.tableWidget.setItem(0, 0, item)

        # 총매입, 총평가, 총손익, 총수익률 데이터 삽입
        for i in range(1, 6):
            item = QTableWidgetItem(self.kiwoom.opw00018_output["single"][i-1])
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.tableWidget.setItem(0, i, item)
            self.tableWidget.resizeRowsToContents()

        # Item Iist
        item_count = len(self.kiwoom.opw00018_output["multi"])
        self.tableWidget_2.setRowCount(item_count)


        for j in range(item_count):
            row = self.kiwoom.opw00018_output["multi"][j]
            for i in range(len(row)):
                item = QTableWidgetItem(row[i])
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
                self.tableWidget_2.setItem(j, i, item)

        self.tableWidget_2.resizeRowsToContents()
        # a = self.tableWidget_2.item(1,1)   #tablewidget의 종목정보 가져오기
        # print(a.text())
        # a = self.tableWidget_2.rowCount()  #행수 카운트
        # print(a)

    # def condition_load(self):
    #     self.kiwoom.getconditionload()

    def check_acc_evaluate(self):  #당일 실현손익 조회 + 현재 보유종목 수익 조회
        # account = int(self.comboBox.currentText())
        self.kiwoom.rq_account_evaluate(self.account_number)
        fufill_profit = int(self.kiwoom.day_profit)
        hold_profit = int(self.change_format_3(self.tableWidget.item(0, 3).text()))
        self.final_day_profit = hold_profit + fufill_profit
        print(self.final_day_profit)

    def condition_load(self):
        self.today_condition_loaded = False   # 첫번째 로딩시만 파일쓰기모드를 덮어쓰기로 이후에는 이어쓰기로
        print(self.now__()+"   전략 1 로딩요청")
        self.kiwoom.sendcondition("0150", "전략1_10~15%_익절5%", "17", 0, condi_num=1)  # 조건식 1로딩  5%익절

        print(self.now__() + "   전략 2 로딩요청")
        self.kiwoom.sendcondition("0151", "전략2_7~10%_익절4%", "11", 0, condi_num=2)  # 조건식 2로딩 ---4% 익절

        # print(self.now__() + "   조건식 3 로딩요청")
        # self.kiwoom.sendcondition("0152", "D2_급등_D0_거래량증가_6퍼익절", "14", 0, condi_num=3)  # 조건식 3로딩


    def _handler_tr_condition(self, strScrNo, strCodeList, strConditionName, nIndex, nNext):
        print(self.now__(), "   조건식 요청에 대한 서버 응답.")

        if strCodeList == "":
            print("조건식 없음")
            print(self.now__(), "  조건식에 부합하는 종목이 없습니다.")
            return

        else:
            print("조건식 있음")
            code_price_list = strCodeList.split(";")[:-1]   # 조건식의 종목 코드,가격 - 종목별로 쪼갠다
            condi_code_list = []
            condi_name_list = []
            condi_price_list = []
            buy_qt_list = []  # 주문 수량

            for code_price in code_price_list:
                one_code_price_list = code_price.split("^")  # 종목코드와 가격을 쪼갠다.
                code = one_code_price_list[0]
                close = int(one_code_price_list[1])   # 현재가
                condi_code_list.append(code)
                condi_price_list.append(close)

            code_price_dic = {"코드": condi_code_list, "현재가": condi_price_list}
            df_code_price = DataFrame(code_price_dic)
            df_sort = df_code_price.sort_values("현재가", ascending=True)   # 가격이 낮은은 순으로 정렬
            df_final = df_sort[0:25]   # 조건식당 25개종목만 선정
            final_code_list = list(df_final["코드"])
            final_price_list = list(df_final["현재가"])
            self.calc_one_stock_invest_money()  # 한 종목당 투자금액 조회

            print(self.one_stock_invest_money)   # 한종목당 투자금액 조회를 반드시 먼저 수행해야함

            for i, code in enumerate(final_code_list):
                name = self.kiwoom.get_master_code_name(code)
                condi_name_list.append(name)
                buy_qt = int(round(self.one_stock_invest_money/final_price_list[i]-0.5, 0))   # 매수수량은 그날 한종목 투자 액수/종목가격
                buy_qt_list.append(buy_qt)
            print(condi_name_list)

        filename = "buy_list.txt"

        if self.today_condition_loaded == False:    # 첫번째 조건식으면 다지우고 새로쓰고 아니면 이어서 쓴다.
            file_mode = "w"
        else:
            file_mode = "a"

        file = open(filename, file_mode)    # 조건 검색 결과 쓰기

        if strScrNo == "0150":   # 스크린넘버 정보를 가지고 익절비율 설정
            stop_profit_ratio = "0.05"
        elif strScrNo == "0151":
            stop_profit_ratio = "0.04"
        else:
            stop_profit_ratio = "0.06"

        for i, con_code in enumerate(final_code_list):
            condi_codestr = str(con_code)
            condi_name = condi_name_list[i]
            condi_price = str(final_price_list[i])
            buy_qt = str(buy_qt_list[i])
            write_stuff = condi_name + ";매수;" + condi_codestr + ";시장가;" + condi_price + ";" + buy_qt + ";0;" + stop_profit_ratio + ";매수전\n"
            file.write(write_stuff)
        file.close()

        filename_2 = "sell_list.txt"
        file_2 = open(filename_2, "wt")
        f = open(filename, "rt")
        sell_list = f.readlines()
        for i in range(len(sell_list)):
            sell_list[i] = sell_list[i].replace("매수", "매도")
        for write_thing in sell_list:
            file_2.write(write_thing)
        file_2.close()

        print(self.now__(), " =================================조건식 함수 실행 완료=================================")
        time.sleep(1)
        self.load_buy_sell_list()  # 화면에 검색 종목 반영
        self.today_condition_loaded = True   # 첫번째 로딩시만 파일쓰기모드를 덮어쓰기로 이후에는 이어쓰기로

    def load_buy_sell_list(self):
        f = open("buy_list.txt", "rt")
        buy_list = f.readlines()
        f.close()

        f = open("sell_list.txt", "rt")
        sell_list = f.readlines()
        f.close()

        row_count = len(buy_list) + len(sell_list)
        self.tableWidget_3.setRowCount(row_count)    # 자동 선정 종목리스트 행 수 결정

        # buylist

        for j in range(len(buy_list)):
            row_data = buy_list[j]
            split_row_data = row_data.split(";")
            # split_row_data[1] = self.kiwoom.get_master_code_name(split_row_data[1].rsplit())
            k = 0  # 필요없는 항목을 건너뛰기 위해 설정한 변수

            for i in range(len(split_row_data)):

                if i == 2 or i == 6:  # 종목코드, 0 건너뛰기 2 = 종목코드, 6 = 시장가
                    continue
                item = QTableWidgetItem(split_row_data[i].rstrip())
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget_3.setItem(j, k, item)
                k = k + 1

        for j in range(len(sell_list)):
            row_data = sell_list[j]
            split_row_data = row_data.split(";")
            k = 0  # 필요없는 항목을 건너뛰기 위해 설정한 변수

            for i in range(len(split_row_data)):
                if i == 2 or i == 6:  # 종목코드, 0 건너뛰기
                    continue
                item = QTableWidgetItem(split_row_data[i].rstrip())
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
                self.tableWidget_3.setItem(len(buy_list)+ j, k, item)
                k = k + 1

        # for j in range(len(sell_list)):
        #     row_data = sell_list[j]
        #     split_row_data = row_data.split(";")
        #     # split_row_data[1] = self.kiwoom.get_master_code_name(split_row_data[1].rsplit())
        #
        #     for i in range(len(split_row_data)):
        #         item = QTableWidgetItem(split_row_data[i].rstrip())
        #         item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
        #         self.tableWidget_3.setItem(len(buy_list)+j, i, item)

        self.tableWidget_3.resizeRowsToContents()

    def buy_reserved(self):  # 예약한 종목의 매수

        # 엑셀로 매수 예정 내역 읽어오기
        base = 'C:\\Users\\백\\PycharmProjects\\gaebal'
        file = 'reserved_trade.xlsx'
        file_path = os.path.join(base, file)
        wb = load_workbook(file_path, data_only=True)
        ws_buy = wb['buy']
        i = 2
        code_list = []
        name_list = []
        price_list = []
        num_list = []
        before_after_list = []
        while 1:   # 무한루프
            i = i + 1
            i = str(i)
            buy_code = ws_buy['B' + i].value
            if buy_code == None:
                print("코드가 없습니다.")
                break
            code_list.append(buy_code)

            buy_name = ws_buy['A' + i].value
            name_list.append(buy_name)
            buy_price = ws_buy['C' + i].value
            price_list.append(buy_price)
            buy_num = ws_buy['E' + i].value
            buy_num = int(buy_num)
            num_list.append(buy_num)
            buy_before_after = ws_buy['F' + i].value
            before_after_list.append(buy_before_after)
            i = int(i)

        account = 5514437210

        # 매수 주문
        for j, code in enumerate(code_list):

            name = name_list[j]
            print(name)
            price = price_list[j]
            price = int(price)
            num = num_list[j]   # 주문 수량
            before_after = before_after_list[j]

            if before_after == "매수전":
                print(self.now__(), "  ", name)
                time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
                print("주문 내역 : %s, 코드; %s, 수량; %s, 가격; %s" % (name, code, num, price))
                self.kiwoom.send_order("send_order_req", "8979", account, 1, code, num, price, "00", "")

                print(self.now__(), "  %s 예약 매수주문 완료" % name)
                print("===================================================================")
                row = j + 3
                row = str(row)
                ws_buy['F' + row] = "매수완료"
        wb.save('reserved_trade.xlsx')
        wb.close()

        # # 엑셀로 매수 예정 내역 읽어오기
        # base = 'C:\\Users\\백\\PycharmProjects\\gaebal'
        # file = 'reserved_trade.xlsx'
        # file_path = os.path.join(base, file)
        # wb = load_workbook(file_path, data_only=True)
        # ws_buy = wb['buy']
        # i = 2
        # code_list = []
        # name_list = []
        # price_list = []
        # num_list = []
        # before_after_list = []
        # while 1:  # 무한루프
        #     i = i + 1
        #     i = str(i)
        #
        #     buy_code = ws_buy['B' + i].value
        #     if buy_code == None:
        #         print("코드가 없습니다.")
        #         break
        #     code_list.append(buy_code)
        #
        #     buy_name = ws_buy['A' + i].value
        #     name_list.append(buy_name)
        #     buy_price = ws_buy['C' + i].value
        #     price_list.append(buy_price)
        #     buy_num = ws_buy['E' + i].value
        #     buy_num = int(buy_num)
        #     num_list.append(buy_num)
        #     buy_before_after = ws_buy['F' + i].value
        #     before_after_list.append(buy_before_after)
        #
        #     i = int(i)
        #
        # account = self.comboBox.currentText()  # 불안정하므로 나중에 수정할 것
        #
        # # 매수 주문
        # for j, code in enumerate(code_list):
        #
        #     name = name_list[j]
        #     print(name)
        #     price = price_list[j]
        #     num = num_list[j]  # 주문 수량
        #     before_after = before_after_list[j]
        #
        #     if before_after == "매수전":
        #         print(self.now__(), "  ", name)
        #         time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
        #         print("주문 내역 : %s, 코드; %s, 수량; %s, 가격; %s" % (name, code, num, price))
        #         # self.kiwoom.send_order("send_order_req", "0101", account, 1, code, num, price, "03", "")
        #         print(self.now__(), "  %s 예약 매수주문 완료" % name)
        #         print("===================================================================")
        #         row = j + 3
        #         row = str(row)
        #         ws_buy['F' + row] = "매수완료"
        # wb.save('reserved_trade.xlsx')
        # wb.close()


    def auto_select_buy(self):    # 조건식의 종목 매수
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        f = open("buy_list.txt", "rt")   # 텍스트 파일에 있는 내용을 읽어서 매수주문
        buy_list = f.readlines()   # 텍스트 파일 전체를 다 읽음
        f.close()
        account = self.comboBox.currentText()

        # 매수 주문
        for row_data in buy_list:
            split_row_data = row_data.split(";")
            name = split_row_data[0]
            hoga = split_row_data[3]
            code = split_row_data[2]
            num = split_row_data[5]  # 주문 수량
            if hoga == "시장가":
                price = 0
            else:
                price = split_row_data[4]  # 주문 가격

            if split_row_data[-1].rstrip() == "매수전":
                print(self.now__(), "  ", name)
                time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
                print("주문 내역 : 호가; %s, 코드; %s, 수량; %s, 가격; %s" %(hoga, code, num, price))
                self.kiwoom.send_order("send_order_req", "0101", account, 1, code, num, price, hoga_lookup[hoga], "")
                print(self.now__(), "  %s 매수주문 완료" % name)
                print("===================================================================")

        # 파일 주문완료로 변경
        for i, row_data in enumerate(buy_list):
            buy_list[i] = buy_list[i].replace("매수전", "주문완료")

        f = open("buy_list.txt", "wt")
        for row_data in buy_list:
            f.write(row_data)
        f.close()

    def auto_select_sell(self):  # 조건식의 종목 매도

        hoga_lookup = {"지정가": "00", "시장가": "03"}

        f = open("sell_list.txt", "rt")  # 텍스트 파일에 있는 내용을 읽어서 매수주문
        buy_list = f.readlines()  # 텍스트 파일 전체를 다 읽음
        f.close()
        account = self.comboBox.currentText()

        # 매수 주문
        for row_data in buy_list:
            split_row_data = row_data.split(";")
            name = split_row_data[0]
            hoga = split_row_data[3]
            code = split_row_data[2]
            num = split_row_data[5]  # 주문 수량
            if hoga == "시장가":
                price = 0
            else:
                price = split_row_data[4]  # 주문 가격

            if split_row_data[-1].rstrip() == "매수전":
                print(self.now__(), "  ", name)
                time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
                print("주문 내역 : 호가; %s, 코드; %s, 수량; %s, 가격; %s" % (hoga, code, num, price))
                self.kiwoom.send_order("send_order_req", "0101", account, 1, code, num, price, hoga_lookup[hoga],
                                       "")
                print(self.now__(), "  %s 매수주문 완료" % name)
                print("===================================================================")

        # 파일 주문완료로 변경
        for i, row_data in enumerate(buy_list):
            buy_list[i] = buy_list[i].replace("매수전", "주문완료")

        f = open("buy_list.txt", "wt")
        for row_data in buy_list:
            f.write(row_data)
        f.close()

    def check_my_stock(self):
        self.check_balance()

        code_list = []
        name_list = []
        quantity_list = []
        purchase_price_list = []
        open_list = []
        open_ratio_list = []

        df = DataFrame()

        for i, data in enumerate(self.kiwoom.opw00018_output['multi']):
            code = data[0]
            code = code[1:]
            code_list.append(code)
            name = data[1]
            print(name)
            name_list.append(name)
            quantity = data[2]
            quantity_list.append(quantity)
            purchase_price = data[3]
            purchase_price = self.change_format_3(purchase_price)
            purchase_price = int(purchase_price)
            purchase_price_list.append(purchase_price)

            self.kiwoom.base_info_request(code)
            time.sleep(0.4)

            open = self.kiwoom.base_stock_info[1]  # 종목명, 시가, 고가, 저가, 현재가 + 보유수량 +매수가 ,,, 앞에 부호제거
            open = self.change_format_4(open)  # 맨앞 부호제거
            open = int(open)
            open_list.append(open)
            # purchase_price = purchase_price_list[i]
            # purchase_price = int(purchase_price)
            open_ratio = round((open - purchase_price) / purchase_price, 4)
            open_ratio_list.append(open_ratio)

        df['code'] = code_list
        df['name'] = name_list
        df['quantity'] = quantity_list
        df['purchase_price'] = purchase_price_list
        df['open'] = open_list
        df['open_R'] = open_ratio_list
        print(df)
        self.check_my_stock_value = df


    ####이 아래부분 작성해야함



    def auto_profit_sell(self, code, name, price, num):    #자동선택 익절 매도 주문  #반드시 보유종목 조회 이후에 실행해야함
        account = self.comboBox.currentText()
        # account = "5514437210"
        print(self.now__(), "  익절주문 실행")
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        jumun_price = self.hoga_function(round(int(price) * 1.04, 0))      #익절값 4%
        hoga = "지정가"
        print(code, name, num)
        self.kiwoom.send_order("send_order_req", "8289", account, 2, code, num, jumun_price, hoga_lookup[hoga], "")
        print(self.now__(), "  %s 익절주문 완료, 수량 : %s, 가격; %s" % (name, num, jumun_price))
        time.sleep(0.4)  # 주문 초당 5회 초과시 문제될 수 있다.

    def auto_profit_sell_from_txt(self):    #자동선택 익절 매도 주문  #반드시 보유종목 조회 이후에 실행해야함
        account = self.comboBox.currentText()
        # account = "5514437210"
        print(self.now__(), "  익절주문 실행")
        hoga = "지정가"
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        f = open("sell_list.txt", "rt")
        sell_list = f.readlines()
        f.close()

        # 익절주문 실행
        for row_data in sell_list:
            print(row_data)
            split_row_data = row_data.split(";")
            name = split_row_data[0]
            code = split_row_data[2]
            num = split_row_data[5]
            sell_price = int(split_row_data[4]) * (1 + float(split_row_data[-2]))   # 익절값 곱하기
            sell_price = self.hoga_function(sell_price)

            self.kiwoom.send_order("send_order_req", "8289", account, 2, code, num, sell_price, hoga_lookup[hoga], "")
            print(self.now__(), "  %s 익절주문 실행합니다. 수량 : %s, 가격; %s" % (name, num, sell_price))
            time.sleep(0.4)  # 주문 초당 5회 초과시 문제될 수 있다.

    def auto_profit_sell_00(self):    #자동선택 익절 매도 주문  #반드시 보유종목 조회 이후에 실행해야함
        account = self.comboBox.currentText()
        row = self.tableWidget_2.rowCount()  # 행수 카운트
        print(self.now__(), "  익절주문 실행")
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        # 전종목 종가 매도 주문
        for i in range(row):    #수정된 주문버전
            code = self.tableWidget_2.item(i, 0).text()
            code = code[1:]
            name = self.tableWidget_2.item(i, 1).text()
            num = int(self.tableWidget_2.item(i, 2).text())  # 주문 수량
            orginal_price = self.tableWidget_2.item(i, 3).text()
            orginal_price_ch_form = self.change_format_3(orginal_price)  #쉼표 제거
            price = self.hoga_function(round(int(orginal_price_ch_form) * 1.04, 0))      #익절값 4%
            hoga = "지정가"
            profit_ratio = "4%"
            print(code, name, num)

            time.sleep(0.4)  # 주문 초당 5회 초과시 문제될 수 있다.
            self.kiwoom.send_order("send_order_req", "8289", account, 2, code, num, price, hoga_lookup[hoga], "")
            print(self.now__(), "  %s %s 익절주문 완료, 수량 : %s, 가격; %s" % (name, profit_ratio, num, price))

        print(self.now__(), "  모든 종목 익절 매도주문 완료")
        print("===========================================================")

    def sell_stock_sijangga(self, code, name, num):    #자동선택 익절 매도 주문  #반드시 보유종목 조회 이후에 실행해야함
        account = self.comboBox.currentText()
        # account = "5514437210"
        print(self.now__(), "  시장가 매도주문 실행")
        hoga_lookup = {"지정가": "00", "시장가": "03"}
        jumun_price = 0
        hoga = "시장가"
        self.kiwoom.send_order("send_order_req", "8285", account, 2, code, num, jumun_price, hoga_lookup[hoga], "")
        print(self.now__(), "  %s 시장가 매도 주문 완료, 수량 : %s" % (name, num))
        time.sleep(0.4)  # 주문 초당 5회 초과시 문제될 수 있다.

    def send_order_auto(self, code, name, price, num):   #매도 주문 전용
        account = self.comboBox.currentText()
        print(self.now__(), "  익절주문 실행")
        hoga_lookup = {"지정가": "00", "시장가": "03"}
        hoga = "지정가"
        self.kiwoom.send_order("send_order_req", "8389", account, 2, code, num, price, hoga_lookup[hoga], "")   # 2 = 매도
        print(self.now__(), "  %s 주문 완료, 수량 : %s, 가격; %s" % (name, num, price))
        time.sleep(0.4)  # 주문 초당 5회 초과시 문제될 수 있다.


    def all_close_sell(self):  # 보유종목 전량 시장가 매도, 반드시 잔고 조회 후 해야함
        # account = self.comboBox.currentText()
        account = 5514437210
        row = self.tableWidget_2.rowCount()  # 행수 카운트
        print(self.now__(), "  시장가 종가 매도 주문 실행")
        print(self.now__(), "  보유종목수 : %s 종목" % row)
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        # 전종목 종가 매도 주문
        for i in range(row):
            code = self.tableWidget_2.item(i, 0).text()  #수정된 종가 매도 주문
            code = code[1:]
            name = self.tableWidget_2.item(i, 1).text()
            num = int(self.tableWidget_2.item(i, 2).text())  # 주문 수량
            hoga = "시장가"
            print(self.now__(), "  %s 시장가 매도주문, 수량: %s" % (name, num))
            time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.

            print(code, name, num)
            self.kiwoom.send_order("send_order_req", "8290", account, 2, code, num, 0, hoga_lookup[hoga], "")

            # name = self.tableWidget_2.item(i, 0).text()   #수정 전 종가 매도 주문
            # code = self.kospi_kosdaq_code_name_dic[name]
            # num = int(self.tableWidget_2.item(i, 1).text()) # 주문 수량
            # hoga = "시장가"
            # print(self.now__(), "  %s 시장가 매도주문, 수량: %s" % (name, num))
            # time.sleep(0.5)  # 주문 초당 5회 초과시 문제될 수 있다.
            # self.kiwoom.send_order("send_order_req", "8290", account, 2, code, num,0, hoga_lookup[hoga], "")
        print(self.now__(), "  모든 종목 시장가 매도주문 완료")
        print("===========================================================")

    def trade_stocks(self):
        hoga_lookup = {"지정가": "00", "시장가": "03"}

        f = open("buy_list.txt", "rt")
        buy_list = f.readlines()
        f.close()

        f = open("sell_list.txt", "rt")
        sell_list = f.readlines()
        f.close()
        account = self.comboBox.currentText()

        # 사자 주문
        for row_data in buy_list:
            split_row_data = row_data.split(";")
            hoga = split_row_data[3]
            code = split_row_data[2]
            num = split_row_data[4]
            price = split_row_data[5]

            if split_row_data[-1].rstrip() == "매수전":
                # self.kiwoom.send_order("send_order_req", "0101", account, 1, code, num, price, hoga_lookup[hoga], "")
                print("가상매수")
        print("매수주문 완료")

        # 팔자 주문
        for row_data in sell_list:
            split_row_data = row_data.split(";")
            hoga = split_row_data[3]
            code = split_row_data[2]
            num = split_row_data[4]
            price = split_row_data[5]

            if split_row_data[-1].rstrip() == "매도전":
                print("가상매도")
                # self.kiwoom.send_order("send_order_req", "0101", account, 2, code, num, price, hoga_lookup[hoga], "")
        print("매도주문 완료")

        # 주문완료 텍스트 파일 업데이트
        # 매수
        for i, row_data in enumerate(buy_list):
            buy_list[i] = buy_list[i].replace("매수전", "주문완료")

        f = open("buy_list.txt", "wt")
        for row_data in buy_list:
            f.write(row_data)
        f.close()

        # 매도
        for i, row_data in enumerate(sell_list):
            sell_list[i] = sell_list[i].replace("매수전", "주문완료")

        f = open("sell_list.txt", "wt")
        for row_data in sell_list:
            f.write(row_data)
        f.close()
        print("매수매도 파일 업데이트 완료")

    def kosdaq_name_code_dic(self):  #코스닥 종목명:코드 딕셔너리 만들기

        kosdaq_code_list = self.kiwoom.get_code_list_by_market("10")  # 코스닥 코드리스트 받기
        kospi_code_list = self.kiwoom.get_code_list_by_market("0")  # 코스닥 코드리스트 받기
        kospi_kosdaq_code_list = kospi_code_list + kosdaq_code_list
        self.kospi_kosdaq_code_name_dic = {}

        name_list = []
        for i, code in enumerate(kospi_kosdaq_code_list):       #삭제예정 코드
            name = self.kiwoom.get_master_code_name(code)
            name_list.append(name)
            self.kospi_kosdaq_code_name_dic[name] = code   #딕셔너리 명 [ "추가할키"] = 추가할값

        # for i, code in enumerate(kospi_kosdaq_code_list):   #추가 예정 코드
        #     name = self.kiwoom.get_master_code_name(code)
        #     name_list.append(name)
        #
        # self.kospi_kosdaq_code_name_dic = {"code": kospi_kosdaq_code_list, "name": name_list}
        # code_data = DataFrame(self.kospi_kosdaq_code_name_dic, columns=["code"], index=self.kospi_kosdaq_code_name_dic["name"])
        # con = sqlite3.connect("c:/users/백/kiwoom_stock_code.db")
        # code_data.to_sql("stock_code", con, if_exists="replace")
        # con.close()
        print(self.now__(), "  종목코드 저장 완료")
        print("===========================================================")

    def change_format_4(self, data):   #부호빼기
        dummy_list = []
        for word in data:
            if word != "+" and word != "-":
                dummy_list.append(word)
        result = ''.join(dummy_list)
        return result

    def change_format_3(self, data):    #쉽표 빼기
        dummy_list = []
        for word in data:
            if word != ",":
                dummy_list.append(word)
        result = ''.join(dummy_list)
        return result

    def now_price_button(self):
        close_list = []
        print(self.code_list)
        for code in self.code_list:
            print(code)
            close = self.kiwoom.base_info_request(code)
            close_list.append(close)
            print(close_list)
        print(close_list)

    def base_info(self):
        # self.kiwoom.base_info_request(code)
        self.kiwoom.base_info_request("019660")

    def request_jisu(self):   #지수와 연승 현황을 업데이트 함
        market = 1        #시장구분 = 0:코스피, 1:코스닥, 2:코스피200
        code = 101 # 업종코드 = 001:종합(KOSPI), 002: 대형주, 003: 중형주, 004: 소형주 101: 종합(KOSDAQ), 201: KOSPI200, 302: KOSTAR, 701: KRX100 나머지 ※ 업종코드 참고
        self.kiwoom.rq_jisu(market, code)
        print(self.kiwoom.today_jisu_data)
        invest_ratio_file = 'C:\\Users\\백\\PycharmProjects\\gaebal\\투자비중(모멘텀,연승).xlsx'
        wb = load_workbook(invest_ratio_file)
        ws = wb['kosdaq_mmt']
        max_row = ws.max_row    #행 최대값(행 수)
        today = int(datetime.datetime.today().strftime("%Y%m%d"))

        for i, data in enumerate(self.kiwoom.today_jisu_data):
            new_data = data[1:]
            float_data = float(new_data)

            ws.cell(max_row+1, i+2, float_data)
        ws.cell(max_row+1, 1, today)
        momentum_list = []

        for j in range(10):
            today_jisu = ws.cell(max_row+1, 5).value
            DX_jisu = ws.cell(max_row - j, 5).value

            if today_jisu > DX_jisu:
                momentum_list.append(1)
            else:
                momentum_list.append(0)
        print(momentum_list)
        sum_momentum = 0

        for momentum in momentum_list:   #평균 모멘텀 구하기
            sum_momentum += momentum
        average_momentum = sum_momentum / len(momentum_list)
        print(average_momentum)
        ws.cell(max_row+1, 19, average_momentum)
        ws.cell(1, 21, average_momentum)

        # 지수 모멘텀 업데이트 완료

        # 연승 업데이트
        # self.check_acc_evaluate()  # 금일 투자손익 확인
        # day_profit = self.final_day_profit
        # ws_v = wb['kosdaq_victory']
        # max_row2 = ws_v.max_row
        # ws_v.cell(max_row2+1, 1, today)
        # ws_v.cell(max_row2 + 1, 2, day_profit)
        # previous_victory = ws_v.cell(max_row2, 3).value
        # limit_victory = ws_v.cell(1, 2).value
        # print("직전 연승 : %s" % previous_victory)
        #
        # if previous_victory == limit_victory:
        #     series_victory = 0
        # elif day_profit <= 0:
        #     series_victory = 0
        # else:
        #     series_victory = previous_victory+1
        # print("현재 연승 : %s" % series_victory)
        #
        # victory_incentive = ws_v.cell(2,2).value
        # final_victory_ratio = 1 + (series_victory * victory_incentive)  #최종 연승으로 인한 투자비율
        #
        # ws_v.cell(max_row2 + 1, 3, series_victory)
        # ws_v.cell(max_row2 + 1, 4, final_victory_ratio)
        # print("최종 연승 투자 비율 : %s" % final_victory_ratio)

        #연승업데이트 완료
        # final_invest_ratio = final_victory_ratio * average_momentum
        # ws.cell(2, 21, final_victory_ratio)
        # ws.cell(3, 21, final_invest_ratio)
        ws.cell(3, 21, average_momentum)   #연승 투자 보류함함
       # print("최종 투자 비율 : %s" % final_invest_ratio)
        wb.save('투자비중(모멘텀,연승).xlsx')
        wb.close()
        self.jisu_update = True
        print("투자 비율 업데이트 최종완료2")

    def quit_(self):
        sys.exit(1)

    def test_5(self):
        self.check_my_stock()
        data = self.check_my_stock_value
        code_list = data['code']
        name_list = data['name']
        price_list = data['purchase_price']
        num_list = data['quantity']

        for i, code in enumerate(code_list):
            name = name_list[i]
            price = price_list[i]
            num = num_list[i]
            self.auto_profit_sell(code, name, price, num)
        print(self.now__() + ";" + "   자동익절 주문 실행 완료")
        print("===============================================================")
        f = open("connection.txt", "a")
        now = self.now__()
        f.write(now + ";" + "자동익절 주문 실행 완료\n")
        f.close()
        self.sell_profit_done = True
    ##@@@@
        # for row_data in sell_list:
        #     row_data = row_data.rstrip("\n")
        #     split_row_data = row_data.split(";")
        #     sell_code = split_row_data[2]
        #     sell_name = split_row_data[0]
        #     sell_num = split_row_data[5]
        #     print(sell_name, sell_code, sell_num)
        #     self.sell_stock_sijangga(sell_code, sell_name, sell_num)  # 시장가 매도 주문 실행

    def test_2(self):
        self.check_my_stock()
        data = self.check_my_stock_value
        code_list = data['code']
        name_list = data['name']
        price_list = data['purchase_price']
        num_list = data['quantity']
        print("여기임")

        for i, code in enumerate(code_list):
            name = name_list[i]
            price = price_list[i]
            num = num_list[i]
            self.auto_profit_sell(code, name, price, num)
        print(self.now__() + ";" + "   자동익절 주문 실행 완료")

    def sell_order_if_gap_up(self, open_sell_condition):   #갭상승시 매도 #시간 조건은 함수 선언 앞에 입력 = 함수안에는 없음, open_sell_condition = 갭상승시 매도 조건 ex) 0.01이면 1%이상 갭상승시 매도
        # 갭상승시 시초가로 매도 주문을 낸다.
        self.check_my_stock()
        data = self.check_my_stock_value
        code_list = data['code']
        name_list = data['name']
        open_ratio_list = data['open_R']
        purchase_price_list = data['purchase_price']
        num_list = data['quantity']
        sell_code_list = []
        sell_name_list = []
        sell_price_list = []
        sell_num_list = []

        for i, open_ratio in enumerate(open_ratio_list):
            if open_ratio > open_sell_condition:  # 시가가 조건보다 크면 코드, 종목명...을 추가한다.
                sell_code = code_list[i]
                sell_code_list.append(sell_code)
                sell_name = name_list[i]
                sell_name_list.append(sell_name)
                sell_price = purchase_price_list[i] * (1 + open_ratio)
                sell_price = self.hoga_function(sell_price)
                sell_price_list.append(sell_price)
                sell_num = num_list[i]       #수량
                sell_num_list.append(sell_num)

        for j, sell_code in enumerate(sell_code_list):
            sell_name = sell_name_list[j]
            sell_price = sell_price_list[j]
            sell_num = sell_num_list[j]

            self.send_order_auto(sell_code, sell_name, sell_price, sell_num)
        self.sell_open_done = True    #재주문하지 않도록
        print(self.now__() + ";" + "   시초가 익절 주문 실행 완료, ")
        print("===============================================================")

    def test_3(self):
        sell_condition = -0.1  # 4%이상 갭상승일때 매도
        self.check_my_stock()
        data = self.check_my_stock_value
        code_list = data['code']
        name_list = data['name']
        print(name_list)
        open_ratio_list = data['open_R']
        purchase_price_list = data['purchase_price']
        num_list = data['quantity']
        sell_code_list = []
        sell_name_list = []
        sell_price_list = []
        sell_num_list = []

        for i, open_ratio in enumerate(open_ratio_list):
            if open_ratio > sell_condition:  # 시가가 조건보다 크면 코드, 종목명...을 추가한다.
                sell_code = code_list[i]
                sell_code_list.append(sell_code)
                sell_name = name_list[i]
                sell_name_list.append(sell_name)
                sell_price = purchase_price_list[i]
                sell_price_list.append(sell_price)
                sell_num = num_list[i]
                sell_num_list.append(sell_num)

        for j, sell_code in enumerate(sell_code_list):
            sell_name = sell_name_list[j]
            sell_price = sell_price_list[j]
            sell_num = sell_num_list[j]
            self.auto_profit_sell(sell_code, sell_name, sell_price, sell_num)
        self.sell_open_done = True
        print(self.now__() + ";" + "   시초가 자동익절 주문 실행 완료")
        print("===============================================================")

    def test____(self):
        if self.spinBox_3.value() == 1:
            self.cancell_order_button()
        elif self.spinBox_3.value() == 2:
            self.auto_select_buy()  # 자동선택 종가 매도 주문
        elif self.spinBox_3.value() == 3:
            self.cancell_order_button()  # 앞선 모든 주문취소
            print(self.now__() + ";" + "   모든 주문 취소 요청 완료")
        elif self.spinBox_3.value() == 4:
            self.condition_load()
        elif self.spinBox_3.value() == 5:
            self.buy_reserved()
        elif self.spinBox_3.value() == 6:
            # print("프로그램을 종료합니다.")
            # f = open("connection.txt", "a")
            # print("111")
            # now = self.now__()
            # print("22")
            # f.write(now + ";" + "프로그램 종료")  # 접속 끊긴시점 기록
            # print("33")
            # f.close()
            print("44")
            sys.exit()

        else:
            self.auto_profit_sell_from_txt()
            print("다시입력")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.show()
    app.exec_()
