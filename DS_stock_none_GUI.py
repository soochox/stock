import sys
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
import win32com.client
import talib as ta
from talib import abstract
from PyQt5.QtCore import *
import numpy as np
import matplotlib.pyplot as plt
import FinanceDataReader as fdr

import time
import datetime
import pandas as pd
from pandas import DataFrame, Series
import sqlite3

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class Daeshin_stock(QAxWidget):   #QAxWidget 클래스를 상속받아 QAxWidget의 모든 기능이 사용가능
    def __init__(self):
        super().__init__()


    def DS_kosdaq_code_load(self):    #대신서버로부터 종목코드 갱신
        instCpCodeMgr = win32com.client.Dispatch("CpUtil.CpCodeMgr")
        kospi_codelist = instCpCodeMgr.GetStockListByMarket(1)
        kospi_namelist = []
        kosdaq_codelist = instCpCodeMgr.GetStockListByMarket(2)
        kosdaq_namelist =[]

        for code in kospi_codelist:
            kospi_namelist.append(instCpCodeMgr.CodeToName(code))


        kospi_codeNname = {"code" : kospi_codelist, "name": kospi_namelist}

        df_kospi_codeNname = DataFrame(kospi_codeNname)


        con = sqlite3.connect("c:/users/백/DS_codedata.db")
        df_kospi_codeNname.to_sql("kospi", con, if_exists="replace")


        for code in kosdaq_codelist:
            kosdaq_namelist.append(instCpCodeMgr.CodeToName(code))

        kosdaq_codeNname = {"code" : kosdaq_codelist, "name" : kosdaq_namelist}

        df_kosdaq_codeNname = DataFrame(kosdaq_codeNname)
        con = sqlite3.connect("c:/users/백/DS_codedata.db")
        df_kosdaq_codeNname.to_sql("kosdaq", con, if_exists="replace")

    def DS_save_jisu__(self, start_date, end_date,kospi_kosdaq,bong):  #지수 정보 받기
        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")
        if kospi_kosdaq == "kosdaq":
            code = "U201" #코스닥 코드임
        else:
            code = "U001"

        instStockChart.SetInputValue(0, code)

        if bong == "D":  # 일봉요청이면
            instStockChart.SetInputValue(1, ord("1"))  # 날짜로 요청
            instStockChart.SetInputValue(2, end_date)  # 요청기간
            instStockChart.SetInputValue(3, start_date)  # 요청기간
        else:
            instStockChart.SetInputValue(1, ord("2"))  # 개수로 요청
            instStockChart.SetInputValue(4, 360)  # 요청개수 144개월 = 12년

        if bong == "m":
            instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  # 1번은 시간
            boon = int(self.comboBox_3.currentText())
            instStockChart.SetInputValue(7, boon)  # 틱/분봉주기

        else:
            instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))  # 1번 시간 제외

        instStockChart.SetInputValue(6, ord(bong))  # 봉 구분
        instStockChart.SetInputValue(9, ord("1"))  # 수정 주가

        # BlockRequest
        instStockChart.BlockRequest()

        time.sleep(0.25)  # 1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

        # GetHeaderValue
        numdata = instStockChart.GetHeaderValue(3)
        numfield = instStockChart.GetHeaderValue(1)

        # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:change_ratio, 7:vol
        date_list = []
        open_list = []
        high_list = []
        low_list = []
        close_list = []
        change_list = []
        change_ratio_list = []
        volume_list = []

        for j in range(numdata):

            date = instStockChart.GetDataValue(0, j)
            date_list.append(date)
            open = instStockChart.GetDataValue(1, j)
            open_list.append(open)
            high = instStockChart.GetDataValue(2, j)
            high_list.append(high)
            low = instStockChart.GetDataValue(3, j)
            low_list.append(low)
            close = instStockChart.GetDataValue(4, j)
            close_list.append(close)
            change = instStockChart.GetDataValue(5, j)
            change_list.append(change)
            change_ratio = change / (close - change)
            change_ratio_list.append(change_ratio)
            volume = instStockChart.GetDataValue(6, j)
            volume_list.append(volume)


        ds_ohlcv = {"date": date_list, "open": open_list, "high": high_list, "low": low_list, "close": close_list,
                    "change": change_list, "change_ratio": change_ratio_list, "volume": volume_list}

        df = pd.DataFrame(ds_ohlcv, columns=["open", "high", "low", "close", "change", "change_ratio", "volume"],
                          index=ds_ohlcv["date"])

        today = datetime.datetime.today().strftime("%Y%m%d")
        todaystr = str(today)
        con = sqlite3.connect("c:/users/백/" + kospi_kosdaq + "_jisu" + todaystr + ".db")

        df.to_sql(kospi_kosdaq, con, if_exists="replace")
        con.close()
        print("지수데이터 다운로드 완료")

    def DS_save_ohlc_(self, start_date, end_date):  #업종, 지수 정보 받기

        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")

        con = sqlite3.connect("c:/users/백/DS_JISU_codedata.db")
        codedata = pd.read_sql("SELECT * FROM kosdaq_JISU", con, index_col="index")

        code_list = codedata["code"]
        name_list = codedata["name"]

        for i, code in enumerate(code_list):
            name = name_list[i]

            str_1 = str(i+1)
            str_2 = str(len(code_list))
            code = "U" + str(code)

            instStockChart.SetInputValue(0, code)
            instStockChart.SetInputValue(1, ord("1"))
            instStockChart.SetInputValue(2, end_date)
            instStockChart.SetInputValue(3, start_date)
            instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))
            instStockChart.SetInputValue(6, ord("D"))
            instStockChart.SetInputValue(9, ord("1"))

            # BlockRequest
            instStockChart.BlockRequest()

            time.sleep(0.25)    #1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

            # GetHeaderValue
            numdata = instStockChart.GetHeaderValue(3)
            numfield = instStockChart.GetHeaderValue(1)

            # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
            date_list = []
            open_list = []
            high_list = []
            low_list = []
            close_list = []
            change_list = []
            volume_list = []
            change_ratio_list = []

            for j in range(numdata):
                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)

            for j in range(numdata):
                open = instStockChart.GetDataValue(1, j)
                open_list.append(open)

            for j in range(numdata):
                high = instStockChart.GetDataValue(2, j)
                high_list.append(high)

            for j in range(numdata):
                low = instStockChart.GetDataValue(3, j)
                low_list.append(low)

            for j in range(numdata):
                close = instStockChart.GetDataValue(4, j)
                close_list.append(close)

            for j in range(numdata):
                change = instStockChart.GetDataValue(5,j)
                change_list.append(change)

            for j in range(numdata):
                volume = instStockChart.GetDataValue(6, j)
                volume_list.append(volume)

            for j in range(numdata):
                if j == numdata-1:
                    change_ratio_list.append(0)

                    break
                change = change_list[j]
                close = close_list[j+1]
                change_ratio = round(change / close, 5)

                change_ratio_list.append(change_ratio)


            ds_ohlcv = {"date": date_list, "open": open_list, "high": high_list, "low": low_list, "close": close_list,
                        "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv, columns=["open", "high", "low", "close", "change", "volume", "change_ratio"],
                              index=ds_ohlcv["date"])


            today = datetime.datetime.today().strftime("%Y%m%d")
            # todaystr = str(today)
            con = sqlite3.connect("c:/users/백/stock_kosdaq.db")
            # con = sqlite3.connect("c:/users/백/DS_kosdaq_" + todaystr + ".db")
            df.to_sql(name, con, if_exists="replace")


            lastmsg_here_def = str_1 + "/" + str_2 + name + " 완료"
            print(lastmsg_here_def)
            #self.statusbar.showMessage(lastmsg_here_def)


    def DS_offline_code_load(self):  #파일로부터 코스닥 코드 정보 로딩
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 종목 데이터 베이스
        jum = "'"
        inputstr = "SELECT * FROM " + jum + "kosdaq" + jum
        df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
        return df


    def check_speedy_rising_volume_2(self, name, low_X, high_X):  #10일 거래량 이평 대비 거래량 X배~X배상승
        df = self.load_stock_data(name)

        vol_list = df["volume"]
        vol_MA10_list = df["vol_MA10"]
        vol_ratio_list = vol_list / vol_MA10_list

        C_speedy_rising_volume_list = []  # C = Condition(조건)


        for i, vol_ratio in enumerate(vol_ratio_list):

            if i > 0:   #다음날 바로 매도 하기때문에 0은 제외
                if high_X > vol_ratio > low_X:
                    C_speedy_rising_volume_list.append(1)
                else:
                    C_speedy_rising_volume_list.append(0)
            else:
                C_speedy_rising_volume_list.append(0)

        df["C_speedy_riging_volume"] = C_speedy_rising_volume_list
        print(df)

        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
        # con = sqlite3.connect("c:/users/백/DS_kosdaq_" + todaystr + ".db")
        df.to_sql(name, con, if_exists="replace")


    def simul_data_to_excel(self, data, file_name = "test__"):
        # 전략 시뮬레이션 결과를 데이터를 엑셀로 저장, 받은 데이터는 반드시 Dataframe형식이고 "date"열을 포함해야함

        wb = Workbook()
        ws = wb.create_sheet()
        #sheet2_name = "수익률 분석"
        #ws1 = wb.create_sheet(sheet2_name)

        # Dataframe을 엑셀로 뿌린다.
        for row in dataframe_to_rows(data, index=True, header=True):

            if len(row) > 1:
                ws.append(row)

        #date_list = data["date"]
        #row_range = str(len(date_list)+1)  # 엑셀 데이터의 행 개수

        # 날짜 중복 제거
        #sdate_list = date_list.unique()
        #sdate_list = pd.DataFrame(sdate_list)

        #for row in dataframe_to_rows(sdate_list, index=True, header=True):
         #   if len(row) > 1:
          #      ws1.append(row)

        wb.save("c:/users/백/" + file_name + ".xlsx")


    def update_buy_list(self, buy_code_list, buy_name_list):   #살종목 텍스트 파일 업데이트
        day_money = 10000000 #1일 최대 투자금액 1000만원
        one_stock_money = int(day_money / self.buy_list_len)
        print("1일 투자금액: " , one_stock_money)

        f = open("buy_list.txt", "wt")
        for i, code in enumerate(buy_code_list):
            name = buy_name_list[i]
            close = self.buy_close_list[i]  #살종목의 오늘 종가

            if len(code) > 6:   ##대신증권 코드에서 키움증권코드로 고치기(맨앞에 A를 제거)
                code = code[1:7]
            buy_quantity = int(one_stock_money / close)   #매수 수량
            f.writelines("%s;매수;%s;시장가;%s;0;매수전\n" %(name, code, buy_quantity))
        f.close()

    def Seek_Today_buy_list(self):
        code_data = self.DS_offline_code_load()

        name_list = code_data["name"]
        code_list = code_data["code"]
        today__ = datetime.datetime.today()  # " 오늘"
        today = today__.strftime("%Y%m%d")
        print("오늘 : ", today)
        #today = 20190315   #나중에 자동으로 되도록 업데이트 할 것
        buy_list = []
        buy_name_list = []

        self.buy_close_list = []  #살종목의 종가

        for i, name in enumerate(name_list):
            print(i, "/", len(name_list), name)
            if self.check_speedy_rising_volume(name, today):
                code = code_list[i]
                buy_name_list.append(name)
                buy_list.append(code)

                self.buy_close_list.append(self.today_close)  # 오늘종가 기록

        print(buy_name_list)
        self.buy_list_len = len(buy_list)    #살 종목 수
        if self.buy_list_len == 0:
            print("살종목이 없습니다.")
            return False
        self.update_buy_list(buy_list, buy_name_list)   #살종목 텍스트 파일 업데이트


    def cal_change(self, codename, day):
        df = self.load_stock_data(codename)

        self.kosdaq_open = df["open"]
        self.kosdaq_close = df["close"]
        self.kosdaq_change = self.rev_kosdaq_vol.rolling(window=day).mean()
        self.ma = self.rev_ma[::-1]
        str_day = str(day)

        self.stock_ohlcv_kosdaq.insert(len(self.stock_ohlcv_kosdaq.columns), "vol_MA" + str_day, self.ma)

        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
        self.stock_ohlcv_kosdaq.to_sql(codename, con, if_exists="replace")

    def test__(self):

        # 현대차(005380) 2018년도부터 현재까지 주가 정보를 가져온다.
        df = fdr.DataReader('005380', '2010')

        print(df.shape)
        print(df.head)


    def load_stock_data(self, codename):   #특정 한종목의 데이터 불러오기, 코드명이 아니고 종목명이다.
    # 종목 정보 불러오기
        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
        jum = "'"
        inputstr = "SELECT * FROM " + jum + codename + jum

        self.stock_ohlcv_kosdaq = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.

        return self.stock_ohlcv_kosdaq

    def calc_ma(self, codename, day):  #거래량이평 구하기
        df = self.load_stock_data(codename)

        self.kosdaq_vol = df["volume"]
        self.rev_kosdaq_vol = self.kosdaq_vol[::-1]
        self.rev_ma = self.rev_kosdaq_vol.rolling(window=day).mean()
        self.ma = self.rev_ma[::-1]
        str_day = str(day)

        #self.stock_ohlcv_kosdaq.insert(len(self.stock_ohlcv_kosdaq.columns), "vol_MA" + str_day, self.ma)
        df.insert(len(self.stock_ohlcv_kosdaq.columns), "vol_MA" + str_day, self.ma)

        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
        self.stock_ohlcv_kosdaq.to_sql(codename, con, if_exists="replace")

    def calc_close_ma_x(self, codename,day):
        df = self.load_stock_data(codename)
        str_day = str(day)
        if "MA" + str_day not in df.columns:
            close_list = df["close"]
            rev_close_list = close_list[::-1]
            rev_close_ma_list = rev_close_list.rolling(window=day).mean()
            close_ma_list = rev_close_ma_list[::-1]
            df.insert(len(df.columns), "MA" + str_day, close_ma_list)

        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
        self.stock_ohlcv_kosdaq.to_sql(codename, con, if_exists="replace")

    def calc_close_ma3_5_10_20_60_120(self, codename):
        ma_x_list = [3,5,10,20,60,120]
        for ma_x in ma_x_list:
            self.calc_close_ma_x(codename, ma_x)


    def calc_vol_ma10_ratio(self, name):
        df = self.load_stock_data(name)
        vol_list = df["volume"]
        vol_ma10_list = df["vol_MA10"]
        vol_ma10_ratio_list = vol_list/vol_ma10_list

        return vol_ma10_ratio_list


    def today_close_buy_tomorrow_close_sell(self, name): #익일 종가 팔기 전략 수익 칼럼 추가
        df = self.load_stock_data(name)
        close_list = df["close"]
        change_list = df["change"]
        change_list_list = list(change_list)
        change_list_list.pop(-1)
        profit_list = change_list_list
        profit_list.insert(0, 0)
        change_index_list = list(change_list.index)

        profit_series = Series(profit_list, index=change_index_list)
        profit_ratio_list = profit_series / close_list

        df1 = {"profit": profit_series, "profit_ratio": profit_ratio_list}
        df2 = DataFrame(df1, index= change_index_list)

        return df2

    #10이평 비율로 매매 전략에 대한 수익률 검증

    def profit_verification(self, name, ref_vol_ratio, ref_vol_ratio2):    #한종목 한Vol_ratio에 대한 검증
        df = self.load_stock_data(name)
        D1_profit_list = list(df["D+1_profit"])
        D1_profit_ratio_list = list(df["D+1_profit_ratio"])
        vol_ratio_list = list(df["vol_ma10_ratio"])
        date_list = list(df["vol_ma10_ratio"].index)

        result_date_list = []
        result_name_list = []
        result_d1_profit_list = []
        result_d1_profit_ratio_list = []

        for i, vol_ratio in enumerate(vol_ratio_list):

            if vol_ratio:   #vol_ratio가 없는 종목을 피하기 위한 코드

                if ref_vol_ratio <= vol_ratio < ref_vol_ratio2:
                    date = date_list[i]
                    result_date_list.append(date)
                    d1_profit = D1_profit_list[i]
                    result_name_list.append(name)
                    result_d1_profit_list.append(d1_profit)
                    d1_profit_ratio = D1_profit_ratio_list[i]
                    result_d1_profit_ratio_list.append(d1_profit_ratio)

        #df1 = {"date" : result_date_list, "name" : result_name_list, "d1_profit" : result_d1_profit_list}
        df1 = {"name": result_name_list, "d1_profit": result_d1_profit_list, "d1_profit_ratio" : result_d1_profit_ratio_list}
        df2 = DataFrame(df1, index= result_date_list)

        return df2

    def profit_verification_all_stock(self, ref_vol_ratio, ref_vol_ratio2): #한 vol_ratio에 대한 모든 종목 검증
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장
        date_list = []
        df1 = {"name": [], "d1_profit": []}
        df2 = DataFrame(df1, index= date_list)

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(i, len(name_list), name)
            df3 = self.profit_verification(name, ref_vol_ratio, ref_vol_ratio2)
            df2 = df2.append(df3)   #df2 를 df3와 합친다.

        df2.sort_index(inplace= True, ascending= False)  #데이터를 인덱스 기준으로 정렬한다. asecending = False <-역순정렬
        print(df2)
        str_ref_vol_ratio = str(ref_vol_ratio)

        str_ref_vol_ratio2 = str(ref_vol_ratio2)
        sheetname = str_ref_vol_ratio + "~" + str_ref_vol_ratio2
        file_name = "거래량 비중 투자 결과(" + sheetname + ").xlsx"

        con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ratio_profit.db")
        df2.to_sql(sheetname , con, if_exists="replace")
        df2.to_excel(file_name, sheet_name= sheetname)   #엑셀로 옮기기


    def add_bong_size(self, name):  #당일 시가 기준 봉사이즈
        df = self.load_stock_data(name)
        open_list = df["open"]
        close_list = df["close"]
        bong_size_list = (close_list - open_list) / open_list

        return bong_size_list

    def add_ma_converge(self, name):  #c_ratio는 이평최대값과 최소값 차의 비율, #이평이 수렴할 때 매입한다.
        df = self.load_stock_data(name)
        ma3_list = list(df["MA3"])
        ma5_list = list(df["MA5"])
        ma10_list = list(df["MA10"])
        ma20_list = list(df["MA20"])
        ma60_list = list(df["MA60"])
        close_list = list(df["close"])
        diff_ma_ratio_list = []

        for i in range(len(ma3_list)):
            ma3 = ma3_list[i]
            ma5 = ma5_list[i]
            ma10 = ma10_list[i]
            ma20 = ma20_list[i]
            ma60 = ma60_list[i]
            close = close_list[i]
            if ma60 == None:
                diff_ma_ratio_list.append(0)
            else:
                diff_ma = max(ma3, ma5, ma10, ma20, ma60) - min(ma3, ma5, ma10, ma20, ma60)
                diff_ma_ratio = diff_ma / close
                diff_ma_ratio_list.append(diff_ma_ratio)

        return diff_ma_ratio_list

    def add_bolband(self):
        df = self.load_stock_data(name)
        close_list = list(df["close"])
        rev_close_list = close_list[::-1]
        df_rev = {
            "close": rev_close_list
        }
        df_rev2 = DataFrame(df_rev)

        #upper, middle, lower = ta.BBANDS(df_rev2["close"], timeperiod=20,nbdevup=2, nbdevdn=2 )
        upper, middle, lower = ta.BBANDS(df["close"], timeperiod=20,nbdevup=2, nbdevdn=2 )

        print(upper)

    def add_gap_size_ratio(self, name):
        df = self.load_stock_data(name)
        open_list = list(df["open"])
        close_list = list(df["close"])
        gap_size_ratio_list = []

        i = 0

        while i < len(close_list)-1:  #마지막 데이터는 갭이 확인 불가하므로 스킵

            d0_open = open_list[i]  # 당일 종가

            d0_close = close_list[i]  # 당일 종가

            d_1_open = open_list[i+1]  #전일 시가f

            d_1_close = close_list[i+1]  #전일 종가

            d_1_max_open_close = max(d_1_open, d_1_close)   #전일 시가, 종가 중 큰값
            d_1_min_open_close = min(d_1_open, d_1_close)   #전일 시가, 종가 중 작은값

            d0_max_open_close = max(d0_open, d0_close)  # 당일 시가, 종가 중 큰값
            d0_min_open_close = min(d0_open, d0_close)  # 당일 시가, 종가 중 작은값

            if d0_min_open_close > d_1_max_open_close:   #상승갭이 있다 = 금일 시가 종가 중 작은값이 전일 시가 종가 중 큰 값보다 크다
                gap_size = d0_min_open_close - d_1_max_open_close
                gap_size_ratio = gap_size / d_1_max_open_close    #전일기준 갭사이즈 비율
                gap_size_ratio_list.append(gap_size_ratio)


            elif d0_max_open_close < d_1_min_open_close:    #하락갭이 있다 = 금일 시가 종가 중 큰값이 전일 시가 종가 중 작은 값보다 작다
                gap_size = d0_max_open_close - d_1_min_open_close
                gap_size_ratio = gap_size / d_1_min_open_close    #전일기준 갭사이즈 비율
                gap_size_ratio_list.append(gap_size_ratio)

            else:
                gap_size_ratio_list.append(0)

            i = i + 1
        gap_size_ratio_list.append(0)  #마지막은 갭이 확인불가하므로 0을 입력

        return gap_size_ratio_list

    #def select_data_date(self, name, start_date, end_date):
     #   df = self.load_stock_data(name)


    def verify_BONG_size(self, name, low_x, high_x):

        df = self.load_stock_data(name)
        close_list = df["close"]
        open_list = df["open"]
        change_ratio_list = df["change_ratio"]
        BONG_size_list = (close_list - open_list) / open_list

        date_list = []
        codename_list = []
        d0_change_ratio_list = []
        d0_close_list = []
        d1_close_list = []
        result_BONG_size_list = []
        profit_list = []
        profit_rate_list = []


        for i, BONG_size in enumerate(BONG_size_list):

            if i > 0:  # 다음날 바로 매도 하기때문에 0은 제외

                if low_x < BONG_size < high_x:

                    date = df.index[i]
                    date_list.append(date)
                    codename_list.append(name)
                    result_BONG_size_list.append(BONG_size)

                    d0_close = list(close_list)[i]  # 당일 종가
                    d0_close_list.append(d0_close)
                    d0_change_ratio = list(change_ratio_list)[i]  # 당일 등락률
                    d0_change_ratio_list.append(d0_change_ratio)
                    d1_close = list(close_list)[i - 1]  # 익일 종가(파는날)
                    d1_close_list.append(d1_close)  # 그다음날 종가(파는날)

                    profit = d1_close - d0_close # 다음날 종가에 팔경우 수익
                    profit_list.append(profit)   #수익

                    profit_rate = profit / d0_close    #수익률
                    profit_rate_list.append(profit_rate)


        result = {"date": date_list, "codename": codename_list, "D+0_close": d0_close_list, "D+0_change_ratio" : d0_change_ratio_list,
        "D+1_close": d1_close_list, "BONG_size_ratio": result_BONG_size_list, "profit": profit_list, "profit_rate": profit_rate_list}

        return result


    def verify_vMA10(self, name, low_X, high_X):  #10일 거래량 이평 대비 거래량 X배~X배상승

        df = self.load_stock_data(name)
        if len(df["open"]) > 0:
            open_list = df["open"]
            high_list = df["high"]
            low_list = df["low"]
            close_list = df["close"]
            change_ratio_list = df["change_ratio"]

            gap_size_list = df["gap_size"]
            vol_ma10_ratio_list = df["vol_ma10_ratio"]

            ma3_list = df["MA3"]
            ma5_list = df["MA5"]
            ma10_list = df["MA10"]
            ma20_list = df["MA20"]
            ma60_list = df["MA60"]
            ma120_list = df["MA120"]

            close_ma3_ratio_list = (close_list - ma3_list) / close_list
            close_ma5_ratio_list = (close_list - ma5_list) / close_list
            close_ma10_ratio_list = (close_list - ma10_list) / close_list
            close_ma20_ratio_list = (close_list - ma20_list) / close_list
            close_ma60_ratio_list = (close_list - ma60_list) / close_list

            close_ma120_ratio_list = (close_list - ma120_list) / close_list

            ma_converge_ratio_list = df["ma_converge_ratio"]

            date_list = []
            codename_list = []
            d0_open_list = []
            d0_high_list = []
            d0_low_list = []
            d0_close_list = []
            d0_change_ratio_list = []

            d0_vol_ma10_ratio_list = []
            d0_gap_size_list = []
            d0_close_ma3_ratio_list = []
            d0_close_ma5_ratio_list = []
            d0_close_ma10_ratio_list = []
            d0_close_ma20_ratio_list = []
            d0_close_ma60_ratio_list = []
            d0_close_ma120_ratio_list = []
            d0_ma_converge_ratio_list = []

            d1_high_list = []
            d1_low_list = []
            d1_close_list = []

            profit_list = []
            profit_rate_list = []

            for i, vol_ma10_ratio in enumerate(vol_ma10_ratio_list):

                if i > 0:   #다음날 바로 매도 하기때문에 0은 제외
                    if vol_ma10_ratio != None:

                        if low_X <= vol_ma10_ratio < high_X:
                            date = df.index[i]
                            date_list.append(date)
                            codename_list.append(name)

                            d0_open = list(open_list)[i]  # 그날 시가(사는날)
                            d0_open_list.append(d0_open)

                            d0_high = list(high_list)[i]
                            d0_high_list.append(d0_high)
                            d0_low = list(low_list)[i]
                            d0_low_list.append(d0_low)

                            d0_close = list(close_list)[i]
                            d0_close_list.append(d0_close)  # 그날 종가(사는날)
                            d0_change_ratio = list(change_ratio_list)[i]  #사는날 등락률
                            d0_change_ratio_list.append(d0_change_ratio)

                            d0_gap_size = list(gap_size_list)[i]
                            d0_gap_size_list.append(d0_gap_size)
                            d0_vol_ma10_ratio_list.append(vol_ma10_ratio)

                            d0_close_ma3_ratio = list(close_ma3_ratio_list)[i]
                            d0_close_ma3_ratio_list.append(d0_close_ma3_ratio)
                            d0_close_ma5_ratio = list(close_ma5_ratio_list)[i]
                            d0_close_ma5_ratio_list.append(d0_close_ma5_ratio)
                            d0_close_ma10_ratio = list(close_ma10_ratio_list)[i]
                            d0_close_ma10_ratio_list.append(d0_close_ma10_ratio)
                            d0_close_ma20_ratio = list(close_ma20_ratio_list)[i]
                            d0_close_ma20_ratio_list.append(d0_close_ma20_ratio)
                            d0_close_ma60_ratio = list(close_ma60_ratio_list)[i]
                            d0_close_ma60_ratio_list.append(d0_close_ma60_ratio)
                            d0_close_ma120_ratio = list(close_ma120_ratio_list)[i]
                            d0_close_ma120_ratio_list.append(d0_close_ma120_ratio)
                            d0_ma_converge_ratio = list(ma_converge_ratio_list)[i]
                            d0_ma_converge_ratio_list.append(d0_ma_converge_ratio)

                            d1_high = list(high_list)[i - 1]
                            d1_high_list.append(d1_high)
                            d1_low = list(low_list)[i - 1]
                            d1_low_list.append(d1_low)
                            d1_close = list(close_list)[i - 1]
                            d1_close_list.append(d1_close)

                            profit = d1_close - d0_close
                            profit_list.append(profit)

                            profit_rate = profit / d0_close
                            profit_rate_list.append(profit_rate)

            result = {"date": date_list, "codename": codename_list, "D+0_open" : d0_open_list,
                      "D+0_high": d0_high_list, "D+0_low" : d0_low_list,
                      "D+0_close": d0_close_list, "D+0_change_ratio" : d0_change_ratio_list,
                      "D+1_high": d1_high_list,"D+1_low": d1_low_list, "D+1_close": d1_close_list,
                      "gap_size_ratio": d0_gap_size_list, "vol_ma10_ratio": d0_vol_ma10_ratio_list,
                      "close/ma3 ratio": d0_close_ma3_ratio_list, "close/ma5 ratio": d0_close_ma5_ratio_list,
                      "close/ma10 ratio": d0_close_ma10_ratio_list, "close/ma20 ratio": d0_close_ma20_ratio_list,
                      "close/ma60 ratio": d0_close_ma60_ratio_list, "close/ma120 ratio": d0_close_ma120_ratio_list,
                      "ma_converge_ratio" : d0_ma_converge_ratio_list,
                      "profit": profit_list, "profit_rate" : profit_rate_list}

        return result

    def verify_change_ratio(self, name, low_X, high_X):

        df = self.load_stock_data(name)
        close_list = df["close"]
        open_list = df["open"]
        change_ratio_list = df["change_ratio"]

        gap_size_list = df["gap_size"]
        vol_ma10_ratio_list = df["vol_ma10_ratio"]

        ma3_list = df["MA3"]
        ma5_list = df["MA5"]
        ma10_list = df["MA10"]
        ma20_list = df["MA20"]
        ma60_list = df["MA60"]
        ma120_list = df["MA120"]
        close_ma3_ratio_list = (close_list - ma3_list) / close_list
        close_ma5_ratio_list = (close_list - ma5_list) / close_list
        close_ma10_ratio_list = (close_list - ma10_list) / close_list
        close_ma20_ratio_list = (close_list - ma20_list) / close_list
        close_ma60_ratio_list = (close_list - ma60_list) / close_list
        close_ma120_ratio_list = (close_list - ma120_list) / close_list
        ma_cvg_ratio_list = df["ma_converge_ratio"]

        date_list = []
        codename_list = []
        d0_open_list = []
        d0_close_list = []
        d0_change_ratio_list = []

        d0_vol_ma10_ratio_list = []
        d0_gap_size_list = []
        d0_close_ma3_ratio_list = []
        d0_close_ma5_ratio_list = []
        d0_close_ma10_ratio_list = []
        d0_close_ma20_ratio_list = []
        d0_close_ma60_ratio_list = []
        d0_close_ma120_ratio_list = []
        d0_ma_cvg_ratio_list = []

        d1_close_list = []
        profit_list = []
        profit_rate_list = []

        for i, change_ratio in enumerate(change_ratio_list):

            vol_ma10_ratio = list(vol_ma10_ratio_list)[i]

            if i > 0:   #다음날 바로 매도 하기때문에 0은 제외
                if vol_ma10_ratio != None:

                    if low_X <= change_ratio < high_X:

                        date = df.index[i]
                        date_list.append(date)
                        codename_list.append(name)
                        d0_open = list(open_list)[i]  # 그날 시가(사는날)
                        d0_open_list.append(d0_open)

                        d0_close = list(close_list)[i]
                        d0_close_list.append(d0_close)  # 그날 종가(사는날)
                        d0_change_ratio = list(change_ratio_list)[i]  #사는날 등락률
                        d0_change_ratio_list.append(d0_change_ratio)

                        d0_gap_size = list(gap_size_list)[i]
                        d0_gap_size_list.append(d0_gap_size)
                        d0_vol_ma10_ratio = list(vol_ma10_ratio_list)[i]
                        d0_vol_ma10_ratio_list.append(d0_vol_ma10_ratio)

                        d0_close_ma3_ratio = list(close_ma3_ratio_list)[i]
                        d0_close_ma3_ratio_list.append(d0_close_ma3_ratio)
                        d0_close_ma5_ratio = list(close_ma5_ratio_list)[i]
                        d0_close_ma5_ratio_list.append(d0_close_ma5_ratio)
                        d0_close_ma10_ratio = list(close_ma10_ratio_list)[i]
                        d0_close_ma10_ratio_list.append(d0_close_ma10_ratio)
                        d0_close_ma20_ratio = list(close_ma20_ratio_list)[i]
                        d0_close_ma20_ratio_list.append(d0_close_ma20_ratio)
                        d0_close_ma60_ratio = list(close_ma60_ratio_list)[i]
                        d0_close_ma60_ratio_list.append(d0_close_ma60_ratio)
                        d0_close_ma120_ratio = list(close_ma120_ratio_list)[i]
                        d0_close_ma120_ratio_list.append(d0_close_ma120_ratio)

                        d0_ma_cvg_ratio = list(ma_cvg_ratio_list)[i]
                        d0_ma_cvg_ratio_list.append(d0_ma_cvg_ratio)

                        d1_close = list(close_list)[i - 1]
                        d1_close_list.append(d1_close)  # 그다음날 종가(파는날)

                        profit = d1_close - d0_close
                        profit_list.append(profit)

                        profit_rate = profit / d0_close
                        profit_rate_list.append(profit_rate)

        result = {"date": date_list, "codename": codename_list, "D+0_open" : d0_open_list,
                  "D+0_close": d0_close_list, "D+0_change_ratio" : d0_change_ratio_list, "D+1_close": d1_close_list,
                  "gap_size_ratio": d0_gap_size_list, "vol_ma10_ratio": d0_vol_ma10_ratio_list,
                  "close/ma3 ratio": d0_close_ma3_ratio_list, "close/ma5 ratio": d0_close_ma5_ratio_list,
                  "close/ma10 ratio": d0_close_ma10_ratio_list, "close/ma20 ratio": d0_close_ma20_ratio_list,
                  "close/ma60 ratio": d0_close_ma60_ratio_list, "close/ma120 ratio": d0_close_ma120_ratio_list,
                  "ma_cvg_ratio": d0_ma_cvg_ratio_list,
                  "profit": profit_list, "profit_rate" : profit_rate_list}

        return result


    def verify_gap_size(self, name, low_x, high_x):

        df = self.load_stock_data(name)
        close_list = df["close"]
        open_list = df["open"]

        change_ratio_list = df["change_ratio"]
        gap_size_list = df["gap_size"]
        vol_ma10_ratio_list = df["vol_ma10_ratio"]

        date_list = []
        codename_list = []
        d0_open_list = []
        d0_close_list = []
        d0_change_ratio_list = []
        d1_close_list = []

        d0_vol_ma10_ratio_list = []
        result_gap_size_list = []
        profit_list = []
        profit_rate_list = []

        for i, gap_size in enumerate(gap_size_list):

            if i > 0:  # 다음날 바로 매도 하기때문에 0은 제외

                if low_x < gap_size < high_x:

                    result_gap_size_list.append(gap_size)
                    date = df.index[i]
                    date_list.append(date)
                    codename_list.append(name)
                    d0_open = list(open_list)[i]
                    d0_open_list.append(d0_open)

                    vol_ma10_ratio = list(vol_ma10_ratio_list)[i]
                    d0_vol_ma10_ratio_list.append(vol_ma10_ratio)

                    d0_close = list(close_list)[i]  # 당일 종가
                    d0_close_list.append(d0_close)
                    d0_change_ratio = list(change_ratio_list)[i]  # 당일 등락률
                    d0_change_ratio_list.append(d0_change_ratio)

                    d1_close = list(close_list)[i - 1]  # 익일 종가(파는날)
                    d1_close_list.append(d1_close)  # 그다음날 종가(파는날)

                    profit = d1_close - d0_close # 다음날 종가에 팔경우 수익
                    profit_list.append(profit)   #수익

                    profit_rate = profit / d0_close    #수익률
                    profit_rate_list.append(profit_rate)


        result = {"date": date_list, "codename": codename_list, "D+0_open" : d0_open_list,
                  "D+0_close": d0_close_list, "D+0_change_ratio" : d0_change_ratio_list,
        "D+1_close": d1_close_list, "gap_size_ratio": result_gap_size_list, "vol_ma10_ratio": d0_vol_ma10_ratio_list,
                  "profit": profit_list, "profit_rate": profit_rate_list}


        return result

    def trace_d1_change(self, name, low_x, high_x):   #low, high는 수익률

        df = self.load_stock_data(name)
        close_list = df["close"]
        open_list = df["open"]

        change_ratio_list = df["change_ratio"]
        gap_size_list = df["gap_size"]
        vol_ma10_ratio_list = df["vol_ma10_ratio"]
        ma3_list = df["MA3"]
        ma5_list = df["MA5"]
        ma10_list = df["MA10"]
        ma20_list = df["MA20"]
        ma60_list = df["MA60"]
        ma120_list = df["MA120"]
        close_ma3_ratio_list = (close_list - ma3_list) / close_list
        close_ma5_ratio_list = (close_list - ma5_list) / close_list
        close_ma10_ratio_list = (close_list - ma10_list) / close_list
        close_ma20_ratio_list = (close_list - ma20_list) / close_list
        close_ma60_ratio_list = (close_list - ma60_list) / close_list
        close_ma120_ratio_list = (close_list - ma120_list) / close_list

        date_list = []
        codename_list = []
        d0_open_list = []
        d0_close_list = []
        d0_change_ratio_list = []

        d0_vol_ma10_ratio_list = []
        d0_gap_size_list = []
        d0_close_ma3_ratio_list = []
        d0_close_ma5_ratio_list = []
        d0_close_ma10_ratio_list = []
        d0_close_ma20_ratio_list = []
        d0_close_ma60_ratio_list = []
        d0_close_ma120_ratio_list = []

        d1_close_list = []
        profit_list = []
        profit_rate_list = []

        trace_d0_close_list = []     #조건에 해당하는 매수시점 가격
        trace_d1_close_list = []        #조건에 해당하는 매도시점 가격

        d0_profit_list = []  # 익일 종가에 팔경우 당일 시점 수익
        d0_profit_rate_list = []  # 익일 종가에 팔경우 당일 시점 수익률


        for i, open in enumerate(open_list):

            d0_close = list(close_list)[i]  # 당일 종가
            d0_close_list.append(d0_close)

            d1_close = list(close_list)[i - 1]  # 익일 종가(파는날)
            d1_close_list.append(d1_close)  # 그다음날 종가(파는날)

            if i == 0:
                profit_list = [0]
                profit_rate_list = [0]

            if i > 0:  # 다음날 바로 매도 하기때문에 0은 제외

                profit = d1_close - d0_close  # 다음날 종가에 팔경우 수익
                profit_list.append(profit)  # 수익

                profit_rate = profit / d0_close  # 수익률
                profit_rate_list.append(profit_rate)


                if low_x <= profit_rate < high_x:

                    date = df.index[i]
                    date_list.append(date)
                    codename_list.append(name)

                    d0_open = list(open_list)[i]
                    d0_open_list.append(d0_open)
                    trace_d0_close = list(d0_close_list)[i]
                    trace_d0_close_list.append(trace_d0_close)
                    trace_d1_close = list(d1_close_list)[i]
                    trace_d1_close_list.append(trace_d1_close)

                    d0_change_ratio = list(change_ratio_list)[i]  # 당일 등락률
                    d0_change_ratio_list.append(d0_change_ratio)

                    vol_ma10_ratio = list(vol_ma10_ratio_list)[i]
                    d0_vol_ma10_ratio_list.append(vol_ma10_ratio)
                    d0_gap_size = list(gap_size_list)[i]
                    d0_gap_size_list.append(d0_gap_size)
                    d0_close_ma3_ratio = list(close_ma3_ratio_list)[i]
                    d0_close_ma3_ratio_list.append(d0_close_ma3_ratio)
                    d0_close_ma5_ratio = list(close_ma5_ratio_list)[i]
                    d0_close_ma5_ratio_list.append(d0_close_ma5_ratio)
                    d0_close_ma10_ratio = list(close_ma10_ratio_list)[i]
                    d0_close_ma10_ratio_list.append(d0_close_ma10_ratio)
                    d0_close_ma20_ratio = list(close_ma20_ratio_list)[i]
                    d0_close_ma20_ratio_list.append(d0_close_ma20_ratio)
                    d0_close_ma60_ratio = list(close_ma60_ratio_list)[i]
                    d0_close_ma60_ratio_list.append(d0_close_ma60_ratio)
                    d0_close_ma120_ratio = list(close_ma120_ratio_list)[i]
                    d0_close_ma120_ratio_list.append(d0_close_ma120_ratio)

                    d0_profit = list(profit_list)[i]
                    d0_profit_list.append(d0_profit)
                    d0_profit_rate = list(profit_rate_list)[i]
                    d0_profit_rate_list.append(d0_profit_rate)


        result = {"date": date_list, "codename": codename_list, "D+0_open": d0_open_list,
                  "D+0_close": trace_d0_close_list, "D+0_change_ratio": d0_change_ratio_list,
                  "D+1_close": trace_d1_close_list, "vol_ma10_ratio": d0_vol_ma10_ratio_list,
                  "gap_size_ratio": d0_gap_size_list,
                  "close/ma3 ratio" : d0_close_ma3_ratio_list, "close/ma5 ratio" : d0_close_ma5_ratio_list,
                  "close/ma10 ratio": d0_close_ma10_ratio_list, "close/ma20 ratio": d0_close_ma20_ratio_list,
                  "close/ma60 ratio": d0_close_ma60_ratio_list, "close/ma120 ratio": d0_close_ma120_ratio_list,
                  "profit": d0_profit_list, "profit_rate": d0_profit_rate_list}

        return result

    def run(self):
        name = "고려제약"
        c_ratio = 0.1
        low_x1 = 3.9
        high_x1 = 4.4
        low_x2 = -0.60
        high_x2 = 0.60
        #ref_vol_ratio = 3.0
        #self.gap_rising_2(name)
        self.check_speedy_rising_volume_2(name, low_x1, high_x1)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dsstock = Daeshin_stock()
    #dsstock.run()
    name = "고려제약"
    a = 20090602
    b = 20181228

    day = 10
    dsstock.DS_save_ohlc_(a,b)

