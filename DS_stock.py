import sys
from pywinauto import application
import win32com.client
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import time
from datetime import datetime, timedelta

from DS_stock_none_GUI import *
from pandas import DataFrame, Series
import sqlite3
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5 import uic
import talib as ta


form_class = uic.loadUiType("DS_stock.ui")[0]

class DS_stock(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.dsstock_none_gui = Daeshin_stock()
        self.setupUi(self)
        self.dstate = 0   # 미접속 상태
        self.timer = QTimer(self)  # 1초마다 시간을 띄우도록
        self.timer.start(1000)  # 1초에 한 번 timeout 시그널이 발생함
        self.timer.timeout.connect(self.timeout)  # timeout시그널 발생시 timeout 함수로 이동
        self.download_ohlc_done = False  #당일 거래량 계산 안됨
        self.lastmsg = ""
        today = datetime.datetime.now()
        todaystring = int(today.strftime('%Y%m%d'))  # 오늘 날짜 yymmdd 형식으로
        # self.spinBox_2.value() = todaystring
        self.spinBox_2.setValue(todaystring)    #spinBox에 값 할당 = setValue(),  lineEdit에 값 할당 = setText()
        self.pushButton.clicked.connect(self.DS_auto_login)  # 접속버튼 클릭시
        self.pushButton_2.clicked.connect(self.DS_kosdaq_code_load)
        self.pushButton_3.clicked.connect(self.DS_save_ohlc_button)
        self.pushButton_4.clicked.connect(self.DS_vol_ma_calc_button)
        self.pushButton_5.clicked.connect(self.Seek_today_buy_list_button)
        self.pushButton_6.clicked.connect(self.verify_vMA10_button)
        self.pushButton_7.clicked.connect(self.add_vol_ratio)
        self.pushButton_8.clicked.connect(self.add_tomorrow_sell_profit)
        self.pushButton_10.clicked.connect(self.add_bong_size_button)
        self.pushButton_15.clicked.connect(self.verify_BONG_size_button)
        self.pushButton_9.clicked.connect(self.verify_change_ratio_button)

        self.pushButton_17.clicked.connect(self.add_close_ma)
        self.pushButton_16.clicked.connect(self.veriyfy_gap_rising_button)
        self.pushButton_18.clicked.connect(self.add_gap_size_ratio_button)
        self.pushButton_19.clicked.connect(self.DS_kosdaq_industry)
        self.pushButton_12.clicked.connect(self.trace_d1_change_button)
        self.pushButton_14.clicked.connect(self.add_ma_converge_button)
        self.pushButton_20.clicked.connect(self.add_ma_converge_button)
        self.pushButton_21.clicked.connect(self.DS_save_today_ohlc)
        self.pushButton_22.clicked.connect(self.add_tech_today)
        self.pushButton_23.clicked.connect(self.execute_query__)
        self.pushButton_24.clicked.connect(self.DS_save_jisu)
        self.pushButton_25.clicked.connect(self.move_to_excel)
        self.pushButton_26.clicked.connect(self.DS_save_ohlc_one_tabel)
        self.pushButton_11.clicked.connect(self.test___)
        self.pushButton_27.clicked.connect(self.DS_marketeye)
        self.pushButton_28.clicked.connect(self.one_stock_ohlc)
        self.pushButton_29.clicked.connect(self.quit_program)

    def timeout(self):
        current_time = QTime.currentTime()
        DS_login_time = QTime(15, 16, 0)
        calc_time = QTime(15, 17, 0)

        text_time = current_time.toString("hh:mm:ss")
        time_msg = "현재시간: "+text_time
        self.dstate = self.DS_connect_confirm()
        if self.dstate == 1:
            state_msg = "대신 서버 연결중"
        else:
            state_msg = "대신 서버 미연결"

        self.statusbar.showMessage(state_msg + " | " + time_msg + " | " + self.lastmsg)

        #자동 로그인 설정
        if current_time > DS_login_time and self.dstate != 1 and self.checkBox_2.isChecked():
            self.DS_auto_login()
            time.sleep(60)

        #자동 살종목 선정 설정
        if current_time > calc_time and self.download_ohlc_done is False and self.dstate == 1 and self.checkBox.isChecked():
            today = datetime.datetime.today()  #" 오늘"
            end_date = today.strftime("%Y%m%d")  #오늘
            start_date = (today - datetime.timedelta(15)).strftime("%Y%m%d")

            self.DS_save_ohlc_button()
            self.statusbar.showMessage("ohlc 다운로드 완료")
            time.sleep(10)
            self.DS_vol_ma_calc_button()
            self.statusbar.showMessage("이평계산 완료")
            time.sleep(10)
            self.Seek_today_buy_list_button()
            self.statusbar.showMessage("살종목 갱신 완료")
            time.sleep(10)

            self.download_ohlc_done = True


    def DS_auto_login(self):  # 대신증권 Cybos plus 자동로그인
        app = application.Application()
        app.start("C:/DAISHIN/STARTER/ncStarter.exe /prj:cp /id:uni9ue /pwd:psyy2758 /pwdcert: /autostart")

    def DS_kosdaq_industry(self):   #코스닥 산업별 코드리스트 저장
        instCpCodeMgr = win32com.client.Dispatch("CpUtil.CpCodeMgr")
        kosdaq_indst_codelist = instCpCodeMgr.GetKosdaqIndustry1List()
        kosdaq_indst_nameist = []
        kospi_indst_codelist = instCpCodeMgr.GetIndustryList()
        kospi_indst_nameist = []

        for code in kospi_indst_codelist:
            kospi_indst_nameist.append(instCpCodeMgr.CodeToName(code))

        for code2 in kosdaq_indst_codelist:
            kosdaq_indst_nameist.append(instCpCodeMgr.CodeToName(code2))

        kospi_indst = {"code" : kospi_indst_codelist, "name": kospi_indst_nameist}
        kosdaq_indst = {"code": kosdaq_indst_codelist, "name": kosdaq_indst_nameist}

        df_kospi_indst = DataFrame(kospi_indst)
        df_kosdaq_indst = DataFrame(kosdaq_indst)

        con = sqlite3.connect("c:/users/백/DS_JISU_codedata.db")
        df_kospi_indst.to_sql("kospi_JISU", con, if_exists="replace")
        df_kosdaq_indst.to_sql("kosdaq_JISU", con, if_exists="replace")
        con.close()
        print("지수코드 다운로드 완료")

    def DS_kosdaq_code_load(self):
        instCpCodeMgr = win32com.client.Dispatch("CpUtil.CpCodeMgr")
        kospi_codelist = instCpCodeMgr.GetStockListByMarket(1)
        kospi_namelist = []
        kosdaq_codelist = instCpCodeMgr.GetStockListByMarket(2)
        kosdaq_namelist =[]

        for code in kospi_codelist:
            kospi_namelist.append(instCpCodeMgr.CodeToName(code))

        kospi_codeNname = {"code" : kospi_codelist, "name": kospi_namelist}

        df_kospi_codeNname = DataFrame(kospi_codeNname)

        #### 엑셀로
        wb = Workbook()
        ws1 = wb.create_sheet("kospi")
        ws2 = wb.create_sheet("kosdaq")

        # Dataframe을 엑셀로 뿌린다.
        for row in dataframe_to_rows(df_kospi_codeNname, index=True, header=True):

            if len(row) > 1:
                ws1.append(row)

        ws1['E1'] = "삭제해야할 데이터"
        ws1['E2'] = "코드가 Q로 시작"
        ws1['E3'] = "KODEX"
        ws1['E4'] = "ARIRANG"
        ws1['E5'] = "TIGER"
        ws1['E6'] = "TIGER"
        ws1['E7'] = "KBSTAR"
        ws1['E8'] = "KBSTAR"
        ws1['E9'] = "KINDEX"
        ws1['E10'] = "HANARO"
        ws1['E11'] = "KOSEF"
        ws1['E12'] = "SMART"
        ws1['E13'] = "FOCUS"
        ws1['E14'] = "마이티"
        ws1['E15'] = "TREX"
        ws1['E16'] = "파워"
        ws1['E17'] = "그외에도 몇개 있음"

        wb.save("c:/users/백/DS_codedata.xlsx")
        print("코스피 코드 엑셀 저장 완료")

        # SQL로
        #con = sqlite3.connect("c:/users/백/DS_codedata.db")
        #df_kospi_codeNname.to_sql("kospi", con, if_exists="replace")

        for code in kosdaq_codelist:
            kosdaq_namelist.append(instCpCodeMgr.CodeToName(code))

        kosdaq_codeNname = {"code" : kosdaq_codelist, "name" : kosdaq_namelist}
        df_kosdaq_codeNname = DataFrame(kosdaq_codeNname)

        # Dataframe을 엑셀로 뿌린다.
        for row in dataframe_to_rows(df_kosdaq_codeNname, index=True, header=True):

            if len(row) > 1:
                ws2.append(row)

        ws2['E1'] = "삭제해야할 데이터"
        ws2['E2'] = "스팩"
        wb.save("c:/users/백/DS_codedata.xlsx")
        print("코스닥 코드 엑셀 저장 완료")


        # SQL로
        #con = sqlite3.connect("c:/users/백/DS_codedata.db")
        #df_kosdaq_codeNname.to_sql("kosdaq", con, if_exists="replace")

        self.lastmsg = "대신증권 코드 갱신 완료"


    def DS_save_ohlc_button(self):

        start_date = self.spinBox.value()
        end_date = self.spinBox_2.value()
        kospi_kosdaq = self.comboBox.currentText()

        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")
        #con = sqlite3.connect("c:/users/백/DS_codedata.db")
        wb = Workbook()
        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")

        sheet = file.get_sheet_by_name(kospi_kosdaq)
        code_list = []
        name_list = []

        for r in sheet.rows:

            code_list.append(r[1].value)
            name_list.append(r[2].value)

        bong__ = self.comboBox_2.currentText()

        if bong__ == "일봉":
            bong = "D"  #일봉
        elif bong__ == "월봉":
            bong = "M"
        else:
            bong = "m"


        for i, code in enumerate(code_list):

            name = name_list[i]
            print(name)
            # code = code[1:]
            # print(code)
            str_1 = str(i+1)
            str_2 = str(len(code_list))

            instStockChart.SetInputValue(0, code)
            if bong == "D": #일봉요청이면
                instStockChart.SetInputValue(1, ord("1"))   # 날짜로 요청
                instStockChart.SetInputValue(2, end_date)  # 요청기간
                instStockChart.SetInputValue(3, start_date)  # 요청기간
            else:
                instStockChart.SetInputValue(1, ord("2"))  # 개수로 요청
                instStockChart.SetInputValue(4, 2220)  # 요청개수 144개월 = 12년 ,30분 1680 = 6개월

            if bong == "m":
                instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  # 1번은 시간
                boon = int(self.comboBox_3.currentText())
                instStockChart.SetInputValue(7, boon)  #틱/분봉주기
            else:
                instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 13, 37))   #1번 시간 제외 #13은 시가 총액

            instStockChart.SetInputValue(6, ord(bong))  #봉 구분
            instStockChart.SetInputValue(9, ord("1"))  #수정 주가

            if self.dstate == 1:
                # BlockRequest
                instStockChart.BlockRequest()
                time.sleep(0.25)  # 1종목 요청 후 X초 쉬고난뒤 다음 종목 요청
            else:
                print("먼저 접속부터 하세요")
                return

            # GetHeaderValue
            numdata = instStockChart.GetHeaderValue(3)
            print(numdata)
            numfield = instStockChart.GetHeaderValue(1)

            # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
            name2_list = []
            date_list = []
            name_date = []
            date_time = []
            time_list = []
            open_list = []
            high_list = []
            low_list = []
            close_list = []
            change_list = []
            volume_list = []
            change_ratio_list = []
            si_chong_list = []   # 시가 총액

            if bong == "m":

                for j in range(numdata):
                    name2_list.append(name)
                    date = instStockChart.GetDataValue(0, j)
                    date_list.append(date)
                    time_ = instStockChart.GetDataValue(1, j)
                    time_list.append(time_)
                    date_time.append(str(date) + ";" + str(time_))

                    open = instStockChart.GetDataValue(2, j)
                    open_list.append(open)
                    high = instStockChart.GetDataValue(3, j)
                    high_list.append(high)
                    low = instStockChart.GetDataValue(4, j)
                    low_list.append(low)
                    close = instStockChart.GetDataValue(5, j)
                    close_list.append(close)
                    change = instStockChart.GetDataValue(6, j)
                    change_list.append(change)
                    volume = instStockChart.GetDataValue(7, j)
                    volume_list.append(volume)
                    change_ratio = round(change / (close - change), 5)
                    change_ratio_list.append(change_ratio)

                ds_ohlcv = {"date_time":date_time, "date": date_list, "name": name2_list, "time": time_list, "open": open_list, "high": high_list, "low": low_list,
                                "close": close_list, "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list}

                df = pd.DataFrame(ds_ohlcv,
                                      columns=["date", "time", "name", "open", "high", "low", "close", "change", "volume", "change_ratio"],
                                      index=ds_ohlcv["date_time"])

            else:
                for j in range(numdata):
                    name2_list.append(name)
                    date = instStockChart.GetDataValue(0, j)
                    date_list.append(date)
                    name_date.append(name + ";" + str(date))

                    open = instStockChart.GetDataValue(1, j)
                    open_list.append(open)
                    high = instStockChart.GetDataValue(2, j)
                    high_list.append(high)
                    low = instStockChart.GetDataValue(3, j)
                    low_list.append(low)
                    close = instStockChart.GetDataValue(4, j)
                    close_list.append(close)
                    change = instStockChart.GetDataValue(5,j)
                    change_list.append(change)
                    volume = instStockChart.GetDataValue(6, j)
                    volume_list.append(volume)
                    si_chong = instStockChart.GetDataValue(7, j)
                    si_chong = si_chong / 100000000
                    si_chong = round(si_chong, 0)
                    si_chong_list.append(si_chong)   # 시가 총액

                    if close-change == 0:
                        change_ratio_list.append(1)
                    else:
                        change_ratio = round(change/(close-change), 5)
                        change_ratio_list.append(change_ratio)


                ds_ohlcv = {"name_date":name_date, "name": name2_list, "date": date_list, "open": open_list, "high": high_list, "low": low_list, "close": close_list,
                        "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list, "si_chong": si_chong_list}

                df = pd.DataFrame(ds_ohlcv, columns=["date", "name", "open", "high", "low", "close", "change", "volume", "change_ratio", "si_chong"],
                              index=ds_ohlcv["name_date"])


            con = sqlite3.connect("c:/users/백/stock_" + kospi_kosdaq + "_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            con.close()


            lastmsg_here_def = str_1 + "/" + str_2 + name + " 완료"
            print(lastmsg_here_def)
            #self.statusbar.showMessage(lastmsg_here_def)

    def DS_marketeye(self):
        instmarketeye = win32com.client.Dispatch("CpSysDib.MarketEye")

        kospi_kosdaq = self.comboBox.currentText()
        file = openpyxl.load_workbook("c:/users/백/DS_codedata.xlsx")
        sheet = file.get_sheet_by_name(kospi_kosdaq)
        codedata2 = []
        name_data = []
        for r in sheet.rows:
            codedata2.append(r[1].value)
            name_data.append(r[2].value)

        # con = sqlite3.connect("c:/users/백/DS_codedata.db")
        # kospi_kosdaq = self.comboBox.currentText()
        # codedata = pd.read_sql("SELECT * FROM " + kospi_kosdaq, con, index_col="index")
        # codedata2 = codedata['code']
        # con.close()
        codedata2 = codedata2[1:]
        name_data = name_data[1:]

        rq_field = [0, 4, 17, 20, 63, 64, 67, 70, 74, 86, 89, 91, 102, 107, 108, 109, 110, 111, 125]

        # 4 : 현재가
        # 17 : 종목명
        # 20 : 총상장주식수
        # 63 : 52주 최고가
        # 64 : 52주 최저가
        # 67 : PER
        # 70 : EPS
        # 74 : 배당수익률
        # 86 : 매출액
        # 89 : 주당순자산
        # 91 : 영업이익(1년)
        # 102 : 분기영업이익
        # 107 : 분기 ROE
        # 108 : 분기이자보상비율
        # 109 : 분기유보율
        # 110 : 분기부채비율
        # 111 : 최근분기년월
        # 125 : EBITDA

        code_list = []
        price_list = []   # 4
        name_list = []   # 17
        jusik_su_list = []   # 20
        high_52week_list = []  # 63
        low_52week_list = []   # 64
        per_Y_list = []    # 67
        EPS_list = []            # 70
        bae_dang_list = []    # 74
        sell_list = []  # 86
        bps_list = []    # 89
        youngub_2ik_Y_list = []   # 91
        youngub_2ik_Q_list = []  # 102
        roe_Q_list = []   # 107
        ija_bosang_R_list = []  # 108
        youbo_Q_list = []   # 109
        buche_Q_list = []    # 110
        recent_month_list = []   #111
        EBITA_list = []  # 125

        for i, code in enumerate(codedata2):

            instmarketeye.SetInputValue(0, rq_field)
            instmarketeye.SetInputValue(1, code)

            # BlockRequest
            instmarketeye.BlockRequest()
            time.sleep(0.25)  # 1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

            # GetHeaderValue
            numdata = instmarketeye.GetHeaderValue(2)
            numfield = instmarketeye.GetHeaderValue(0)

            for j in range(numdata):
                code = instmarketeye.GetDataValue(0, j)
                code_list.append(code)
                price = instmarketeye.GetDataValue(1, j)
                price_list.append(price)
                name = instmarketeye.GetDataValue(2, j)
                print("다운로드 진행 중 %d / %d %s" % (i+1, len(codedata2), name))
                name_list.append(name)
                jusik_su = instmarketeye.GetDataValue(3, j)
                jusik_su_list.append(jusik_su)
                high_52week = instmarketeye.GetDataValue(4, j)
                high_52week_list.append(high_52week)
                low_52week = instmarketeye.GetDataValue(5, j)
                low_52week_list.append(low_52week)
                per = instmarketeye.GetDataValue(6, j)
                per_Y_list.append(per)
                eps = instmarketeye.GetDataValue(7, j)
                EPS_list.append(eps)
                bae_dang = instmarketeye.GetDataValue(8, j)
                bae_dang_list.append(bae_dang)
                sell = instmarketeye.GetDataValue(9, j)
                sell_list.append(sell)
                bps = instmarketeye.GetDataValue(10, j)
                bps_list.append(bps)
                youngub_2ik_Y = instmarketeye.GetDataValue(11, j)
                youngub_2ik_Y_list.append(youngub_2ik_Y)
                youngub_2ik_Q = instmarketeye.GetDataValue(12, j)
                youngub_2ik_Q_list.append(youngub_2ik_Q)
                roe_Q = instmarketeye.GetDataValue(13, j)
                roe_Q_list.append(roe_Q)
                ijabosang = instmarketeye.GetDataValue(14, j)
                ija_bosang_R_list.append(ijabosang)
                youbo_Q = instmarketeye.GetDataValue(15, j)
                youbo_Q_list.append(youbo_Q)
                buche_Q = instmarketeye.GetDataValue(16, j)
                buche_Q_list.append(buche_Q)
                recent_month = instmarketeye.GetDataValue(17, j)
                recent_month_list.append(recent_month)
                EBITA = instmarketeye.GetDataValue(18, j)
                EBITA_list.append(EBITA)

        ds_info = {"code": code_list, "price":price_list, "name": name_list, "jusik_su": jusik_su_list, "per_Y": per_Y_list,
                  "EPS": EPS_list, "high_52week": high_52week_list, "low_52week": low_52week_list, "sell": sell_list, "yungub_profit_Y": youngub_2ik_Y_list,
                   "yungub_profit_Q": youngub_2ik_Q_list, "bae_dang": bae_dang_list, "ija_bosnag": ija_bosang_R_list, "EBITA": EBITA_list,
                   "bps": bps_list, "roe_Q": roe_Q_list, "youbo_Q": youbo_Q_list, "buche": buche_Q_list, "recent_M": recent_month_list}
        df = pd.DataFrame(ds_info)
        print(df)
        self.dsstock_none_gui.simul_data_to_excel(df, "market_eye")
        print("엑셀저장 완료")

    def DS_save_ohlc_one_tabel(self):

        start_date = self.spinBox.value()
        end_date = self.spinBox_2.value()

        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")

        con = sqlite3.connect("c:/users/백/DS_codedata.db")
        kospi_kosdaq = self.comboBox.currentText()
        codedata = pd.read_sql("SELECT * FROM " + kospi_kosdaq, con, index_col="index")

        code_list = codedata["code"]
        name_list = codedata["name"]
        bong1 = "D"  #일봉
        bong2 = "M"  #월봉
        bong3 = "m"  # 분봉
        con.close()

        for i, code in enumerate(code_list):
            name = name_list[i]
            str_1 = str(i+1)
            str_2 = str(len(code_list))

            instStockChart.SetInputValue(0, code)
            instStockChart.SetInputValue(1, ord("1"))
            #instStockChart.SetInputValue(1, ord("2"))  # 개수로 요청
            instStockChart.SetInputValue(2, end_date)  #요청기간
            instStockChart.SetInputValue(3, start_date)  #요청기간
            #instStockChart.SetInputValue(4, 1950)  # 요청개수
            #instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  # 분봉전용
            instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))
            instStockChart.SetInputValue(6, ord(bong1))  #일봉
            #instStockChart.SetInputValue(6, ord(bong3))   #분봉
            #instStockChart.SetInputValue(7, 10)  #틱/분봉주기
            instStockChart.SetInputValue(9, ord("1"))

            # BlockRequest
            instStockChart.BlockRequest()

            time.sleep(0.25)    #1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

            # GetHeaderValue
            numdata = instStockChart.GetHeaderValue(3)
            numfield = instStockChart.GetHeaderValue(1)

            # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
            date_list = []
            open_list = []
            high_list = []
            low_list = []
            close_list = []
            change_list = []
            volume_list = []
            change_ratio_list = []
            index_list = []
            name2_list = []

            for j in range(numdata):
                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)
                open = instStockChart.GetDataValue(1, j)
                open_list.append(open)
                high = instStockChart.GetDataValue(2, j)
                high_list.append(high)
                low = instStockChart.GetDataValue(3, j)
                low_list.append(low)
                close = instStockChart.GetDataValue(4, j)
                close_list.append(close)
                change = instStockChart.GetDataValue(5,j)
                change_list.append(change)
                volume = instStockChart.GetDataValue(6, j)
                volume_list.append(volume)
                index_list.append(str(date) + name)
                name2_list.append(name)
                if j == numdata-1:
                    change_ratio_list.append(0)
                    break
                change = change_list[j]
                change_ratio = round(change / (close-change), 4)
                change_ratio_list.append(change_ratio)

            ds_ohlcv = {"index": index_list, "date": date_list, "name": name2_list, "open": open_list,
                        "high": high_list, "low": low_list,
                        "close": close_list, "change": change_list, "volume": volume_list,
                        "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv, columns=["date", "name", "open", "high", "low", "close", "change", "volume",
                                                 "change_ratio"],index=ds_ohlcv["index"])
            con1 = sqlite3.connect("c:/users/백/stock_" + kospi_kosdaq + "_vol_ma.db")

            #df.to_sql(name, con1, if_exists="replace")
            df.to_sql(kospi_kosdaq, con1, if_exists="append")
            lastmsg_here_def = str_1 + "/" + str_2 + name + " 완료"
            con1.close()
            print(lastmsg_here_def)
            #self.statusbar.showMessage(lastmsg_here_def)


    def DS_save_today_ohlc(self):

        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")

        con = sqlite3.connect("c:/users/백/DS_codedata.db")
        kospi_kosdaq = self.comboBox.currentText()
        codedata = pd.read_sql("SELECT * FROM " + kospi_kosdaq, con, index_col="index")
        con.close()

        code_list = codedata["code"]
        name_list = codedata["name"]
        bong1 = "D"  #일봉


        for i, code in enumerate(code_list):
            name = name_list[i]

            print(name + str(i) + "/" + str(len(name_list)))

            str_1 = str(i+1)
            str_2 = str(len(code_list))

            instStockChart.SetInputValue(0, code)
            #instStockChart.SetInputValue(1, ord("1"))
            instStockChart.SetInputValue(1, ord("2"))  # 개수로 요청
            #instStockChart.SetInputValue(2, end_date)  #요청기간
            #instStockChart.SetInputValue(3, start_date)  #요청기간
            instStockChart.SetInputValue(4, 60)  # 요청개수
            #instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  # 분봉전용
            instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))
            instStockChart.SetInputValue(6, ord(bong1))  #일봉
            #instStockChart.SetInputValue(6, ord(bong3))   #분봉
            #instStockChart.SetInputValue(7, 10)  #틱/분봉주기
            instStockChart.SetInputValue(9, ord("1"))

            # BlockRequest
            instStockChart.BlockRequest()

            time.sleep(0.25)    #1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

            # GetHeaderValue
            numdata = instStockChart.GetHeaderValue(3)
            print(numdata)
            numfield = instStockChart.GetHeaderValue(1)

            # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
            index_list = []
            name2_list = []
            date_list = []
            open_list = []
            high_list = []
            low_list = []
            close_list = []
            change_list = []
            volume_list = []
            change_ratio_list = []

            for j in range(numdata):
                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)
                open = instStockChart.GetDataValue(1, j)
                open_list.append(open)
                high = instStockChart.GetDataValue(2, j)
                high_list.append(high)
                low = instStockChart.GetDataValue(3, j)
                low_list.append(low)
                close = instStockChart.GetDataValue(4, j)
                close_list.append(close)
                change = instStockChart.GetDataValue(5,j)
                change_list.append(change)
                volume = instStockChart.GetDataValue(6, j)
                volume_list.append(volume)
                index_list.append(str(date) + name)
                name2_list.append(name)
                if j == numdata-1:
                    change_ratio_list.append(0)

                    break
                change = change_list[j]
                close = close_list[j+1] #전일 종가
                change_ratio = round(change / close, 5)

                change_ratio_list.append(change_ratio)


            ds_ohlcv = {"index" : index_list, "date": date_list, "name" : name2_list, "open": open_list, "high": high_list, "low": low_list,
                        "close": close_list, "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv, columns=["date", "name", "open", "high", "low", "close", "change", "volume", "change_ratio"],
                             index=ds_ohlcv["index"])

            today = datetime.datetime.today().strftime("%Y%m%d")
            todaystr = str(today)
            con = sqlite3.connect("c:/users/백/" + todaystr + kospi_kosdaq + "_data.db")
            df.to_sql(kospi_kosdaq, con, if_exists='append')
            con.close()

        print("오늘OHCLV 다운완료")

    def add_tech_today(self, name):    #기술적 지표 계산
        kospi_kosdaq = self.comboBox.currentText()
        today = datetime.datetime.today().strftime("%Y%m%d")
        todaystr = str(today)

        con = sqlite3.connect("c:/users/백/" + todaystr + kospi_kosdaq + "_data.db")

        df = pd.read_sql("SELECT * FROM "+ kospi_kosdaq, con, index_col=None) #df = pd.read_sql("SELECT * FROM CMG제약 ", con , index_col = None)
        #df = pd.read_sql("SELECT * FROM " + kospi_kosdaq, con, index_col="index")
        df.sort_index(inplace= True, ascending= False)  #인덱스 기준으로 역으로 정렬

        op = df['open'] * 0.1 * 10
        cl = df['close'] * 0.1 * 10
        hi = df['high'] * 0.1 * 10
        lo = df['low'] * 0.1 * 10
        vo = df['volume'] * 0.1 * 10

        dfma3 = ta.SMA(cl, 3)
        dfma5 = ta.SMA(cl, 5)
        dfma10 = ta.SMA(cl, 10)
        dfma20 = ta.SMA(cl, 20)
        dfma60 = ta.SMA(cl, 60)

        df['vma10_R'] = vo / ta.SMA(vo, 10)

        df['dp_ma3'] = cl / dfma3
        df['dp_ma5'] = cl / dfma5
        df['dp_ma10'] = cl / dfma10
        df['dp_ma20'] = cl / dfma20
        df['dp_ma60'] = cl / dfma60


        bbands20 = pd.Series(ta.BBANDS(cl, timeperiod=20, nbdevup=2, nbdevdn=2))  # 밴드는 또 Series를 만들었다가
        df['bbands20_upR'] = cl/bbands20[0]  # 하나씩 일일이 인덱싱해줘야 함.

        #df['bbands20_mov'] = bbands20[1]
        df['bbands20_downR'] = cl/bbands20[2]
        macd = pd.Series(ta.MACD(cl, 10, 15, 7))  # 밴드는 또 Series를 만들었다가
        df['macd_line'] = macd[0]  # 하나씩 일일이 인덱싱해줘야 함.
        df['macd_sig'] = macd[1]
        df['macd_histo'] = macd[2]

        df['rsi10'] = ta.RSI(cl, 10)
        df['rsi20'] = ta.RSI(cl, 20)
        df['rsi60'] = ta.RSI(cl, 60)
        df.to_sql(kospi_kosdaq, con, if_exists='replace')
        con.close()
        print("최종완료")


    def execute_query__(self):
        kospi_kosdaq = self.comboBox.currentText()

        #today = datetime.datetime.today().strftime("%Y%m%d")
        test = "20200129"
        today = datetime.datetime.today()
        todaystr = str(today.strftime("%Y%m%d"))

        year_ = datetime.datetime.today().year
        month_ = datetime.datetime.today().month
        day_ = datetime.datetime.today().day
        week_day = datetime.date(year_, month_, day_).weekday()   #요일 구하기 5=토, 6=일, 0 = 월
        print(week_day)

        if week_day == 5:
            print("완료")
            today = today + timedelta(days = -1)    #토요일이면 하루전
        elif week_day == 6:
            today = today + timedelta(days=-2)      #일요일이면 이틀전

        modified_todaystr = str(today.strftime("%Y%m%d"))
        print(modified_todaystr)
        #query_ = "change_ratio > 0.04 AND change_ratio < 0.22 AND vma10_R <2 AND bbands20_upR > 0.38 AND bbands20_upR <0.76 AND date == " + modified_todaystr  # 쿼리 입력
        query_ = "change_ratio > 0.04 AND change_ratio < 0.22 AND vma10_R <2 AND bbands20_upR > 0.38 AND bbands20_upR <0.76 AND date == " + test  # 쿼리 입력
        con = sqlite3.connect("c:/users/백/" + todaystr + kospi_kosdaq + "_data.db")  # 키움증권 다운로드 종목 데이터 베이스

        #if self.kospi_kosdaq == "kospi":
         #   df = pd.read_sql("SELECT * FROM 동화약품 WHERE " + query_, con, index_col=None)   #코스피

        #else:
         #   df = pd.read_sql("SELECT * FROM 삼천당제약 WHERE " + query_, con, index_col=None)
        df = pd.read_sql("SELECT * FROM " + kospi_kosdaq + " WHERE " + query_, con, index_col=None)

        con.close()
        self.dsstock_none_gui.simul_data_to_excel(df)
        print("엑셀저장 완료")

    def DS_save_jisu(self):  #지수 데이터
        start_date = self.spinBox.value()
        end_date = self.spinBox_2.value()
        kospi_kosdaq = self.comboBox.currentText()
        bong__ = self.comboBox_2.currentText()
        if bong__ == "일봉":
            bong = "D"  #일봉
        elif bong__ == "월봉":
            bong = "M"
        else:
            bong = "m"
        if self.dstate == 1:
            self.dsstock_none_gui.DS_save_jisu__(start_date, end_date, kospi_kosdaq, bong)
        else:
            print("먼저 접속부터 하세요")

    def DS_vol_ma_calc_button(self):
        day = self.spinBox_3.value()
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기
            self.dsstock_none_gui.calc_ma(codename, day)
            print(i+1, "/", len(codename_list), codename, " 완료")

    def add_close_ma(self):   #가격 이평 추가

        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기
            self.dsstock_none_gui.calc_close_ma3_5_10_20_60_120(codename)
            print(i+1, "/", len(codename_list), codename, " 완료")

    def add_vol_ratio(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            vol_ma10_ratio_list = self.dsstock_none_gui.calc_vol_ma10_ratio(name)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
            jum = "'"
            inputstr = "SELECT * FROM " + jum + name + jum

            df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
            df.insert(len(df.columns), "vol_ma10_ratio", vol_ma10_ratio_list)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")

    def add_bong_size_button(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            bong_size_list = self.dsstock_none_gui.add_bong_size(name)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
            jum = "'"
            inputstr = "SELECT * FROM " + jum + name + jum

            df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
            df.insert(len(df.columns), "bong_size", bong_size_list)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")

    def add_gap_size_ratio_button(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            gap_size_ratio_list = self.dsstock_none_gui.add_gap_size_ratio(name)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
            jum = "'"
            inputstr = "SELECT * FROM " + jum + name + jum

            df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
            df.insert(len(df.columns), "gap_size", gap_size_ratio_list)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")

        print("최종완료")

    def add_ma_converge_button(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            print(name + " 진입")
            diff_ma_ratio_list = self.dsstock_none_gui.add_ma_converge(name)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
            jum = "'"
            inputstr = "SELECT * FROM " + jum + name + jum

            df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.
            df.insert(len(df.columns), "ma_converge_ratio", diff_ma_ratio_list)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")
        print("최종완료")

    def add_bolband(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

    def add_tomorrow_sell_profit(self):  #익일 매도시 수익률 데이터에 추가하기
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권용 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")
        name_list = df["name"]  # 코드네임을 리스트로 저장

        for i, name in enumerate(name_list):  # 뺑뺑이 돌리기
            df1 = self.dsstock_none_gui.today_close_buy_tomorrow_close_sell(name)

            D1_profit_list = df1["profit"]
            D1_profit_ratio_list = df1["profit_ratio"]

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")  # 키움증권 다운로드 종목 데이터 베이스
            jum = "'"
            inputstr = "SELECT * FROM " + jum + name + jum

            df = pd.read_sql(inputstr, con, index_col="index")  # 여기서 데이터형식은 DataFrame 객체다.

            #df.drop("D+1_profit", axis=1, inplace=True)   #column을 삭제하는 가장 좋은 방법은 drop을 사용하는 것입니다.
                                                            #dataframe.drop("컬럼 이름", axis=1, inplace=True)
                                                            #여기서 axis는 축을 의미하고 0은 row를 1은 column을 의미하게 됩니다.
                                                            #inplace=True라고 설정을 해야지 바로 drop의 동작을 실행합니다.
            df.insert(len(df.columns), "D+1_profit", D1_profit_list)

            df.insert(len(df.columns), "D+1_profit_ratio", D1_profit_ratio_list)

            con = sqlite3.connect("c:/users/백/stock_kosdaq_vol_ma.db")
            df.to_sql(name, con, if_exists="replace")
            print(i + 1, "/", len(name_list), name, " 완료")


    def DS_connect_confirm(self):
        instCpCybos = win32com.client.Dispatch("CpUtil.CpCybos")
        return instCpCybos.IsConnect


    def Seek_today_buy_list_button(self):
        self.dsstock_none_gui.Seek_Today_buy_list()

    def verify_BONG_size_button(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권 기준 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기
        all_codename_list = df["name"]  # 코드네임을 리스트로 저장

        low_X = self.doubleSpinBox_9.value()
        high_X = self.doubleSpinBox_10.value()

        final_result_date = []
        final_codename = []
        final_result_d0_close = []
        final_result_d0_change_ratio = []
        final_result_d1_close = []
        final_result_profit = []
        final_result_profit_rate = []
        final_result_BONG_size_ratio = []

        for i, name in enumerate(all_codename_list):  # 뺑뺑이 돌리기

            print(name)

            final_result = self.dsstock_none_gui.verify_BONG_size(name, low_X, high_X)

            #if i == 0:
             #   final_result_date = final_result["date"]
              #  final_codename = final_result["codename"]
               # final_result_d0_close = final_result["D+0_close"]
                #final_result_d0_change_ratio = final_result["D+0_change_ratio"]
                #final_result_d1_close = final_result["D+1_close"]
                #final_result_BONG_size_ratio = final_result["BONG_size_ratio"]
                #final_result_profit = final_result["profit"]
                #final_result_profit_rate = final_result["profit_rate"]


            final_result_date += final_result["date"]
            final_codename += final_result["codename"]
            final_result_d0_close += final_result["D+0_close"]
            final_result_d0_change_ratio += final_result["D+0_change_ratio"]
            final_result_d1_close += final_result["D+1_close"]
            final_result_BONG_size_ratio += final_result["BONG_size_ratio"]
            final_result_profit += final_result["profit"]
            final_result_profit_rate += final_result["profit_rate"]

            print(i + 1, "/", len(all_codename_list), " 완료")

        final_result2 = {"date": final_result_date, "codename": final_codename, "D+0_close": final_result_d0_close,
                         "D+0_change_ratio" : final_result_d0_change_ratio,
                         "D+1_close": final_result_d1_close, "BONG_size_ratio": final_result_BONG_size_ratio,
                         "profit": final_result_profit, "profit_rate": final_result_profit_rate}

        final_result2_df = DataFrame(final_result2)
        sorted_result_df = final_result2_df.sort_values(by="date", ascending=False)

        print("OK2")

        # print(sorted_result_df)
        print(sorted_result_df["codename"])

        # sqlite3로 저장
        # con = sqlite3.connect("c:/users/백/kosdaq_result"+ low_X_str + "~" + high_X_str + ".db")
        # final_result2_df.to_sql("결과", con, if_exists="replace")

        # 엑셀로 결과 저장
        self.dsstock_none_gui.simul_data_to_excel(sorted_result_df)


    def verify_vMA10_button(self):  #거래량 이평 급등종목 시뮬레이션

        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권 기준 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        low_X = self.doubleSpinBox.value()
        high_X = self.doubleSpinBox_2.value()
        low_X_str = str(low_X)
        high_X_str = str(high_X)
        final_result_date = []
        final_codename = []
        final_result_d0_open = []
        final_result_d0_high = []
        final_result_d0_low = []
        final_result_d0_close = []
        final_result_d0_change_ratio = []
        final_result_d1_high = []
        final_result_d1_low = []
        final_result_d1_close = []
        final_result_vol_ma10_ratio = []
        final_result_gap_size_ratio = []
        final_close_ma3_ratio = []
        final_close_ma5_ratio = []
        final_close_ma10_ratio = []
        final_close_ma20_ratio = []
        final_close_ma60_ratio = []
        final_close_ma120_ratio = []
        final_ma_converge_ratio = []
        final_result_profit = []
        final_result_profit_rate = []


        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기

            print(str(codename)+"진입")

            final_result = self.dsstock_none_gui.verify_vMA10(codename, low_X, high_X)

            final_date = list(final_result["date"])
            print(len(final_date))

            if len(final_date) > 0:
                final_result_date += final_result["date"]
                final_codename += final_result["codename"]
                final_result_d0_open += final_result["D+0_open"]
                final_result_d0_high += final_result["D+0_high"]
                final_result_d0_low += final_result["D+0_low"]
                final_result_d0_close += final_result["D+0_close"]
                final_result_d0_change_ratio += final_result["D+0_change_ratio"]
                final_result_d1_high += final_result["D+1_high"]
                final_result_d1_low += final_result["D+1_low"]
                final_result_d1_close += final_result["D+1_close"]
                final_result_gap_size_ratio += final_result["gap_size_ratio"]
                final_result_vol_ma10_ratio += final_result["vol_ma10_ratio"]
                final_close_ma3_ratio += final_result["close/ma3 ratio"]
                final_close_ma5_ratio += final_result["close/ma5 ratio"]
                final_close_ma10_ratio += final_result["close/ma10 ratio"]
                final_close_ma20_ratio += final_result["close/ma20 ratio"]
                final_close_ma60_ratio += final_result["close/ma60 ratio"]
                final_close_ma120_ratio += final_result["close/ma120 ratio"]
                final_ma_converge_ratio += final_result["ma_converge_ratio"]
                final_result_profit += final_result["profit"]
                final_result_profit_rate += final_result["profit_rate"]

            print(i + 1, "/", len(codename_list), " 완료")

        final_result2 = {"date": final_result_date, "codename": final_codename, "D+0_open": final_result_d0_open,
                         "D+0_high": final_result_d0_high, "D+0_low": final_result_d0_low,
                         "D+0_close": final_result_d0_close, "D+0_change_ratio": final_result_d0_change_ratio,
                         "D+1_high": final_result_d1_high, "D+1_low": final_result_d1_low, "D+1_close": final_result_d1_close,
                         "Vol_ratio": final_result_vol_ma10_ratio,
                         "gap_size_ratio" : final_result_gap_size_ratio,
                         "close/ma3 ratio": final_close_ma3_ratio, "close/ma5 ratio": final_close_ma5_ratio,
                         "close/ma10 ratio": final_close_ma10_ratio, "close/ma20 ratio": final_close_ma20_ratio,
                         "close/ma60 ratio": final_close_ma60_ratio, "close/ma120 ratio": final_close_ma120_ratio,
                         "ma_converge_ratio" : final_ma_converge_ratio,
                         "profit": final_result_profit, "profit_rate": final_result_profit_rate}

        print(len(final_result_date))

        final_result2_df = DataFrame(final_result2)

        sorted_result_df = final_result2_df.sort_values(by="date", ascending=False)

        #sqlite3로 저장
        #con = sqlite3.connect("c:/users/백/kosdaq_result"+ low_X_str + "~" + high_X_str + ".db")
        #final_result2_df.to_sql("결과", con, if_exists="replace")

        #엑셀로 결과 저장
        self.dsstock_none_gui.simul_data_to_excel(sorted_result_df)
        print("엑셀저장 완료")


    def verify_change_ratio_button(self):  #등락률 검증 시뮬레이션

        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권 기준 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        low_X = self.doubleSpinBox_13.value()
        high_X = self.doubleSpinBox_14.value()

        final_result_date = []
        final_codename = []
        final_result_d0_open = []
        final_result_d0_close = []
        final_result_d0_change_ratio = []
        final_result_d1_close = []
        final_result_vol_ma10_ratio = []
        final_result_gap_size_ratio = []
        final_close_ma3_ratio = []
        final_close_ma5_ratio = []
        final_close_ma10_ratio = []
        final_close_ma20_ratio = []
        final_close_ma60_ratio = []
        final_close_ma120_ratio = []
        final_ma_converge_ratio = []
        final_result_profit = []
        final_result_profit_rate = []


        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기

            print(str(codename)+"진입")

            final_result = self.dsstock_none_gui.verify_change_ratio(codename, low_X, high_X)  #날짜
            final_date = list(final_result["date"])
            print(len(final_date))

            if len(final_date) > 0:
                final_result_date += final_result["date"]
                final_codename += final_result["codename"]

                final_result_d0_open += final_result["D+0_open"]
                final_result_d0_close += final_result["D+0_close"]
                final_result_d0_change_ratio += final_result["D+0_change_ratio"]
                final_result_d1_close += final_result["D+1_close"]
                final_result_gap_size_ratio += final_result["gap_size_ratio"]
                final_result_vol_ma10_ratio += final_result["vol_ma10_ratio"]
                final_close_ma3_ratio += final_result["close/ma3 ratio"]
                final_close_ma5_ratio += final_result["close/ma5 ratio"]
                final_close_ma10_ratio += final_result["close/ma10 ratio"]
                final_close_ma20_ratio += final_result["close/ma20 ratio"]
                final_close_ma60_ratio += final_result["close/ma60 ratio"]
                final_close_ma120_ratio += final_result["close/ma120 ratio"]
                final_ma_converge_ratio += final_result["ma_cvg_ratio"]
                final_result_profit += final_result["profit"]
                final_result_profit_rate += final_result["profit_rate"]

            print(i + 1, "/", len(codename_list), " 완료")

        final_result2 = {"date": final_result_date, "codename": final_codename, "D+0_open": final_result_d0_open,
                         "D+0_close": final_result_d0_close, "D+0_change_ratio": final_result_d0_change_ratio,
                         "D+1_close": final_result_d1_close, "Vol_ratio": final_result_vol_ma10_ratio,
                         "gap_size_ratio" : final_result_gap_size_ratio,
                         "close/ma3 ratio": final_close_ma3_ratio, "close/ma5 ratio": final_close_ma5_ratio,
                         "close/ma10 ratio": final_close_ma10_ratio, "close/ma20 ratio": final_close_ma20_ratio,
                         "close/ma60 ratio": final_close_ma60_ratio, "close/ma120 ratio": final_close_ma120_ratio,
                         "ma_cvg_ratio": final_ma_converge_ratio,
                         "profit": final_result_profit, "profit_rate": final_result_profit_rate}

        print(len(final_result_date))

        final_result2_df = DataFrame(final_result2)

        sorted_result_df = final_result2_df.sort_values(by="date", ascending=False)

        #sqlite3로 저장
        #con = sqlite3.connect("c:/users/백/kosdaq_result"+ low_X_str + "~" + high_X_str + ".db")
        #final_result2_df.to_sql("결과", con, if_exists="replace")

        #엑셀로 결과 저장
        self.dsstock_none_gui.simul_data_to_excel(sorted_result_df)
        print("엑셀저장 완료")


    def veriyfy_gap_rising_button(self):  #갭상승 종목 시뮬레이션

        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권 기준 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        low_X = self.doubleSpinBox_11.value()
        high_X = self.doubleSpinBox_12.value()

        final_result_date = []
        final_codename = []
        final_result_d0_open = []
        final_result_d0_close = []
        final_result_d0_change_ratio = []
        final_result_d1_close = []
        final_result_gap_size_ratio = []
        final_result_vol_ma10_ratio = []
        final_result_profit = []
        final_result_profit_rate = []

        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기

            print(codename)

            final_result = self.dsstock_none_gui.verify_gap_size(codename,low_X, high_X)

            final_result_date += final_result["date"]
            final_codename += final_result["codename"]
            final_result_d0_open += final_result["D+0_open"]
            final_result_d0_close += final_result["D+0_close"]
            final_result_d0_change_ratio += final_result["D+0_change_ratio"]
            final_result_d1_close += final_result["D+1_close"]
            final_result_gap_size_ratio += final_result["gap_size_ratio"]
            final_result_vol_ma10_ratio += final_result["vol_ma10_ratio"]
            final_result_profit += final_result["profit"]
            final_result_profit_rate += final_result["profit_rate"]

            print(i + 1, "/", len(codename_list), " 완료")

        final_result2 = {"date": final_result_date, "codename": final_codename, "D+0_open" : final_result_d0_open,
                         "D+0_close": final_result_d0_close,
                         "D+0_change_ratio" : final_result_d0_change_ratio,
                        "D+1_close": final_result_d1_close, "gap_size_ratio" : final_result_gap_size_ratio,
                         "vol_ma10_ratio" : final_result_vol_ma10_ratio,
                         "profit": final_result_profit, "profit_rate": final_result_profit_rate}

        print(final_result2)

        final_result2_df = DataFrame(final_result2)
        print("OK3")
        sorted_result_df = final_result2_df.sort_values(by="date", ascending=False)

        print(sorted_result_df)
        print("File is saved as kosdaq_result_gap_rising")

        #sqlite3로 저장
        #con = sqlite3.connect("c:/users/백/kosdaq_result"+ low_X_str + "~" + high_X_str + ".db")
        #final_result2_df.to_sql("결과", con, if_exists="replace")

        #엑셀로 저장
        self.dsstock_none_gui.simul_data_to_excel(sorted_result_df)

    def move_to_excel(self):
        wb = Workbook()
        ws = wb.create_sheet()

        filename = self.lineEdit_3.text()
        tablename = self.lineEdit_4.text()

        con = sqlite3.connect("c:/users/백/" + filename)
        data = pd.read_sql("SELECT * FROM " + tablename, con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기

        # Dataframe을 엑셀로 뿌린다.
        for row in dataframe_to_rows(data, index=True, header=True):
            if len(row) > 1:
                ws.append(row)
        wb.save("c:/users/백/DB_to_Excel.xlsx")
        con.close()
        print("엑셀전달 완료")

    def trace_d1_change_button(self):
        con = sqlite3.connect("c:/users/백/DS_codedata.db")  # 대신증권 기준 코드와 코드네임 불러오기
        df = pd.read_sql("SELECT * FROM kosdaq", con, index_col="index")  # 대신증권 기준 코드와 코드네임 불러오기
        codename_list = df["name"]  # 코드네임을 리스트로 저장

        low_X = self.doubleSpinBox_16.value()
        high_X = self.doubleSpinBox_15.value()

        final_result_date = []
        final_codename = []
        final_result_d0_open = []
        final_result_d0_close = []
        final_result_d0_change_ratio = []
        final_result_d1_close = []
        final_result_vol_ma10_ratio = []
        final_result_gap_size_ratio = []
        final_close_ma3_ratio = []
        final_close_ma5_ratio = []
        final_close_ma10_ratio = []
        final_close_ma20_ratio = []
        final_close_ma60_ratio = []
        final_close_ma120_ratio = []
        final_result_profit = []
        final_result_profit_rate = []

        for i, codename in enumerate(codename_list):  # 뺑뺑이 돌리기

            print(codename)

            final_result = self.dsstock_none_gui.trace_d1_change(codename, low_X, high_X)


            final_result_date += final_result["date"]
            final_codename += final_result["codename"]
            final_result_d0_open += final_result["D+0_open"]
            final_result_d0_close += final_result["D+0_close"]
            final_result_d0_change_ratio += final_result["D+0_change_ratio"]
            final_result_d1_close += final_result["D+1_close"]
            final_result_gap_size_ratio += final_result["gap_size_ratio"]
            final_result_vol_ma10_ratio += final_result["vol_ma10_ratio"]
            final_close_ma3_ratio += final_result["close/ma3 ratio"]
            final_close_ma5_ratio += final_result["close/ma5 ratio"]
            final_close_ma10_ratio += final_result["close/ma10 ratio"]
            final_close_ma20_ratio += final_result["close/ma20 ratio"]
            final_close_ma60_ratio += final_result["close/ma60 ratio"]
            final_close_ma120_ratio += final_result["close/ma120 ratio"]
            final_result_profit += final_result["profit"]
            final_result_profit_rate += final_result["profit_rate"]

            print(i + 1, "/", len(codename_list), " 완료")

        final_result2 = {"date": final_result_date, "codename": final_codename, "D+0_open": final_result_d0_open,
                         "D+0_close": final_result_d0_close,
                         "D+0_change_ratio": final_result_d0_change_ratio,
                         "D+1_close": final_result_d1_close, "vol_ma10_ratio": final_result_vol_ma10_ratio,
                         "gap_size_ratio": final_result_gap_size_ratio,
                         "close/ma3 ratio" : final_close_ma3_ratio, "close/ma5 ratio" : final_close_ma5_ratio,
                         "close/ma10 ratio": final_close_ma10_ratio, "close/ma20 ratio": final_close_ma20_ratio,
                         "close/ma60 ratio": final_close_ma60_ratio, "close/ma120 ratio": final_close_ma120_ratio,
                         "profit": final_result_profit, "profit_rate": final_result_profit_rate}

        print(final_result2)

        final_result2_df = DataFrame(final_result2)
        sorted_result_df = final_result2_df.sort_values(by="date", ascending=False)

        print(sorted_result_df)
        print("File is saved as kosdaq_result_gap_rising")

        # sqlite3로 저장
        # con = sqlite3.connect("c:/users/백/kosdaq_result"+ low_X_str + "~" + high_X_str + ".db")
        # final_result2_df.to_sql("결과", con, if_exists="replace")

        # 엑셀로 저장
        self.dsstock_none_gui.simul_data_to_excel(sorted_result_df)
        print("최종완료")

    def one_stock_ohlc(self):
        instCpCodeMgr = win32com.client.Dispatch("CpUtil.CpCodeMgr")
        start_date = self.spinBox.value()
        end_date = self.spinBox_2.value()
        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")

        bong__ = self.comboBox_2.currentText()

        if bong__ == "일봉":
            bong = "D"  # 일봉
        elif bong__ == "월봉":
            bong = "M"
        else:
            bong = "m"

        code = self.lineEdit_5.text()
        name = instCpCodeMgr.CodeToName(code)
        instStockChart.SetInputValue(0, code)
        if bong == "D":  # 일봉요청이면
            instStockChart.SetInputValue(1, ord("1"))  # 날짜로 요청
            instStockChart.SetInputValue(2, end_date)  # 요청기간
            instStockChart.SetInputValue(3, start_date)  # 요청기간
        else:
            instStockChart.SetInputValue(1, ord("2"))  # 개수로 요청
            instStockChart.SetInputValue(4, 3360)  # 요청개수 144개월 = 12년 ,30분 1680 = 6개월

        if bong == "m":
            instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  # 1번은 시간
            boon = int(self.comboBox_3.currentText())
            instStockChart.SetInputValue(7, boon)  # 틱/분봉주기

        else:
            instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))  # 1번 시간 제외

        instStockChart.SetInputValue(6, ord(bong))  # 봉 구분
        instStockChart.SetInputValue(9, ord("1"))  # 수정 주가

        # BlockRequest
        instStockChart.BlockRequest()

        # GetHeaderValue
        numdata = instStockChart.GetHeaderValue(3)
        numfield = instStockChart.GetHeaderValue(1)

        # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
        name2_list = []
        date_time_list = []
        name_date = []
        date_list = []
        time_list = []
        open_list = []
        high_list = []
        low_list = []
        close_list = []
        change_list = []
        volume_list = []
        change_ratio_list = []
        print(numdata)
        if bong == "m":

            for j in range(numdata):
                name2_list.append(name)

                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)
                time_ = instStockChart.GetDataValue(1, j)
                time_list.append(time_)

                date_time_list.append(str(date) +";"+str(time_))

                open = instStockChart.GetDataValue(2, j)
                open_list.append(open)
                high = instStockChart.GetDataValue(3, j)
                high_list.append(high)
                low = instStockChart.GetDataValue(4, j)
                low_list.append(low)
                close = instStockChart.GetDataValue(5, j)
                close_list.append(close)
                change = instStockChart.GetDataValue(6, j)
                change_list.append(change)
                volume = instStockChart.GetDataValue(7, j)
                volume_list.append(volume)
                change_ratio = round(change / (close - change), 5)
                change_ratio_list.append(change_ratio)
            ds_ohlcv = {"date_time": date_time_list, "date": date_list, "name": name2_list, "time": time_list, "open": open_list,
                        "high": high_list, "low": low_list,
                        "close": close_list, "change": change_list, "volume": volume_list,
                        "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv,
                              columns=["date", "time", "name", "open", "high", "low", "close", "change", "volume",
                                       "change_ratio"],
                              index=ds_ohlcv["date_time"])

        else:
            for j in range(numdata):
                name2_list.append(name)
                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)
                name_date.append(name+";"+str(date))

                open = instStockChart.GetDataValue(1, j)
                open_list.append(open)
                high = instStockChart.GetDataValue(2, j)
                high_list.append(high)
                low = instStockChart.GetDataValue(3, j)
                low_list.append(low)
                close = instStockChart.GetDataValue(4, j)
                close_list.append(close)
                change = instStockChart.GetDataValue(5, j)
                change_list.append(change)
                volume = instStockChart.GetDataValue(6, j)
                volume_list.append(volume)

                if close - change == 0:
                    change_ratio_list.append(1)
                else:
                    change_ratio = round(change / (close - change), 5)
                    change_ratio_list.append(change_ratio)

            ds_ohlcv = {"name_date": name_date, "name": name2_list, "date": date_list, "open": open_list, "high": high_list,
                        "low": low_list, "close": close_list,
                        "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv, columns=["date", "name", "open", "high", "low", "close", "change", "volume",
                                                 "change_ratio"],
                              index=ds_ohlcv["name_date"])
        print("여기임")
        con = sqlite3.connect("c:/users/백/" + code + ".db")
        df.to_sql(name, con, if_exists="replace")

        print("%s 다운로드 완료" % name)
        con.close()
        # self.statusbar.showMessage(lastmsg_here_def)


    def test___(self):
        start_date = "201801"
        end_date = "201902"

        instStockChart = win32com.client.Dispatch("CpSysDib.StockChart")

        con = sqlite3.connect("c:/users/백/DS_codedata.db")
        kospi_kosdaq = self.comboBox.currentText()
        codedata = pd.read_sql("SELECT * FROM " + kospi_kosdaq, con, index_col="index")

        code_list = codedata["code"]
        name_list = codedata["name"]
        bong1 = "D"  # 일봉

        bong2 = "M"  # 월봉
        bong3 = "m"  # 분봉

        for i, code in enumerate(code_list):
            name = name_list[i]

            str_1 = str(i + 1)
            str_2 = str(len(code_list))

            instStockChart.SetInputValue(0, code)
            #instStockChart.SetInputValue(1, ord("1"))  #기간으로 요청
            instStockChart.SetInputValue(1, ord("2")) #개수로 요청
            #instStockChart.SetInputValue(2, end_date)
            #instStockChart.SetInputValue(3, start_date)
            instStockChart.SetInputValue(4, 1950)   #요청개수
            #instStockChart.SetInputValue(5, (0, 2, 3, 4, 5, 6, 8, 37))
            instStockChart.SetInputValue(5, (0, 1, 2, 3, 4, 5, 6, 8, 37))  #분봉전용
            instStockChart.SetInputValue(6, ord(bong3))
            instStockChart.SetInputValue(7, 10)  # 분틱차트 주기
            instStockChart.SetInputValue(9, ord("1"))

            # BlockRequest
            instStockChart.BlockRequest()

            time.sleep(0.25)  # 1종목 요청 후 X초 쉬고난뒤 다음 종목 요청

            # GetHeaderValue
            numdata = instStockChart.GetHeaderValue(3)
            print(numdata)
            numfield = instStockChart.GetHeaderValue(1)

            # GetDataValue : 0번열 : 날짜, 1: open, 2:high, 3:low, 4: close, 5:change, 6:vol
            date_list = []
            open_list = []
            high_list = []
            low_list = []
            close_list = []
            change_list = []
            volume_list = []
            change_ratio_list = []

            for j in range(numdata):
                date = instStockChart.GetDataValue(0, j)
                date_list.append(date)

            for j in range(numdata):
                open = instStockChart.GetDataValue(1, j)
                open_list.append(open)

            for j in range(numdata):
                high = instStockChart.GetDataValue(2, j)
                high_list.append(high)

            for j in range(numdata):
                low = instStockChart.GetDataValue(3, j)
                low_list.append(low)

            for j in range(numdata):
                close = instStockChart.GetDataValue(4, j)
                close_list.append(close)

            for j in range(numdata):
                change = instStockChart.GetDataValue(5, j)
                change_list.append(change)

            for j in range(numdata):
                volume = instStockChart.GetDataValue(6, j)
                volume_list.append(volume)

            for j in range(numdata):
                if j == numdata - 1:
                    change_ratio_list.append(0)

                    break
                change = change_list[j]
                close = close_list[j + 1]
                change_ratio = round(change / close, 5)

                change_ratio_list.append(change_ratio)

            ds_ohlcv = {"date": date_list, "open": open_list, "high": high_list, "low": low_list, "close": close_list,
                        "change": change_list, "volume": volume_list, "change_ratio": change_ratio_list}

            df = pd.DataFrame(ds_ohlcv, columns=["open", "high", "low", "close", "change", "volume", "change_ratio"],
                              index=ds_ohlcv["date"])

            today = datetime.datetime.today().strftime("%Y%m%d")
            # todaystr = str(today)

            con = sqlite3.connect("c:/users/백/stock_" + kospi_kosdaq + "_vol_ma.db")

            df.to_sql(name, con, if_exists="replace")

            lastmsg_here_def = str_1 + "/" + str_2 + name + " 완료"
            print(lastmsg_here_def)
            # self.statusbar.showMessage(lastmsg_here_def)

    def quit_program(self):
        sys.exit()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ds_stock = DS_stock()
    ds_stock.show()
    app.exec_()